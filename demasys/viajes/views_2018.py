# coding: utf-8
from django.shortcuts import render_to_response, get_object_or_404, render, redirect
from django.template.context import RequestContext
from django.db.models import Q, Max
from django.contrib.auth.decorators import permission_required, login_required
from django.contrib import messages
from django.template import loader
from django.conf import settings

from django.http.response import HttpResponseRedirect, HttpResponse
from django.core.urlresolvers import reverse
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from easy_pdf.views import PDFTemplateView
from django.core.mail import send_mail, EmailMessage
from PyPDF2 import PdfFileMerger
from solicitudes.models import Solicitudes

from sepomex.models import Sepomex
from viajes.models import Viaje, Destino, Ruta, Gasto, MovUnidad, CONCEPTO_GASTO,\
    Comprobante, ViajeStatus, ExcelViaje, CargaPdf, ReporteUnidad, Archivos, Facturas, Pagos, SolicitudesViaje, ArchivoPrefacturas, Prefacturas
from viajes.forms import NuevoViajeForm, RutaForm, AutorizarViajeForm, repUnidad, editarMovForm,\
    CambioForm, GastoForm, EntregadoForm, ComprobanteGastoForm,SearchFormFact,repIncidenteForm, UploadFilesForm,\
    FacturadoForm, SearchForm, ComprobanteForm, CargaViajeForm, movimientoForm, libviajeForm,movunidadForm, revisarMovForm, archivosForm,\
    GastoHistoricoForm, SalidaForm, ReportForm, CAMPOS, UploadExcelForm, Movimiento, MovUnidad, DocumentoForma, MensajeCorreoForm, reporteForm,\
    UploadForm, SolicitudesViajeForm, prefacturasFileForm, prefacturasForm
from workflow.models import Workflow, WorkflowHistory, WorkflowActivity, Participant, Role,\
    Transition
from django.forms.models import modelformset_factory
from django.contrib.auth.models import User
from cuentas.models import Perfil
from clientes.models import Cliente, Departamento
from catalogos.models import Economico, Operador, Imagen, movimientosUnidad
from util import to_word
from empleados.models import empleado
from django.contrib.auth.models import Group
##librerias para la descarga de xls 
from django.core.servers.basehttp import FileWrapper
import mimetypes
from django.utils.encoding import smart_str
## Finalizan nuevas librerias
#from sorl.thumbnail import get_thumbnail, delete

import smtplib
import time
import datetime
from email.mime.text import MIMEText


import json
import xlwt
from xlwt import Workbook,XFStyle,Borders, Pattern, Font, easyxf
from django.views.decorators.csrf import csrf_exempt
from django.http import QueryDict

import logging
from clientes.forms import ClienteForm
from decimal import Decimal
import time
import os

import openpyxl
from contable.forms import MovContaForm
from contable.models import Movimientos, Area, Partidas
from contable.views import saldoPartidaXArea
from django.db.models import Sum, Max
from datetime import datetime, date, timedelta

from django.core.exceptions import ObjectDoesNotExist

logger = logging.getLogger(__name__)

@login_required
def index(request):        
    #Obtener catálogos    
    searchform = SearchForm(request.GET)
    
    filters = {}    
    if ('pk' in request.GET) and request.GET['pk'].strip():        
        filters['workflowactivity__pk'] = request.GET['pk']
    if ('status' in request.GET) and request.GET['status'].strip():
        filters['state__pk'] = request.GET['status']
    if ('fecha_from' in request.GET) and request.GET['fecha_from'].strip():
        filters['workflowactivity__fecha_salida__gte'] = request.GET['fecha_from']
    if ('fecha_to' in request.GET) and request.GET['fecha_to'].strip():
        filters['workflowactivity__fecha_salida__lte'] = request.GET['fecha_to']
    if ('cliente' in request.GET) and request.GET['cliente'].strip():
        filters['workflowactivity__cliente__pk'] = request.GET['cliente']
    if ('departamento' in request.GET) and request.GET['departamento'].strip():
        filters['workflowactivity__departamento__pk'] = request.GET['departamento']        
    if ('referencia' in request.GET) and request.GET['referencia'].strip():
        filters['workflowactivity__referencia__icontains'] = request.GET['referencia']
    if ('economico' in request.GET) and request.GET['economico'].strip():
        filters['workflowactivity__economico__pk'] = request.GET['economico']
    if ('operador' in request.GET) and request.GET['operador'].strip():
        filters['workflowactivity__operador__pk'] = request.GET['operador']
    if ('destino' in request.GET) and request.GET['destino'].strip():
        destinos = Destino.objects.filter(destino_clave_municipio=request.GET['destino']).values_list('viaje_id', flat=True)        
        filters['workflowactivity__id__in'] = destinos        
                
        
    viajes_list = ViajeStatus.objects.filter(**filters)  

    if not bool(filters):
        viajes_list = viajes_list.order_by('-workflowactivity__pk')          
         
    paginator = Paginator(viajes_list, 250)
    page = request.GET.get('page')
    
    try:
        viajes = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        viajes = paginator.page(1)
    except EmptyPage:
        # If page is out of range (e.g. 9999), deliver last page of results.
        viajes = paginator.page(paginator.num_pages)
    
        
    context = RequestContext(request,{             
        'searchform': searchform,  
        'viajes': viajes,
    })    
    
    return render_to_response('viajes.html', context)

@login_required
def historial(request, viaje_id):
    viaje = get_object_or_404(Viaje, pk=viaje_id)  
    eventos = viaje.flujo.history.filter(log_type=1)  
    
    context = RequestContext(request,{             
        'viaje': viaje,
        'eventos': eventos,
    })     
    return render_to_response('historial.html', context)

@login_required
def actualizar(request, viaje_id):
    viaje = get_object_or_404(Viaje, pk=viaje_id) 
    
    user = request.user
    profile = user.get_profile()
    participant = user.participant_set.first()
    
    current = viaje.flujo.current_state()
    transiciones = current.state.transitions_from.filter(roles__in=participant.roles.all())
    #Si el usuario no tiene opciones en este status, manda mensaje
    
    template = ''
    view_context = {             
        'viaje': viaje,    
        'transiciones': transiciones,  
        'current': current,  
    }
    
    #Solicitado
    if current.state.pk == 1:
        template = 'flujos/solicitado.html'
        destinos = viaje.destino_set.values_list('destino_clave_municipio', flat=True)
        destinos_str = ','.join(destinos)        
        form = AutorizarViajeForm(initial={'destinos':destinos_str},instance=viaje)
        view_context['form'] = form
    #Autorizado
    elif current.state.pk == 2:
        form = CambioForm(instance=viaje)
        view_context['form'] = form
        template = 'flujos/autorizado.html'
    #Ruta de Salida
    elif current.state.pk == 3:
        #TODO: Se pueden agregar gastos de salida, depende del grupo se llena el campo
        gastos_salida=[{'tipo': 1,'viaje': viaje_id,},{'tipo': 1,'viaje': viaje_id,},]
        gastos_regreso=[{'tipo': 2,'viaje': viaje_id,},{'tipo': 2,'viaje': viaje_id,},]
        
        conceptos = json.dumps(CONCEPTO_GASTO)
        GastosFormSet = modelformset_factory(Gasto, form=GastoForm, max_num=10, extra=2, can_delete=True)
        #formset_salida = GastosFormSet(initial=gastos_salida, queryset=viaje.gasto_set.filter(tipo=1), prefix='salida')
        #formset_regreso = GastosFormSet(initial=gastos_regreso, queryset=viaje.gasto_set.filter(tipo=2), prefix='regreso')
        form = SalidaForm(instance=viaje)
        view_context['form'] = form
        view_context['conceptos'] = conceptos
        ##view_context['formset_salida'] = formset_salida
        #view_context['formset_regreso'] = formset_regreso
        template = 'flujos/salida.html'
    #Entregado
    elif current.state.pk == 4:
        #TODO: Se puede actualziar la fecha de entrega
        gastos_salida=[{'tipo': 1,'viaje': viaje_id,},{'tipo': 1,'viaje': viaje_id,},]
        gastos_regreso=[{'tipo': 2,'viaje': viaje_id,},{'tipo': 2,'viaje': viaje_id,},]
        
        conceptos = json.dumps(CONCEPTO_GASTO)
        GastosFormSet = modelformset_factory(Gasto, form=GastoForm, max_num=10, extra=2, can_delete=True)
        EntregasFormSet = modelformset_factory(Destino, form=EntregadoForm, max_num=10, extra=0) ### Formulario para cambiar la fecha de entrega.
        
        formset_salida = GastosFormSet(initial=gastos_salida, queryset=viaje.gasto_set.filter(tipo=1), prefix='salida')
        formset_regreso = GastosFormSet(initial=gastos_regreso, queryset=viaje.gasto_set.filter(tipo=2), prefix='regreso')
        formset_entregado = EntregasFormSet(queryset=viaje.destino_set.all(), prefix='destino')
        form = CambioForm(instance=viaje) ## Formulario para actualizar el Bill o referencia.
        view_context['form'] = form
        view_context['conceptos'] = conceptos
        view_context['formset_salida'] = formset_salida
        view_context['formset_regreso'] = formset_regreso
        view_context['formset_entregado'] = formset_entregado
        template = 'flujos/entregado.html'  
    #Ruta de llegada
    elif current.state.pk == 5:
        gastos_salida=[{'tipo': 1,'viaje': viaje_id,},{'tipo': 1,'viaje': viaje_id,},]
        gastos_regreso=[{'tipo': 2,'viaje': viaje_id,},{'tipo': 2,'viaje': viaje_id,},]
        
        conceptos = json.dumps(CONCEPTO_GASTO)
        GastosFormSet = modelformset_factory(Gasto, form=GastoForm, max_num=10, extra=2, can_delete=True)
        formset_salida = GastosFormSet(initial=gastos_salida, queryset=viaje.gasto_set.filter(tipo=1), prefix='salida')
        formset_regreso = GastosFormSet(initial=gastos_regreso, queryset=viaje.gasto_set.filter(tipo=2), prefix='regreso')
        form = CambioForm()
        view_context['form'] = form
        view_context['conceptos'] = conceptos
        view_context['formset_salida'] = formset_salida
        view_context['formset_regreso'] = formset_regreso
        template = 'flujos/llegada.html'
    #Liberación de documentos
    elif current.state.pk == 6:
        #TODO: Se pueden cargar comprobante
        ComprobanteFormSet = modelformset_factory(Destino, form=ComprobanteForm, max_num=10, extra=0)
        formset_comprobante = ComprobanteFormSet(queryset=viaje.destino_set.all(), prefix='destino')                
        form = CambioForm(instance=viaje)
        view_context['formset_comprobante'] = formset_comprobante
        view_context['form'] = form
        template = 'flujos/liberacion.html' 
    #Facturacion
    elif current.state.pk == 7:
        #TODO: Se puede asignar el no de factura y verificar los importes
        form = FacturadoForm(instance=viaje)
        view_context['form'] = form
        template = 'flujos/facturacion.html'
    #Liquidado
    elif current.state.pk == 8:
        #TODO: Consulta
        is_finanzas = request.user.groups.filter(name='Finanzas').exists()
        view_context['is_finanzas'] = is_finanzas
        template = 'flujos/liquidado.html'
    #Cancelado
    elif current.state.pk == 9:
        Movimientos.objects.filter(ref_viaje=viaje_id).delete()
        eco = Economico.objects.filter(id = viaje.economico_id)
        Viaje.objects.filter(pk=viaje_id).update(factura = 'Cancelado')
        
        for e in eco:
            estado = e.status
            viajes = e.enviaje
            print 'Viajes antes de restar: ' + str(viajes)
        viajes = viajes - 1
        print 'Viajes despues de restar: ' + str(viajes)
        if viajes <= 0:
            estado = 10 ##Coloca la unidad en disponible.
            MovUnidad.objects.filter(id = viaje.economico_id).update(tipo = estado, obs = '', fechai = None, fechaf = None, operador = '', usuario = '', ts = None, destino = '', destino2 = '', departamento = '', cliente = '', fechamcia = None)
        elif viajes == 1:
            estado = 9
        Economico.objects.filter(id = viaje.economico_id).update(status =estado, enviaje = viajes) 
        MovUnidad.objects.filter(unidad_id= viaje.economico_id).update(enviaje=viajes)  

        ope = empleado.objects.filter (id = viaje.operador_id)
        for o in ope:
            envia = o.enviaje
            via = o.servicios
        via = via - 1
        envia = envia -1 
        empleado.objects.filter(id = viaje.operador_id).update(enviaje = via, servicios = envia)
        print '<---Llega hasta aca....'
        template = 'flujos/cancelado.html'
        
    elif current.state.pk  == 10:
        #TODO: Se pueden agregar gastos de salida, depende del grupo se llena el campo
        gastos_salida=[{'tipo': 1,'viaje': viaje_id,},{'tipo': 1,'viaje': viaje_id,},]
        gastos_regreso=[{'tipo': 2,'viaje': viaje_id,},{'tipo': 2,'viaje': viaje_id,},]
        conceptos = json.dumps(CONCEPTO_GASTO)
        GastosFormSet = modelformset_factory(Gasto, form=GastoForm, max_num=10, extra=2, can_delete=True)
        #formset_salida = GastosFormSet(initial=gastos_salida, queryset=viaje.gasto_set.filter(tipo=1), prefix='salida')
        #formset_regreso = GastosFormSet(initial=gastos_regreso, queryset=viaje.gasto_set.filter(tipo=2), prefix='regreso')
        form = SalidaForm(instance=viaje)
        view_context['form'] = form
        view_context['conceptos'] = conceptos
        ##view_context['formset_salida'] = formset_salida
        #view_context['formset_regreso'] = formset_regreso
        template = 'flujos/carga_gastos.html'    
    
    context = RequestContext(request, view_context)
                            
    return render_to_response(template, context)
    
@login_required    
def solicitado(request, viaje_id): 
    #TODO: Validar que tienen acceso a esta vista       
    user = request.user
    profile = user.get_profile()
    participant = user.participant_set.first()
    
    viaje = get_object_or_404(Viaje, pk=viaje_id)
    destinos = viaje.destino_set.values_list('destino_clave_municipio', flat=True)
    destinos_str = ','.join(destinos)    
        
    current = viaje.flujo.current_state()
    transiciones = current.state.transitions_from.filter(roles__in=participant.roles.all())
            
    if request.method == 'POST':                                 
        form = AutorizarViajeForm(request.POST, instance=viaje)
        if form.is_valid():
            viaje = form.save()
            viaje.destino_set.all().delete()
            
            status = form.cleaned_data['status']
            nota = form.cleaned_data['nota']
            destinos = form.cleaned_data['destinos'].split(',')
            #Salvar destinos
            for destino in destinos:
                Destino.objects.create(viaje=viaje, destino_clave_municipio=destino)
                
            #Agregar gatos de casetas
            if viaje.economico.pasa_como == 1:
                #PASA COMO AUTO
                casetas_salida = viaje.ruta.salida_total_auto() + viaje.ruta.regreso_total_auto()                
            else:
                #PASA COMO CAMIONETA
                casetas_salida = viaje.ruta.salida_total_camioneta() + viaje.ruta.regreso_total_camioneta()                
            
            if casetas_salida > 0:
                Gasto.objects.create(tipo=1, concepto='Casetas', viaje=viaje, calculado=casetas_salida, pagado_contabilidad=casetas_salida)
                
            #Cambia status            
            transition = current.state.transitions_from.get(pk=status)
            viaje.flujo.progress(transition, user, nota)
            messages.success(request, 'Se ha cambiado el status del viaje: %s'%viaje)
            return HttpResponseRedirect(reverse('viajes.views.actualizar', args=[viaje.pk]))
        else:
            messages.error(request, 'Ha ocurrido un error. Por favor verifique los datos.')
    else:
        form = AutorizarViajeForm(initial={'destinos':destinos_str}, instance=viaje)
            
    context = RequestContext(request,{ 
        'viaje': viaje,    
        'transiciones': transiciones,
        'form': form,
        'current': current,
    }) 
    
    return render_to_response('flujos/solicitado.html', context)
        
@login_required            
def cambio(request, viaje_id):        
    user = request.user
    profile = user.get_profile()
    participant = user.participant_set.first()
    
    viaje = get_object_or_404(Viaje, pk=viaje_id)
    current = viaje.flujo.current_state()
    transiciones = current.state.transitions_from.filter(roles__in=participant.roles.all())
            
    if request.method == 'POST':                        
        form = CambioForm(request.POST, instance=viaje)
        if form.is_valid():
            
            status = form.cleaned_data['status']
            nota = form.cleaned_data['nota']
                
            #Cambia status            
            transition = current.state.transitions_from.get(pk=status)
            viaje.flujo.progress(transition, user, nota)
            
            messages.success(request, 'Se ha cambiado el status del viaje: %s'%viaje)
            return HttpResponseRedirect(reverse('viajes.views.actualizar', args=[viaje.pk]))
        else:
            messages.error(request, 'Ha ocurrido un error. Por favor verifique los datos.')
    else:
        form = CambioForm(instance=viaje)
            
    context = RequestContext(request,{ 
        'viaje': viaje,    
        'transiciones': transiciones,
        'form': form,
    }) 
    
    return render_to_response('flujos/autorizado.html', context)           

@login_required
def salida(request, viaje_id):
    
    user = request.user
    profile = user.get_profile()
    participant = user.participant_set.first()
    
    viaje = get_object_or_404(Viaje, pk=viaje_id)
    current = viaje.flujo.current_state()
    transiciones = current.state.transitions_from.filter(roles__in=participant.roles.all())
                
    conceptos = json.dumps(CONCEPTO_GASTO)
    
    GastosFormSet = modelformset_factory(Gasto, form=GastoForm, max_num=10, extra=2, can_delete=True)
    
    

    if request.method == 'POST':
        formset_salida = GastosFormSet(request.POST, queryset=viaje.gasto_set.filter(tipo=1), prefix='salida')
        formset_regreso = GastosFormSet(request.POST, queryset=viaje.gasto_set.filter(tipo=2), prefix='regreso')
        form = SalidaForm(request.POST, instance=viaje)
        
        if form.is_valid():## and formset_salida.is_valid() and formset_regreso.is_valid():
            #formset_salida.save()
            #formset_regreso.save()
            form.save()
            
            status = form.cleaned_data['status']
            nota = form.cleaned_data['nota']
            
            if status == 0:
                messages.success(request, 'Los gasto han sido guardados')
            else:
                #Cambia status            
                transition = current.state.transitions_from.get(pk=status)
                viaje.flujo.progress(transition, user, nota)
                #print 'Aca ya esta cancelado, aqui se debe de liberar al Operador y a la unidad....'
                #viajes = Viaje.objects.filter(id = viaje_id)
                #for v in viajes:
                #    operador = v.operador_id
                #    economico = v.economico_id
                #    print 'Este es el operador' + str(operador)
                #    print 'Este es el economico' + str(economico)
                #    empleado.objects.filter(id = str(operador)).update(servicios =-1)
                #    Economico.objects.filter(id = str(economico)).update(enviaje =-1)

                
                #messages.success(request, 'Se ha cambiado el status del viaje: %s'%viaje)
		
		messages.success(request, 'Datos guardados correctamente')
            
            return HttpResponseRedirect(reverse('viajes.views.actualizar', args=[viaje.pk]))
        else:            
            messages.error(request, 'Ha ocurrido un error. Por favor verifique los datos.')
            
    else:
        gastos_salida=[{'tipo': 1,'viaje': viaje_id,},{'tipo': 1,'viaje': viaje_id,},]
        gastos_regreso=[{'tipo': 2,'viaje': viaje_id, },{'tipo': 2,'viaje': viaje_id,},]
        formset_salida = GastosFormSet(initial=gastos_salida, queryset=viaje.gasto_set.filter(tipo=1), prefix='salida')
        formset_regreso = GastosFormSet(initial=gastos_regreso, queryset=viaje.gasto_set.filter(tipo=2), prefix='regreso')
        form = SalidaForm(instance=viaje)
        
    context = RequestContext(request,{ 
        'viaje': viaje,    
        'transiciones': transiciones,
        'form': form,
        'conceptos': conceptos,
        'formset_salida': formset_salida,
        'formset_regreso': formset_regreso,
    }) 
    
    return render_to_response('flujos/salida.html', context) 

@login_required
def llegada(request, viaje_id):    
    user = request.user
    profile = user.get_profile()
    participant = user.participant_set.first()
    
    viaje = get_object_or_404(Viaje, pk=viaje_id)
    current = viaje.flujo.current_state()
    transiciones = current.state.transitions_from.filter(roles__in=participant.roles.all())
                
    conceptos = json.dumps(CONCEPTO_GASTO)
    
    GastosFormSet = modelformset_factory(Gasto, form=GastoForm, max_num=10, extra=2, can_delete=True)
    
    if request.method == 'POST':
        formset_salida = GastosFormSet(request.POST, queryset=viaje.gasto_set.filter(tipo=1), prefix='salida')
        formset_regreso = GastosFormSet(request.POST, queryset=viaje.gasto_set.filter(tipo=2), prefix='regreso')
        form = CambioForm(request.POST)
        
        if form.is_valid(): #and formset_salida.is_valid() c:
            #formset_salida.save()
            #formset_regreso.save()
            
            status = form.cleaned_data['status']
            nota = form.cleaned_data['nota']
            
            if status == 0:
                messages.success(request, 'Los gasto han sido guardados')
            else:
                #Cambia status            
                transition = current.state.transitions_from.get(pk=status)
                viaje.flujo.progress(transition, user, nota)
                
                messages.success(request, 'Se ha cambiado el status del viaje: %s'%viaje)
            
            return HttpResponseRedirect(reverse('viajes.views.actualizar', args=[viaje.pk]))
        else:            
            messages.error(request, 'Ha ocurrido un error. Por favor verifique los datos.')
            
    else:
        gastos_salida=[{'tipo': 1,'viaje': viaje_id,},{'tipo': 1,'viaje': viaje_id,},]
        gastos_regreso=[{'tipo': 2,'viaje': viaje_id, },{'tipo': 2,'viaje': viaje_id,},]
        formset_salida = GastosFormSet(initial=gastos_salida, queryset=viaje.gasto_set.filter(tipo=1), prefix='salida')
        formset_regreso = GastosFormSet(initial=gastos_regreso, queryset=viaje.gasto_set.filter(tipo=2), prefix='regreso')
        form = CambioForm()
        
    context = RequestContext(request,{ 
        'viaje': viaje,    
        'transiciones': transiciones,
        'form': form,
        'conceptos': conceptos,
        'formset_salida': formset_salida,
        'formset_regreso': formset_regreso,
    }) 
    
    return render_to_response('flujos/llegada.html', context) 


@login_required
def entrega(request, viaje_id):
    
    user = request.user
    profile = user.get_profile()
    participant = user.participant_set.first()
    
    viaje = get_object_or_404(Viaje, pk=viaje_id)
    current = viaje.flujo.current_state()
    transiciones = current.state.transitions_from.filter(roles__in=participant.roles.all())
                
    conceptos = json.dumps(CONCEPTO_GASTO)
    
    GastosFormSet = modelformset_factory(Gasto, form=GastoForm, max_num=10, extra=2, can_delete=True)
    EntregasFormSet = modelformset_factory(Destino, form=EntregadoForm, max_num=10, extra=0)
    
    if request.method == 'POST':
        formset_salida = GastosFormSet(request.POST, queryset=viaje.gasto_set.filter(tipo=1), prefix='salida')
        formset_regreso = GastosFormSet(request.POST, queryset=viaje.gasto_set.filter(tipo=2), prefix='regreso')
        formset_entregado = EntregasFormSet(request.POST, request.FILES, queryset=viaje.destino_set.all(), prefix='destino')
        form = CambioForm(request.POST, instance=viaje)
        
        if form.is_valid()and formset_entregado.is_valid(): ##and formset_salida.is_valid() and formset_regreso.is_valid() and formset_entregado.is_valid():
            #formset_salida.save()
            #formset_regreso.save()
            formset_entregado.save()
            form.save()
            
            status = form.cleaned_data['status']
            nota = form.cleaned_data['nota']
            
            if status == 0:
                messages.success(request, 'Los gasto han sido guardados')                
            else:
                #Cambia status            
                transition = current.state.transitions_from.get(pk=status)
                viaje.flujo.progress(transition, user, nota)                
                messages.success(request, 'Se ha cambiado el status del viaje: %s'%viaje)
                
            return HttpResponseRedirect(reverse('viajes.views.actualizar', args=[viaje.pk]))
        else:            
            messages.error(request, 'Ha ocurrido un error. Por favor verifique los datos.')
            
    else:
        gastos_salida=[{'tipo': 1,'viaje': viaje_id,},{'tipo': 1,'viaje': viaje_id,},]
        gastos_regreso=[{'tipo': 2,'viaje': viaje_id, },{'tipo': 2,'viaje': viaje_id,},]
        formset_salida = GastosFormSet(initial=gastos_salida, queryset=viaje.gasto_set.filter(tipo=1), prefix='salida')
        formset_regreso = GastosFormSet(initial=gastos_regreso, queryset=viaje.gasto_set.filter(tipo=2), prefix='regreso')
        formset_entregado = EntregasFormSet(queryset=viaje.destino_set.all(), prefix='destino')
        form = CambioForm(instance=viaje)
        
    context = RequestContext(request,{ 
        'viaje': viaje,    
        'transiciones': transiciones,
        'form': form,
        'conceptos': conceptos,
        'formset_salida': formset_salida,
        'formset_regreso': formset_regreso,
        'formset_entregado': formset_entregado,
    }) 
    
    return render_to_response('flujos/entregado.html', context) 

@login_required
def comprobante(request, viaje_id):
    
    user = request.user
    profile = user.get_profile()
    participant = user.participant_set.first()
    
    viaje = get_object_or_404(Viaje, pk=viaje_id)
    current = viaje.flujo.current_state()
    transiciones = current.state.transitions_from.filter(roles__in=participant.roles.all())
    ComprobanteFormSet = modelformset_factory(Destino, form=ComprobanteForm, max_num=10, extra=0)              
        
    if request.method == 'POST':               
        formset_comprobante = ComprobanteFormSet(request.POST, request.FILES, queryset=viaje.destino_set.all(), prefix='destino') 
        form = CambioForm(request.POST, instance=viaje)
        
        if form.is_valid():            
            status = form.cleaned_data['status']
            nota = form.cleaned_data['nota']
            
            if nota == '' and formset_comprobante.is_valid():                             
                formset_comprobante.save()
                form.save()
            
                if status == 0:
                    messages.success(request, 'Los comprobantes han sido guardados')                
                else:
                    #Cambia status            
                    transition = current.state.transitions_from.get(pk=status)
                    viaje.flujo.progress(transition, user, nota)                
                    messages.success(request, 'Se ha cambiado el status del viaje: %s'%viaje)
            else:
                form.save()
                transition = current.state.transitions_from.get(pk=status)
                viaje.flujo.progress(transition, user, nota)                
                messages.success(request, 'Se ha cambiado el status del viaje: %s'%viaje)                
                
            return HttpResponseRedirect(reverse('viajes.views.actualizar', args=[viaje.pk]))
        else:                   
            messages.error(request, 'Ha ocurrido un error. Por favor verifique los datos.')
            
    else:        
        formset_comprobante = ComprobanteFormSet(queryset=viaje.destino_set.all(), prefix='destino')
        form = CambioForm(instance=viaje)
        
    context = RequestContext(request,{ 
        'viaje': viaje,    
        'transiciones': transiciones,
        'form': form,
        'formset_comprobante': formset_comprobante,
        'current': current,
    }) 
    
    return render_to_response('flujos/liberacion.html', context) 

@login_required
def comprobantegasto(request, gasto_id):
    user = request.user
    profile = user.get_profile()
    
    gasto = get_object_or_404(Gasto, pk=gasto_id)
    viaje = gasto.viaje
    
    context = RequestContext(request,{ 
        'gasto': gasto,
        'viaje': viaje,
    }) 
    return render_to_response('comprobantes.html', context) 

@login_required
@csrf_exempt
def comprobante_upload(request, gasto_id):
    data = []
    if request.method == 'POST':
        post_data = QueryDict('gasto=%s'%gasto_id)
        form = ComprobanteGastoForm(post_data, request.FILES)
        if form.is_valid():
            comprobante = form.save()
            file_data = {
                'url': comprobante.archivo.url,
                'name': comprobante.filename(),
                'type': '',
                'thumbnailUrl': '',
                'size': comprobante.archivo.size,
                'deleteUrl': reverse('viajes.views.comprobante_delete', args=[comprobante.pk]),
                'deleteType': 'DELETE',
            }
            data.append(file_data)
    else:
        gasto = Gasto.objects.get(pk=gasto_id)
        for comprobante in gasto.comprobante_set.all(): 
            #thumb = get_thumbnail(comprobante.archivo, '80x80', crop='center', quality=99)                   
            file_data = {
                'url': comprobante.archivo.url,
                'name': comprobante.filename(),
                'type': '',
                'thumbnailUrl': '',
                'size': comprobante.archivo.size,
                'deleteUrl': reverse('viajes.views.comprobante_delete', args=[comprobante.pk]),
                'deleteType': 'DELETE',
            }
            data.append(file_data)
        
    return HttpResponse(json.dumps({'files':data}), mimetype="application/json")

@login_required
@csrf_exempt
def comprobante_delete(request, comprobante_id):
    comprobante = get_object_or_404(Comprobante, pk=comprobante_id);
    result = {'files':[{comprobante.filename(): True}]}    
    comprobante.delete()
    
    return HttpResponse(json.dumps(result), mimetype="application/json")
    
    
@login_required
def facturado(request, viaje_id):
    user = request.user
    profile = user.get_profile()
    participant = user.participant_set.first()
    
    viaje = get_object_or_404(Viaje, pk=viaje_id)
    current = viaje.flujo.current_state()
    transiciones = current.state.transitions_from.filter(roles__in=participant.roles.all())                            
    
    if request.method == 'POST':        
        form = FacturadoForm(request.POST, instance=viaje)
        
        if form.is_valid():            
            status = form.cleaned_data['status'] 
	    print status      
            nota = form.cleaned_data['nota']                 
            form.save()
            
            if status == 0:
                messages.success(request, 'El número de factura ha sido guardado')                
            else:
                #Cambia status            
                transition = current.state.transitions_from.get(pk=status)
                viaje.flujo.progress(transition, user, nota)                
                messages.success(request, 'Se ha cambiado el status del viaje: %s'%viaje)
                
            return HttpResponseRedirect(reverse('viajes.views.actualizar', args=[viaje.pk]))
        else:            
            messages.error(request, 'Ha ocurrido un error. Por favor verifique los datos.')
            
    else:        
        form = FacturadoForm(instance=viaje)
        
    context = RequestContext(request,{ 
        'viaje': viaje,    
        'transiciones': transiciones,
        'form': form,        
    }) 
    
    return render_to_response('flujos/facturacion.html', context)
            
@login_required
def agregar(request):
    userid = request.user.id  ### obtengo el id del usurio 
    user = request.user    ### obtengo el usuario
    profile = user.get_profile()  ### obtengo el nombre completo del usuario
    u1 = Perfil.objects.get(user_id = userid)  ## Obtengo el user_id en base al id del usurio
  
    
    if request.method == 'POST':
        form = NuevoViajeForm(request.POST)
        econo = request.POST['economico']
        ecosta = Economico.objects.filter(id = econo)
        for eco in ecosta:
            esta = eco.status
            valviajes = eco.enviaje


        if form.is_valid():
            viaje = form.save(commit=False)
            eco = viaje.economico
            destinos = form.cleaned_data['destinos'].split(',')
	    operador = request.POST['operador']
            fecha_viaje = request.POST['fecha_salida']
            hoy = date.today()
            h = datetime.strptime(fecha_viaje, "%d-%m-%Y")
            ho = str(hoy)
            ho = datetime.strptime(ho, "%Y-%m-%d")

            if h >= ho:
                    
               lic = empleado.objects.filter(id = operador)
               for l in lic:
                  vig = l.lic_vigencia
                  c_opalo = l.cer_opalo
                  c_rc = l.cer_rc
                  f_viaje = str(fecha_viaje)        
                  date_viaje = datetime.strptime(f_viaje, "%d-%m-%Y")
                  var = date_viaje.date()
                  sta = int(l.enviaje)
                  serv = l.servicios

               if vig > var and (c_opalo is True or c_rc is True)  and (esta == 10 or esta == 9 or esta == 1) and valviajes <= 1 and serv <= 1:  
            #Crear workflow
            ### colocar nuevo codigo aqui.
                   wf = Workflow.objects.get(pk=1)
                   wa = WorkflowActivity(workflow=wf, created_by=user)
                   wa.save()            
                   viaje.flujo = wa
                   serv = serv + 1 
                   stan=empleado.objects.filter(id= operador).update(servicios = serv)
                   valviajes = valviajes + 1
                   print valviajes
                   estan=Economico.objects.filter(id = econo).update(enviaje = valviajes)
                   MovUnidad.objects.filter(unidad_id=econo).update(enviaje = valviajes)
                   viaje.save()      
	
            #Asignar usuarios 
                   periles = Perfil.objects.filter(
                       Q(cliente__isnull=True) | Q(cliente=viaje.cliente)
               )
            
                   coordinador = Role.objects.get(pk=1)
                   finanzas = Role.objects.get(pk=2)
                   administrador = Role.objects.get(pk=3)
                   cliente = Role.objects.get(pk=4)
            
                   Participant.objects.create(user=user, workflowactivity=wa)
            
                   for perfil in periles:
                       if perfil.user != user:
                           Participant.objects.create(user=perfil.user, workflowactivity=wa)
                
                       if perfil.user.is_superuser:
                           wa.assign_role(user, perfil.user, administrador)
                       grupos = perfil.user.groups.values_list('name',flat=True)
                       if 'Cliente' in grupos:
                           wa.assign_role(user, perfil.user, cliente)
                       if 'Coordinador' in grupos:
                           wa.assign_role(user, perfil.user, coordinador)
                       if 'Finanzas' in grupos:
                           wa.assign_role(user, perfil.user, finanzas)
                    
            #Salvar destinos
                   if destinos[0]:
                      print 'pasa el if de destinos'
                      for destino in destinos:
                         Destino.objects.create(viaje=viaje, destino_clave_municipio=destino)
                         des1= destino.split("-")
                         des11= des1[0]
                         des12= des1[1]
                         municipio = Sepomex.objects.filter(clave_estado = des11, id_asenta_cpcons = des12)#.aggregate(Max('municipio'))
                         #print municipio[0].municipio
                         des1=municipio[0].municipio
                   else:
                      des1='Sin Destino'
                   uni=request.POST['economico']
                   o=request.POST['operador']
                   t=9
                   ob=request.POST['observaciones']
                   c=request.POST['cliente']
                   d=request.POST['departamento']
                   con=request.POST['contiene']
                   obs= con + ',' + ob
                   fecha=request.POST['fecha_salida']
                   fi = datetime.strptime(fecha,"%d-%m-%Y").strftime("%Y-%m-%d") 
                   ts1= time.time()
                   ##st = datetime.fromtimestamp(ts1).strftime('%Y-%m-%d %H:%M:%S')
                   st = hoy = date.today()
                   model=Economico.objects.filter(id=uni)
                   
                   for modelos in model:
                       m=modelos.modelo

                   MovUnidad.objects.filter(unidad=uni).update(
                   usuario = u1.id, operador = o, tipo = t, modelo=m, obs = obs, destino = des1, ts=st, destino2= viaje, cliente= c, departamento= d, fechai= fi, email = 0 )     
            
                   viaje.flujo.start(user)
            
                   grupos_usuario = user.groups.values_list('name',flat=True)
                   if 'Cliente' not in grupos_usuario:
                       time.sleep(1)                
                       autorizado = Transition.objects.get(pk=1)
                       viaje.flujo.progress(autorizado, user)
                       time.sleep(1)                
                       salida = Transition.objects.get(pk=2)
                       viaje.flujo.progress(salida, user)    
                       #Agregar gatos de casetas
                       if viaje.economico.pasa_como == 1:
                           #PASA COMO AUTO
                           casetas_salida = viaje.ruta.salida_total_auto() + viaje.ruta.regreso_total_auto()                
                       else:
                           #PASA COMO CAMIONETA
                           casetas_salida = viaje.ruta.salida_total_camioneta() + viaje.ruta.regreso_total_camioneta()                
                
                       if casetas_salida > 0:
                           Gasto.objects.create(tipo=1, concepto='Casetas', viaje=viaje, calculado=casetas_salida, pagado_contabilidad=casetas_salida, vale_operaciones = 0, vale_contabilidad=0, efectivo_operaciones = 0, efectivo_contabilidad = 0)
                           saldo_partida_actual = saldoPartidaXArea(3, 3)
                           afectaOpera = Movimientos(idPartida=Partidas.objects.get(id = 3), idArea = Area.objects.get(id = 3), id_auth_user=user,
                                              importe=casetas_salida, tipo='C', montoPartida=saldo_partida_actual, ref_viaje=viaje.pk, concepto='Casetas')
                           afectaOpera.save()
                   
                   viaje= get_object_or_404(Viaje, pk=str(viaje))
                   current = viaje.flujo.current_state()
                   viaje2 = viaje.id 
                   print viaje2
                   noFlujo = viaje.flujo_id
                   wfh = WorkflowHistory.objects.filter(workflowactivity_id = noFlujo)
                   wfhf = WorkflowHistory.objects.filter(id = wfh[0].id)
                   for wf in wfhf:
                            wfh_id=wf.id
                            wfh_st=wf.state_id
                            wf.state_id = 1
                            wf.save()
                   #cstate = WorkflowHistory.objects.
                   messages.success(request, 'Se ha agregado un nuevo viaje: %s'%viaje)
                   return HttpResponseRedirect(reverse('viajes.views.index',))               
                
               else:
                   print 'entro al else de la validacion del operador y unidad'
	           messages.error(request, 'El Viaje no se puede Asignar al Operador por que no cumple con los requisitos.Licencia Vigente, Certificacion Opalo, Certificacion Recurso Confiable o Se Encuentra en Viaje Favor de Seleccionar otro operador.')
            else:
               print 'entro al else de la fecha'
               messages.error(request, 'La fecha no puede ser menor a hoy %s' %hoy)   
        else:
           messages.error(request, 'Ha ocurrido un error. Por favor verifique los datos.')
    else:
        form = NuevoViajeForm()        
    
    context = RequestContext(request,{ 
        'form': form,
        'action': 'Agregar',
    })    
    
    return render_to_response('nuevoviaje_form.html', context)

def editar(request, cliente_id):
    cliente = get_object_or_404(Cliente, pk=cliente_id)
    
    if request.method == 'POST':
        form = ClienteForm(request.POST, instance=cliente)
        if form.is_valid():
            form.save()
            messages.success(request, 'Se ha editado el cliente: %s'%cliente)
            return HttpResponseRedirect(reverse('clientes.views.index',))
        else:
            messages.error(request, 'Ha ocurrido un error. Por favor verifique los datos.')
    else:
        form = ClienteForm(instance=cliente)
        
    context = RequestContext(request,{ 
        'form': form,
        'action': 'Editar',
    })    
    
    return render_to_response('clientes_form.html', context)

@login_required
def carga(request):    
    user = request.user
    profile = user.get_profile()
        
    conceptos = json.dumps(CONCEPTO_GASTO)
    
    GastosFormSet = modelformset_factory(Gasto, form=GastoHistoricoForm, max_num=10, extra=2, can_delete=True)
    EntregasFormSet = modelformset_factory(Destino, form=EntregadoForm, max_num=10, extra=1)
    
    if request.method == 'POST':
        formset_salida = GastosFormSet(request.POST, queryset=Gasto.objects.none(),  prefix='salida')
        formset_regreso = GastosFormSet(request.POST, queryset=Gasto.objects.none(), prefix='regreso')
        formset_entregado = EntregasFormSet(request.POST, request.FILES, queryset=Destino.objects.none(), prefix='destino')        
        form = CargaViajeForm(request.POST)
                
        if form.is_valid() and formset_salida.is_valid() and formset_regreso.is_valid() and formset_entregado.is_valid():
            viaje = form.save(commit=False)
            viaje.pk = form.cleaned_data['id']        
                        
            #Crear workflow
            wf = Workflow.objects.get(pk=1)
            wa = WorkflowActivity(workflow=wf, created_by=user)
            wa.save()            
            
            viaje.flujo = wa
            viaje.save()
            
            #Asignar usuarios 
            periles = Perfil.objects.filter(
                Q(cliente__isnull=True) | Q(cliente=viaje.cliente)
            )
            
            coordinador = Role.objects.get(pk=1)
            finanzas = Role.objects.get(pk=2)
            administrador = Role.objects.get(pk=3)
            cliente = Role.objects.get(pk=4)
            
            Participant.objects.create(user=user, workflowactivity=wa)
            
            for perfil in periles:
                if perfil.user != user:
                    Participant.objects.create(user=perfil.user, workflowactivity=wa)
                
                if perfil.user.is_superuser:
                    wa.assign_role(user, perfil.user, administrador)
                grupos = perfil.user.groups.values_list('name',flat=True)
                if 'Cliente' in grupos:
                    wa.assign_role(user, perfil.user, cliente)
                if 'Coordinador' in grupos:
                    wa.assign_role(user, perfil.user, coordinador)
                if 'Finanzas' in grupos:
                    wa.assign_role(user, perfil.user, finanzas)
                                                        
            viaje.flujo.start(user)
            
            grupos_usuario = user.groups.values_list('name',flat=True)
            if 'Cliente' not in grupos_usuario:
                time.sleep(1)                
                liquidado = Transition.objects.get(pk=11)
                viaje.flujo.progress(liquidado, user)                                
                
            #Guardar destinos
            destinos = formset_entregado.save(commit=False)
            for destino in destinos:
                destino.viaje = viaje
                destino.save()
            
            #Guardar gastos de salida
            gastoss = formset_salida.save(commit=False)
            for gasto in gastoss:
                gasto.viaje = viaje
                gasto.save()
            
            #Guardar gastos de regreso
            gastosr = formset_regreso.save(commit=False)
            for gasto in gastosr:
                gasto.viaje = viaje
                gasto.save()
                
            messages.success(request, 'Se ha agregado un nuevo viaje: %s'%viaje)
            return HttpResponseRedirect(reverse('viajes.views.index',))
        else:
            messages.error(request, 'Ha ocurrido un error. Por favor verifique los datos.')
    else:        
        gastos_salida=[{'tipo': 1,},{'tipo': 1,},]
        gastos_regreso=[{'tipo': 2,},{'tipo': 2,},]
        formset_salida = GastosFormSet(queryset=Gasto.objects.none(), initial=gastos_salida, prefix='salida')
        formset_regreso = GastosFormSet(queryset=Gasto.objects.none(), initial=gastos_regreso, prefix='regreso')
        formset_entregado = EntregasFormSet(queryset=Destino.objects.none(), prefix='destino')
        form = CargaViajeForm()        
        
    context = RequestContext(request,{             
        'form': form,
        'conceptos': conceptos,
        'formset_salida': formset_salida,
        'formset_regreso': formset_regreso,
        'formset_entregado': formset_entregado,
        'action': 'Agregar',
    }) 
                    
    return render_to_response('carga_viaje.html', context)

@login_required
def cargaedit(request, viaje_id):    
    user = request.user
    profile = user.get_profile()
    viaje = get_object_or_404(Viaje, pk=viaje_id)
    conceptos = json.dumps(CONCEPTO_GASTO)
    
    GastosFormSet = modelformset_factory(Gasto, form=GastoHistoricoForm, max_num=14, extra=2, can_delete=True)
    EntregasFormSet = modelformset_factory(Destino, form=EntregadoForm, max_num=10, extra=0) ### Se habilita para guardar fecha.
    movimientoFormSet = modelformset_factory(Movimientos, form=MovContaForm, max_num=17, extra=10, can_delete=True)
    fi= None
    nestatus = None

    pila = []
    pilaO = []
    print 'carga la pila'
    for x in range(1,5):
        saldoC = saldoPartidaXArea(x,2)
        pila.append(saldoC)
        saldoO = saldoPartidaXArea(x,3)
        pilaO.append(saldoO)

    saldoTransferencia = pila.pop()
    pila.pop()
    saldoVale = pila.pop()
    saldoEfectivo = pila.pop()

    saldoTO = pilaO.pop()
    pilaO.pop()
    saldoVO = pilaO.pop()
    saldoEO = pilaO.pop()
    print 'Termina de hacer las pilas '
    

    print 'Inicia los destinos'
    destinos = Destino.objects.filter(viaje = viaje_id)
    print str(viaje_id)
    print destinos
    for dest in destinos:
       print 'destino:' + str(dest.destino_clave_municipio)
    
    if request.method == 'POST':
        formset_regreso = movimientoFormSet(request.POST, prefix='regreso')
        #print formset_regreso.cleaned_data
        formset_entregado = EntregasFormSet(request.POST, request.FILES, queryset=viaje.destino_set.all(), prefix='destino') # este manda a llamar los datos del otro formulario 
        form = CargaViajeForm(request.POST, request.FILES, instance=viaje)
        opera=request.POST['operador']
        econo=request.POST['economico']
        fechaent=request.POST['fecha_ent']
        fechamcia = request.POST['fecha_entmcia']
        v=request.POST['id'] 
        ##print fechaent
        print '*************************************'
        ###print formset_regreso.deleted_forms
        if fechamcia:
            fm = datetime.strptime(fechamcia,"%d/%m/%Y").strftime("%Y-%m-%d") 
            mov2=MovUnidad.objects.filter(id = econo).update(fechamcia = fm)
            print 'actualizo la entrega de mercancia '

        if form.is_valid():
            destinos = form.cleaned_data['destinos'].split(',')
            if destinos[0]:
               for destino in destinos:
                   dest=[]
                   Destino.objects.create(viaje=viaje, destino_clave_municipio=destino)
                   des1= destino.split("-")
                   des11= des1[0]
                   des12= des1[1]
                   municipio = Sepomex.objects.filter(clave_estado = des11, id_asenta_cpcons = des12)#.aggregate(Max('municipio'))
                   des1=municipio[0].municipio
                   d = Destino.objects.filter(viaje=viaje)
                   for des in d:
                      de1= des.destino_clave_municipio.split("-")
                      de11=de1[0]
                      de12=de1[1]
                      mun = Sepomex.objects.filter(clave_estado = de11, id_asenta_cpcons = de12)#.aggregate(Max('municipio'))
                      des1=mun[0].municipio
                      dest.append(des1)
                   MovUnidad.objects.filter(id=econo).update(destino='-'.join(dest))
            else:
                des1='Sin Destino'
       #for modelos in model:
                   #    m=modelos.modelo
                   #MovUnidad.objects.filter(unidad=uni).update(
                   #usuario = u1.id, operador = o, tipo = t, modelo=m, obs = obs, destino = des1, ts=st, destino2= viaje, cliente= c, departamento= d, fechai= fi, email = 0 )  
        ##if  formset_regreso.is_valid() and form.is_valid() and formset_entregado.is_valid(): ##se modifica para inluir la fecha de entrega
        if formset_regreso.is_valid() and form.is_valid():
            print 'pasa el segundo if 2do'
            viaje = form.save(commit= False)
            #print fechaent
            #viaje.fecha_entrega = fechaent
            movConta = formset_regreso.save(commit=False)
            destinos = Destino.objects.filter(viaje = viaje)

            for movC in movConta:
                print movC.pk
                saldo_partida_actual = saldoPartidaXArea(movC.idPartida, movC.idArea)
                movC.id_auth_user = user
                movC.tipo='C'
                movC.ref_viaje=viaje.pk
                movC.montoPartida = saldo_partida_actual
                movC.save()

            for destino in destinos:
                destino.viaje = viaje
                destino.fecha_entrega = fi
                destino.save()
            viaje= get_object_or_404(Viaje, pk=str(viaje))
            current = viaje.flujo.current_state()
            viaje2 = viaje.id 
            print viaje2
            noFlujo = viaje.flujo_id
            wfh = WorkflowHistory.objects.filter(workflowactivity_id = noFlujo)
            wfhf = WorkflowHistory.objects.filter(id = wfh[0].id)
            for wf in wfhf:
                wfh_id=wf.id
                wfh_st=wf.state_id
                wf.state_id = 10
                wf.save()

            messages.success(request, 'Se ha editado el viaje: %s'%viaje)

        try: 
            print 'Entra al Try '
            IMG1=str(request.FILES['evidencia'])

            cut_IMG1 = IMG1.split('.')
            extIMG1 = cut_IMG1[1]
            #idm = ReporteUnidad
            if extIMG1 !='jpg':
                messages.error(request, 'Error Foto 1: el archivo (%s) no es de extension .jpg '%request.FILES['imagen1'])
            else:
                if request.FILES['evidencia'] is not None or extIMG1=='jpg':
                    clie = request.POST['cliente']
                    referencia=request.POST['referencia']
                    fecha=datetime.today()
                    via=request.POST['id']
                    c = int(clie)
                    cli=Cliente.objects.get(id=c)
                    CargaPdf.objects.create(docfile=IMG1, activo = 1, viaje= via, idCliente=cli, bill = referencia, fecha_ts = fecha)
                    messages.success(request, 'La Foto 1: (%s) se agrego correctamente '%request.FILES['evidencia'])    
                    lista1 = os.listdir(settings.MEDIA_ROOT + 'imgEvidencia/')

                    for lis1 in lista1:
                        PARTE1 = str(lis1)
                        print 'Valor de Parte 1:' + PARTE1
                        num1=500
                        if PARTE1 == str(request.FILES['evidencia']):
                            llave1 = str(viaje.pk)
                            NUEVAPARTE1 = PARTE1.split('.')
                            newimg1=llave1+'_'+'1.'+NUEVAPARTE1[1]
                            os.rename (settings.MEDIA_ROOT + 'imgEvidencia/%s'%PARTE1, settings.MEDIA_ROOT + 'imgEvidencia/%s'%newimg1)
                        num1 +=1
        except Exception:
            print "error foto 1"
        else:
            messages.error(request, 'Ha ocurrido un error. Por favor verifique los datos.')

        val_des = Destino.objects.filter(viaje = viaje)
        print 'Longitud de val_des : '
        print len(val_des)

        if fechaent and len(val_des) != 0:
            print 'Paso la validacion del destino'
            oper = empleado.objects.filter(id = opera)
            for l in oper:
                serv = l.servicios
                serv = serv - 1
            econ=Economico.objects.filter(id=econo)
            for e in econ:
                envia = e.enviaje
                envia = envia - 1
            if envia >= 1:
                nestatus = 9
            elif envia <= 0: 
                nestatus = 10 
            fi = datetime.strptime(fechaent,"%d/%m/%Y").strftime("%Y-%m-%d")
            stan=empleado.objects.filter(id = opera).update(servicios = serv)
            estan=Economico.objects.filter(id = econo).update(status = nestatus, enviaje = envia)
            mov=MovUnidad.objects.filter(id = econo).update(tipo=nestatus, fechaf=None, enviaje=envia, operador = None, destino = None, destino2 = None, 
                fechai = None, usuario = None, obs = None, departamento = None, cliente = None, fechamcia= None)
            print 'actualizo los valores'
            des=Destino.objects.filter(viaje_id= v).update(fecha_entrega = fi)        
            viaje=form.save()
            viaje= get_object_or_404(Viaje, pk=str(viaje))
            current = viaje.flujo.current_state()
            viaje2 = viaje.id 
            print viaje2
            noFlujo = viaje.flujo_id
            wfh = WorkflowHistory.objects.filter(workflowactivity_id = noFlujo)
            wfhf = WorkflowHistory.objects.filter(id = wfh[0].id)
            for wf in wfhf:
                wfh_id=wf.id
                wfh_st=wf.state_id
                wf.state_id = 11
                wf.save()            


        else:
            print 'No econtro destinos'
            fi = None;
            viaje.fecha_ent = fi
            viaje=form.save()

            messages.error(request, 'No Se ha Cerrado el Viaje: %s , por que no cuenta con destinos.'%viaje)

        return HttpResponseRedirect(reverse('viajes.views.index',))
    else:
        gastos_salida=[{'tipo': 1,'viaje': viaje_id,},{'tipo': 1,'viaje': viaje_id,},]
        formset_salida = GastosFormSet(initial=gastos_salida, queryset=viaje.gasto_set.filter(tipo=1), prefix='salida')
        formset_regreso = movimientoFormSet(queryset=Movimientos.objects.filter(ref_viaje=viaje.pk), prefix='regreso')
	formset_entregado = EntregasFormSet(queryset=viaje.destino_set.all(), prefix='destino')
 	form = CargaViajeForm(instance=viaje) 
       
        
    context = RequestContext(request,{ 
	'viaje': viaje,            
        'form': form,
        'conceptos': conceptos,
        'formset_regreso': formset_regreso,
        'formset_entregado': formset_entregado,
        'saldoEfectivo':saldoEfectivo,
        'saldoVale':saldoVale,
        'saldoTransferencia':saldoTransferencia,
        'saldoEO':saldoEO,
        'saldoVO':saldoVO,
        'saldoTO':saldoTO,
        'action': 'Agregar',
    }) 
    ##view_context['formset_entregado'] = formset_entregado                
    return render_to_response('carga_gastos.html', context)

def rutas(request):
    rutas_list = Ruta.objects.all()
    paginator = Paginator(rutas_list, 10)
    page = request.GET.get('page')

    try:
        rutas = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        rutas = paginator.page(1)
    except EmptyPage:
        # If page is out of range (e.g. 9999), deliver last page of results.
        rutas = paginator.page(paginator.num_pages)
    
    context = RequestContext(request,{ 
        'rutas': rutas,                         
    })    
    
    return render_to_response('rutas.html', context)

def rutasadd(request, ruta_id = None):

    if request.method == 'POST':
        form = RutaForm(request.POST)

        if form.is_valid():
            ruta = form.save()
            messages.success(request, 'Se ha agregado la ruta %s'%ruta)
            return HttpResponseRedirect(reverse('viajes.views.rutas',))
        else:
            messages.error(request, 'Ha ocurrido un error. Por favor verifique los datos.')
    else:        
        if ruta_id:
            ruta = get_object_or_404(Ruta, pk=ruta_id)
            ruta.nombre = ''
            form = RutaForm(instance=ruta)
        else:
            form = RutaForm()
    
    context = RequestContext(request,{ 
        'form': form,
        'action': 'Agregar',
    })    
    
    return render_to_response('rutas_form.html', context)

@login_required
@csrf_exempt
def rutasmodal(request):    
    status = 'OK'
    use_json = False
    ruta = None
    if request.method == 'POST':
        use_json = True
        form = RutaForm(request.POST)
        if form.is_valid():
            ruta = form.save()
        else:
            status = 'Error'
            #messages.error(request, 'Ha ocurrido un error. Por favor verifique los datos.')
    else:
        form = RutaForm()
    
    context = RequestContext(request,{ 
        'ruta': ruta,
        'form': form,
        'action': 'Agregar',
    })    
    
    if use_json:
        t = loader.get_template('rutamodal_form.html')
        html = t.render(context)
        if status != 'OK':
            return HttpResponse(json.dumps({'status':status, 'html':html}), mimetype="application/json")
        else:
            return HttpResponse(json.dumps({'status':status, 'ruta':ruta.nombre, 'id':ruta.pk}), mimetype="application/json")
    else:    
        return render_to_response('rutamodal_form.html', context)

def rutasedit(request, ruta_id):
    ruta = get_object_or_404(Ruta, pk=ruta_id)
    
    if request.method == 'POST':
        form = RutaForm(request.POST, instance=ruta)
        if form.is_valid():
            form.save()
            messages.success(request, 'Se ha editado la ruta %s'%ruta)
            return HttpResponseRedirect(reverse('viajes.views.rutas',))
        else:
            messages.error(request, 'Ha ocurrido un error. Por favor verifique los datos.')
    else:
        form = RutaForm(instance=ruta)
        
    context = RequestContext(request,{ 
        'form': form,
        'action': 'Editar',
    })    
    
    return render_to_response('rutas_form.html', context)

class CartaPorteView(PDFTemplateView):
    template_name = "carta_porte.html"
    
    def get_context_data(self, pk, **kwargs):

        viaje = Viaje.objects.get(pk=pk)
        movimientos = Movimientos.objects.filter(ref_viaje=viaje.pk)
        totalMov = 0
        for mov in movimientos:
            totalMov += mov.importe

        des = Destino.objects.filter(viaje=pk)
        if des:
           for dest in des:
               cve = dest.destino_clave_municipio
               print 'Este es el municipio'
               print cve
           clave= cve.split("-")
           ce = clave[0]
           iac= clave[1]
           ciudades = Sepomex.objects.filter(clave_estado= ce, id_asenta_cpcons = iac)
           c1=ciudades[0].ciudad
           ##print str(c1)
        else:
           c1 = 'Local'    
        #for ciudad in ciudades:
        #    c = ciudad.ciudad
        #    print 'Estas son todas las ciudades'

        #importe = viaje.gastos_salida_total()*Decimal(1.20)
        importe = totalMov*Decimal(1.08)
        iva = importe*Decimal(0.16)
        subtotal = importe+iva
        retencion = importe * Decimal(0.04)
        total = subtotal - retencion
        importe_letra = to_word(total)


        
        return super(CartaPorteView, self).get_context_data(
            pagesize="A4",
            title="Carta Porte " + pk,
            viaje=viaje,
            importe=importe,
            iva=iva,
            subtotal=subtotal,
            retencion=retencion,
            total=total,
            importe_letra=importe_letra,
            movimientos = movimientos,
            totalMov = totalMov,
            ciudad= c1,
            **kwargs
        )
        
        
def reportes(request):
        #Obtener catálogos    
    searchform = ReportForm()
    
    context = RequestContext(request,{ 
        'searchform': searchform
    })    
    
    return render_to_response('reportes.html', context)
 

@csrf_exempt
def reportesresults(request):
    
    #if request.method != 'POST':
    campos = []
    searchform = ReportForm(request.POST)
    if searchform.is_valid():    
        campos = searchform.cleaned_data['campos']
            
          
    filters = {}    
    if ('pk' in request.POST) and request.POST['pk'].strip():        
        filters['workflowactivity__pk'] = request.POST['pk']
    if ('status' in request.POST) and request.POST['status'].strip():
        filters['state__pk'] = request.POST['status']
    if ('fecha_from' in request.POST) and request.POST['fecha_from'].strip():
        filters['workflowactivity__fecha_salida__gte'] = request.POST['fecha_from']
    if ('fecha_to' in request.POST) and request.POST['fecha_to'].strip():
        filters['workflowactivity__fecha_salida__lte'] = request.POST['fecha_to']
    if ('cliente' in request.POST) and request.POST['cliente'].strip():
        filters['workflowactivity__cliente__pk'] = request.POST['cliente']
    if ('departamento' in request.POST) and request.POST['departamento'].strip():
        filters['workflowactivity__departamento__pk'] = request.POST['departamento']        
    if ('referencia' in request.POST) and request.POST['referencia'].strip():
        filters['workflowactivity__referencia__icontains'] = request.POST['referencia']
    if ('economico' in request.POST) and request.POST['economico'].strip():
        filters['workflowactivity__economico__pk'] = request.POST['economico']
    if ('operador' in request.POST) and request.POST['operador'].strip():
        filters['workflowactivity__operador__pk'] = request.POST['operador']
    if ('destino' in request.POST) and request.POST['destino'].strip():
        destinos = Destino.objects.filter(destino_clave_municipio=request.GET['destino']).values_list('viaje_id', flat=True)        
        filters['workflowactivity__id__in'] = destinos        
                
        
    viajes = ViajeStatus.objects.filter(**filters)          
    context = RequestContext(request,{ 
        'viajes': viajes,
        'campos': campos,
        'action': 'Editar',
    })    
    
    return render_to_response('results.html', context)   

@csrf_exempt

def reportsexcel(request):
    response = HttpResponse(mimetype="application/vnd.ms-excel")
    response['Content-Disposition'] = 'attachment; filename=reporte.xls'
    
    campos = []
    searchform = ReportForm(request.POST)
    if searchform.is_valid():    
        campos = searchform.cleaned_data['campos']
            
          
    filters = {}    
    if ('pk' in request.POST) and request.POST['pk'].strip():        
        filters['workflowactivity__pk'] = request.POST['pk']
    if ('status' in request.POST) and request.POST['status'].strip():
        filters['state__pk'] = request.POST['status']
    if ('fecha_from' in request.POST) and request.POST['fecha_from'].strip():
        filters['workflowactivity__fecha_salida__gte'] = request.POST['fecha_from']
    if ('fecha_to' in request.POST) and request.POST['fecha_to'].strip():
        filters['workflowactivity__fecha_salida__lte'] = request.POST['fecha_to']
    if ('cliente' in request.POST) and request.POST['cliente'].strip():
        filters['workflowactivity__cliente__pk'] = request.POST['cliente']
    if ('departamento' in request.POST) and request.POST['departamento'].strip():
        filters['workflowactivity__departamento__pk'] = request.POST['departamento']        
    if ('referencia' in request.POST) and request.POST['referencia'].strip():
        filters['workflowactivity__referencia__icontains'] = request.POST['referencia']
    if ('economico' in request.POST) and request.POST['economico'].strip():
        filters['workflowactivity__economico__pk'] = request.POST['economico']
    if ('operador' in request.POST) and request.POST['operador'].strip():
        filters['workflowactivity__operador__pk'] = request.POST['operador']
    if ('destino' in request.POST) and request.POST['destino'].strip():
        destinos = Destino.objects.filter(destino_clave_municipio=request.GET['destino']).values_list('viaje_id', flat=True)        
        filters['workflowactivity__id__in'] = destinos        
                
        
    viajes = ViajeStatus.objects.filter(**filters)      
    if viajes.count():        
        wb = xlwt.Workbook()
        ws = wb.add_sheet('Reporte')
        
        date_format = xlwt.XFStyle()
        date_format.num_format_str = 'dd/mm/yyyy'
                
        columna = 0
        for campo in campos:        
            ws.write(1, columna, dict(CAMPOS)[campo])
            columna += 1    
        
        primitive = [int, str, bool, long, Decimal]
        renglon = 2
        fechaXls = False
        for viaje in viajes:                         
            columna = 0
            for campo in campos:                    
                campo = campo.replace("__", ".")
                print 'campo ' + str(campo) + ' >> '+str(viaje)
                if campo == 'workflowactivity.fecha_salida' or campo == 'workflowactivity.facturacion_fecha_fac' or campo == 'workflowactivity.facturacion_fecha_pago':
                    fechaXls = True
                value = multi_getattr(viaje, campo)
                #print 'resultado ' + str(value)
                if type(value) not in primitive:
                    try:
                        value = value.__unicode__()
                    except AttributeError:
                        value = value
                if fechaXls:
                    ws.write(renglon, columna, value, date_format)
                else:
                    ws.write(renglon, columna, value)
                columna += 1                             
            renglon += 1
    
        wb.save(response)
        return response     
    else:
        return HttpResponse("<script>alert('No se encontrarón registros con los criterios seleccionados.');window.close();</script>")

def multi_getattr(obj, attr, default = None):
    """
    Get a named attribute from an object; multi_getattr(x, 'a.b.c.d') is
    equivalent to x.a.b.c.d. When a default argument is given, it is
    returned when any attribute in the chain doesn't exist; without
    it, an exception is raised when a missing attribute is encountered.

    """
    attributes = attr.split(".")
    for i in attributes:
        try:
            if i == 'solicitante_id':
                obj = obj.solicitante.user.get_full_name()
            elif i == 'beneficiario_id':
                obj = obj.beneficiario. __unicode__()
            elif i == 'benef_otros_id':
                obj = obj.benef_otros.__unicode__()
            elif i == 'concepts_id':
                obj = obj.concepts.__unicode__()
            elif i == 'sol_bill':
                obj = obj.sol_bill.pk
            elif i == 'referencia':
                obj = obj.referencia
            else:
                obj = getattr(obj, i)

        except (AttributeError, Exception):
                return default
                #if default:
                #    return default
                #else:
                #    raise

    return obj




#Modificacion Carga Excel
def subirExcel(request): 
    user = request.user
    profile = user.get_profile()
    participant = user.participant_set.first()
    user = request.user

    if request.method == 'POST':
        form = UploadExcelForm(request.POST, request.FILES)
        #form.save()
        if form.is_valid():
            newdoc = ExcelViaje(docfile = request.FILES['docfile'], usuario = user)
            nombreArchivo = str(newdoc.docfile)
            nombreArchivo = nombreArchivo.upper()


            if nombreArchivo.endswith('.XLSX'):
                newdoc.save()
                actualizaFacturas(request, str(newdoc.docfile))
                ##ExcelViaje.objects.all().delete() Evitar que se borre el registro.
            else: 
                print 'Archivo con formato invalido'
                txtmsg = 'Archivo con formato inválido.'
                messages.error(request, txtmsg)  
              
        return render(request,'subirExcel.html', {'form': form,})
    else:
        form = UploadExcelForm()
    return render(request,'subirExcel.html', {'form': form,})


def actualizaFacturas(request, datoArchivo):

    user = request.user
    profile = user.get_profile()
    participant = user.participant_set.first()
    user = request.user

  
    workbook = openpyxl.load_workbook(filename=settings.MEDIA_ROOT + datoArchivo, use_iterators=True)

    try:
        sheet = workbook.get_sheet_by_name('fac_cm')
        exitos = 0
        numError = 0
        total = 0    
        for row in range(14, sheet.get_highest_row()):
            total += 1
            print sheet['B' + str(row)].value            
            if sheet['B' + str(row)].value == 'FIN': break

            g=sheet['G'+str(row)].value
            if type(g) not in (int, float, long, complex) or g == None:
                g=0
            else:
                g=g
            print g
            h=sheet['H'+str(row)].value
            if type(h) not in (int, float, long, complex) or h == None:
                h=0
            else:
                h=h
            i=sheet['I'+str(row)].value
            if type(i) not in (int, float, long, complex) or i == None:
                i=0
            else:
                i=i
            print i
            j=sheet['J'+str(row)].value
            if type(j) not in (int, float, long, complex) or j == None:
                j=0
            else:
                j=j
            k=sheet['K'+str(row)].value
            if type(k) not in (int, float, long, complex) or k == None:
                k=0
            else:
                k=j
            l=sheet['L'+str(row)].value
            if type(l) not in (int, float, long, complex) or l == None:
                l=0
            else:
                l=l
            m=sheet['M'+str(row)].value
            if type(m) not in (int, float, long, complex) or m== None:
                m=0
            else:
                m=m
            n=sheet['N'+str(row)].value
            if type(n) not in (int, float, long, complex) or n == None:
                n=0
            else:
                n=n
            p=sheet['P' + str(row)].value
            if type(p) not in (int, float, long, complex) or n == None:
                p = 0
            else:
                p=p 
            x=sheet['X'+str(row)].value
            if x == None:
                x='NoCapturado'
            else:
                x=x
            y=sheet['Y'+str(row)].value
            if y == None:
                y='NoCapturado'
            else:
                y=y
            z=sheet['Z'+str(row)].value
            if  z == None:
                z='NoCapturado'
            else:
                z=z
            aa=sheet['AA'+str(row)].value
            if type(aa) not in (int, float, long, complex) and aa == None:
                aa=0
            else:
                aa=aa
            ab=sheet['AB'+str(row)].value
            if type(ab) not in (int, float, long, complex) and ab == None:
                ab=0
            else:
                ab=ab
            ac=sheet['AC'+str(row)].value
            if type(ac) not in (int, float, long, complex) and ac == None:
                ac=0
            else:
                ac=ac
            ad=sheet['AD'+str(row)].value
            if type(ad) not in (int, float, long, complex) and ad == None:
                ad=0
            else:
                ad=ad
            w=sheet['W'+str(row)].value
            
            print "Valor de g" + str(g)
            print "Valor de h" + str(h)
            print "Valor de i" + str(i)
            print "Valor de j" + str(j)
            print "Valor de k" + str(k)    
            print "Valor de l" + str(l)
            print "Valor de m" + str(m)
            print "Valor de n" + str(n)
            print "Valor de w" + str(w)
            print "Valor de x" + str(x)
            print "Valor de y" + str(y)
            print "Valor de z" + str(z)
            print "Valor de aa" + str(aa)
            print "Valor de ab" + str(ab)
            print "Valor de ac" + str(ac)
            
            
            ##o=sheet['O'+str(row)].value
            #if type(o) not in (int, float, long, complex):
            #    o=0
            #else:
            #    o=o

            subtotal = g+h+i+j+k+l+m+n
            iva = subtotal * .16
            retencion = g *.04
            total = subtotal + iva - retencion
            print 'Este es el total:'
            print total
            try:
                participant = user.participant_set.first()

                if Viaje.objects.filter(referencia=sheet['B' + str(row)].value).exists():
                       viaje =get_object_or_404(Viaje, referencia=sheet['B' + str(row)].value)
                       current = viaje.flujo.current_state()
                       viaje2 = viaje.id 
                       print viaje2
                       noFlujo = viaje.flujo_id
                       wfh = WorkflowHistory.objects.filter(workflowactivity_id = noFlujo)
                       wfhf = WorkflowHistory.objects.filter(id = wfh[0].id)                       

                       #for wf in wfhf:
                       #     wfh_id=wf.id
                       #     wfh_st=wf.state_id
                       #     wf.state_id = 8
                       #     wf.save()

                       Viaje.objects.filter(referencia=sheet['B' + str(row)].value).update(
                            factura = sheet['A' + str(row)].value,
			                 facturacion_flete = g,
                             facturacion_maniobra = h,
                             facturacion_casetas = i,
                             facturacion_otros = j + p , 
                           ##facturacion_reparto = sheet['??' + str(row)].value,
                             facturacion_incentivo = k,
                             facturacion_penalizacion = l,
                             facturacion_detencion = m,
                             facturacion_desvio = n,
                             factura_subtotal= subtotal,
                             factura_iva = iva, 
                             factura_retenciones = retencion,
                             factura_total= total,  
                           ##facturacion_ferri = sheet['??' + str(row)].value,
                             facturacion_fecha_fac = sheet['W' + str(row)].value,
                             facturacion_forma_pago = str(x),
                             facturacion_banco = str(y),
                             facturacion_documento = str(z), 
                             maniobras_locales = ab,
			                 maniobras_foraneas = ac,
			                 maniobras_retrabajos = ad,
			                 casetas_lg = aa)

                       Movimientos.objects.filter(ref_viaje=viaje.id).update(
			                 maniobras_locales = ab,
                             maniobras_foraneas =ac,
                             maniobras_retrabajos = ad,
                             casetas_lg = aa)
                       print 'Termina de actulizar la tabla Movimientos'
                            #Cambia de facturacion a facturado
                            #transition = current.state.transitions_from.get(pk=7)
                            #viaje.flujo.progress(transition, user)                	
                       exitos += 1
                       for wf in wfhf:
                            wfh_id=wf.id
                            wfh_st=wf.state_id
                            wf.state_id = 8
                            wf.save()
                       facturas=Facturas.objects.filter(factura = sheet['A'+str(row)].value)
                       if facturas:
                            for factura in facturas:
                                monto= long(factura.monto)
                                monton = monto + total
                                Facturas.objects.filter(factura= sheet['A'+str(row)].value).update(monto= monton, saldo = monton) 
                       else:
                            Facturas.objects.create(factura = sheet['A'+str(row)].value, monto=total, saldo = total, fecha_emision = sheet['W' + str(row)].value) 




                else:
                    messages.warning(request, u'No existe la referencia ' + str(sheet['B' + str(row)].value))
                    numError += 1
            except Exception:
                messages.error(request, u'Error al registrar la referencia ' + str(sheet['B' + str(row)].value) + u' La referencia pertenece a 2 viajes diferentes, favor de reportar a sistemas')
                numError += 1
    
        archivoF = settings.MEDIA_ROOT + 'upload/facturas.xlsx'
        #os.remove(archivoF)  ## Evitar que se elimine el Excel.
        if exitos == 0: 
            messages.error(request, 'No se realizaron actualizaciones')
        if numError == 0:
            messages.success(request, 'Registros actualizados exitosamente.')
        if exitos > 0 and numError > 0:
            messages.success(request, 'Actualización parcial realizada.')
       
    except KeyError:
        messages.error(request, 'Archivo no contiene la hoja fac_cm')

def index_fact(request):       ###Modificacion Facturas 
    #Obtener catálogos    
    searchform = SearchFormFact(request.GET)
    
    filters = {}    
    if ('pk' in request.GET) and request.GET['pk'].strip():        
        filters['workflowactivity__pk'] = request.GET['pk']
    if ('status' in request.GET) and request.GET['status'].strip():
        filters['state__pk'] = request.GET['status']
    if ('fecha_from' in request.GET) and request.GET['fecha_from'].strip():
        filters['workflowactivity__fecha_salida__gte'] = request.GET['fecha_from']
    if ('fecha_to' in request.GET) and request.GET['fecha_to'].strip():
        filters['workflowactivity__fecha_salida__lte'] = request.GET['fecha_to']
    if ('cliente' in request.GET) and request.GET['cliente'].strip():
        filters['workflowactivity__cliente__pk'] = request.GET['cliente']
    if ('factura' in request.GET) and request.GET['factura'].strip():
	filters['workflowactivity__factura']=request.GET['factura']
    if ('departamento' in request.GET) and request.GET['departamento'].strip():
        filters['workflowactivity__departamento__pk'] = request.GET['departamento']        
    if ('referencia' in request.GET) and request.GET['referencia'].strip():
        filters['workflowactivity__referencia__icontains'] = request.GET['referencia']
    if ('economico' in request.GET) and request.GET['economico'].strip():
        filters['workflowactivity__economico__pk'] = request.GET['economico']
    if ('operador' in request.GET) and request.GET['operador'].strip():
        filters['workflowactivity__operador__pk'] = request.GET['operador']
    if ('destino' in request.GET) and request.GET['destino'].strip():
        destinos = Destino.objects.filter(destino_clave_municipio=request.GET['destino']).values_list('viaje_id', flat=True)        
        filters['workflowactivity__id__in'] = destinos        
                
        
    viajes_list = ViajeStatus.objects.filter(**filters)  
    if not bool(filters):
        viajes_list = viajes_list.order_by('-workflowactivity__pk')          
         
    paginator = Paginator(viajes_list, 500)
    page = request.GET.get('page')
    
    try:
        viajes = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        viajes = paginator.page(1)
    except EmptyPage:
        # If page is out of range (e.g. 9999), deliver last page of results.
        viajes = paginator.page(paginator.num_pages)
    
        
    context = RequestContext(request,{             
        'searchform': searchform,  
        'viajes': viajes,
    })    
    
    return render_to_response('viajes_fact.html', context)        
#####################################################################################
#masivo PDF (uno en uno)
def subirPDFMasivo(request):
    if request.method == 'POST':
        form = DocumentoForma(request.POST, request.FILES)
        if form.is_valid():
            newdoc = CargaPdf(docfile = request.FILES['docfile'])
            nombreArchivo = str(newdoc.docfile)
            newnombre= nombreArchivo.replace('.pdf', '')
            checaVal = Viaje.objects.filter(referencia=newnombre)
            if nombreArchivo.endswith('.pdf'):
                for nom in request.FILES.getlist('docfile'):
                    NOM = str(nom)
                    cut_NOM = NOM.split('.')
                    ext = cut_NOM[1]
                    if ext == 'pdf':
                        checaVal = Viaje.objects.filter(referencia=cut_NOM[0])
                        try:
                            checaVall= checaVal[0].referencia
                            viajeval= checaVal[0].id
                            print 'Este es el viaje'
                            print viajeval

                            Tab = CargaPdf.objects.filter(docfile=checaVall + ".pdf")
                            try:
                                checaTab= Tab[0].docfile
                                if checaTab:
                                    txtmsg = 'El archivo %s,  ya existe no se puede duplicar' % checaTab
                                    messages.error(request, txtmsg)
                                else:
                                    print ""
                            except Exception:
                    ### se carga el archivo.
                                subirArchivoPdf(nom)
                                cliente = Cliente(id=checaVal[0].cliente.pk)
                                newdoc2 = CargaPdf(docfile = checaVall + ".pdf", idCliente= cliente, viaje=viajeval, bill=checaVall)
                                print newdoc2
                                newdoc2.save()
                                txtmsg = 'El Archivo %s, se cargo correctamente!!!' % nom
                                messages.success(request, txtmsg)
                        except Exception:
                            txtmsg = 'Archivo %s,  no tiene referencia con viajes' % nom
                            messages.error(request, txtmsg)
                    else:
                        txtmsg = 'Archivo %s,  con formato inválido' % nom
                        messages.error(request, txtmsg)
            else:
                txtmsg = 'Archivo con formato inválido.'
                messages.error(request, txtmsg)
        else:
            messages.error(request, 'Por favor seleccione un archivo (.pdf)')  
        return render(request,'subirPDFMasivo.html', {'form': form,})
    else:
        form = DocumentoForma()
    return render(request,'subirPDFMasivo.html', {'form': form,})

def subirArchivoPdf(f):
    #destino = open('/var/www/PDFS/%s'%f.name, 'wb+')
    destino = open(settings.MEDIA_ROOT + 'pdfs/%s'%f.name, 'wb+')
    for chunk in f.chunks():
        destino.write(chunk)
    destino.close()
############################

### LISTADO PDF
@login_required
def listadoPdf(request):

    #Obtener catálogos    
    searchform2 = DocumentoForma(request.FILES)
    
    filters = {}    
    if ('docfile1' in request.GET) and request.GET['docfile1'].strip():
       filters['id'] = request.GET['docfile1']
             
    viajes_list = CargaPdf.objects.filter(**filters)  
    if not bool(filters):
        viajes_list = viajes_list.order_by('id')
         
    paginator = Paginator(viajes_list, 100)
    page = request.GET.get('page')
    
    try:
        viajes = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        viajes = paginator.page(1)
    except EmptyPage:
        # If page is out of range (e.g. 9999), deliver last page of results.
        viajes = paginator.page(paginator.num_pages)
    
        
    context = RequestContext(request,{             
        'searchform2': searchform2,  
        'viajes': viajes,
    })    
    
    return render_to_response('listadoPdf.html', context)


def reportPdf(request, viaje_nombre):
    nombreArchivo = viaje_nombre
    try:
     #with open('/var/www/PDFS/%s' % nombreArchivo , 'r') as pdf:
     with open(settings.MEDIA_ROOT + 'pdfs/%s' % nombreArchivo , 'r') as pdf:
        response = HttpResponse(pdf.read(), mimetype='application/pdf')
        return response
    except IOError:
        return HttpResponse("<script>alert('No se puede descargar el Archivo');window.close();</script>")    
    pdf.closed   

def pagPdf(request):     
    return render_to_response('subirPDFMasivo.html')

def listaPdf(request):

    return render_to_response('listadoPdf.html')

@login_required
def colocarMensaje(request, cliente_id, archivo):

    if request.method == 'POST':
        form = MensajeCorreoForm(request.POST)
        if form.is_valid():
            destinatario = request.POST['destinatario']
            nomArchivo = request.POST['archivoAdjunto']
            mensaje = request.POST['mensaje']
            varios = destinatario.split(",")
            correo = EmailMessage('Información desde Vitesse', mensaje, 'controladministrativo@logisticavitesse.com.mx', varios) #### Cambio del envio del correo de ofarias0424@gmail.com a 'controladministrativo@logisticavitesse.com.mx'
            ruta = settings.MEDIA_ROOT + "pdfs/"+ str(nomArchivo)
            correo.attach_file(ruta)
            correo.send()
            messages.success(request, 'Correo enviado exitosamente a  %s' % destinatario)

            return HttpResponseRedirect(reverse('viajes.views.listadoPdf',))
        else:
            messages.error(request, 'Ha ocurrido un error. Por favor verifique que el mensaje no este vacío.')
    else:

        cliente = Cliente.objects.get(pk=cliente_id)
        datos = {'destinatario' : cliente.correo, 'archivoAdjunto' : archivo}
        form = MensajeCorreoForm(initial=datos)

    context = RequestContext(request,{
        'action': 'Enviar Correo',
        'form' : form
    })
    return render_to_response('mensajeCorreo.html', context)

@login_required
def reporteManiobras(request):
    form = reporteForm()
    conceptos = json.dumps(CONCEPTO_GASTO)
    return render(request,'reporteManiobras.html', {'form': form,'conceptos':conceptos,})


@csrf_exempt
@login_required
def generaReporteMani(request):
    filters = {}  # Genera Variables de tipo diccionario
    filtersViaje = {}
    filtersSol = {}
    filterMov = {}
    filterVia={}
    inicial = None #inicia variable para la fecha 
    final = None  #inicia variable para la fecha
    porCP = False # inicia variable para la Carta porte como Booleano 
  
    if ('fecha_ini') in request.POST and request.POST['fecha_ini'].strip():
        inicial = request.POST['fecha_ini']

    if ('fecha_fin') in request.POST and request.POST['fecha_fin'].strip():
        final = request.POST['fecha_fin']

    #Validacion de los campos
    if inicial and final:
        filters['fecha__range'] = (inicial, final)
        filterMov['fecha__range'] = (inicial, final)
        filterVia['fecha_salida__range'] = (inicial, final)
    elif inicial or final:
        messages.error(request, 'Seleccione completo rango de fechas.')
        form = reporteForm(request.POST)
        return render(request,'reporteManiobras.html', {'form': form,})

    if not len(filters):
        messages.error(request, 'Seleccione por lo menos un filtro de búsqueda o un rango de fechas')
        form = reporteForm(request.POST)
        return render(request,'reporteManiobras.html', {'form': form,})

    response = HttpResponse(mimetype="application/vnd.ms-excel")
    response['Content-Disposition'] = 'attachment; filename=reporte.xls'

    wb = xlwt.Workbook(encoding='utf8')
    ws = wb.add_sheet('Reporte de maniobras')
    date_format = xlwt.XFStyle()
    date_format.num_format_str = 'dd/mm/yyyy'
    formatoNum = XFStyle()
    formatoNum.num_format_str = '$#,##0.00'
    font = xlwt.Font() # Crear Font
    font.name = 'Arial'
    font.height = 20 * 12  #22pt
    font.bold = True
    font.italic = True
    style = xlwt.XFStyle() # Crar Style
    style = xlwt.easyxf('pattern: pattern solid, fore_colour light_blue;') #background color
    style.font = font # Aplicar Font al Style

    columnas = [
            (u"Viaje", 70),
            (u"Fecha Viaje", 70),
            (r'Facturacion Locales', 70),
            (r'Facturacion Foraneas',70),
            (r'Fcturacion Retrabajos', 70),
            (u"Gasto Solicitud", 70),
            (u"Gasto Maniobras Locales", 70),
            (u"Gasto Maniobras Foraneas", 70),
            (u"Gasto Maniobras Retrabajos", 70),

    ]

    posx = 0
    for col in xrange(len(columnas)):
        ws.write(0, posx, columnas[col][0])
        posx += 1
    filterMov['tipo'] = 'C'
    movimientos = None

    if porCP:
        print 'Busqueda por carta porte'
    else:
        movimientos = Movimientos.objects.filter(**filterMov)#.values_list('ref_viaje',flat=True).distinct() Al parecer hasta aqui va bien.
        for movd in movimientos:
            print str(movd)
        fact=Viaje.objects.filter(**filterVia).order_by('-id')
        for v in fact:
            vi=v.id
    renglon = 1
    if fact:#viajes.count():
        for v in fact:
            print str(vi) + ' <--'
            vi = v.id
            soli = vi
            print soli
            solicitudes = Solicitudes.objects.filter(Q(concepts_id = 1)| Q(concepts_id = 2)| Q(concepts_id = 24), sol_bill_id = soli).aggregate(Sum('importe_asig'))
            sols = solicitudes.keys()
            print sols
            valores = solicitudes.values()
            for valor in valores: 
                a = valor
                if a == None:
                    a = 0

            viagastosml = Movimientos.objects.filter(concepto = 'Maniobras', ref_viaje = soli).aggregate(Sum('importe'))
            vml = viagastosml.values()
            for valvl in vml:
                valml = valvl
                if valml == None:
                    valml = 0                 
            viagastosmf = Movimientos.objects.filter(concepto = 'Maniobras Foraneas', ref_viaje = soli).aggregate(Sum('importe'))
            vmf = viagastosmf.values()
            for valvf in vmf:
                valmf = valvf
                if valmf == None:
                    valmf = 0
            viagastosmr = Movimientos.objects.filter(concepto = 'Maniobras Retrabajos', ref_viaje = soli).aggregate(Sum('importe'))
            vmr = viagastosmr.values()
            for valvr in vmr:
                valmr = valvr
                if valmr == None:
                    valmr = 0
                
            print 'Paso el filtro de solicitudes'
            viaje = vi
            fecha = str(v.fecha_salida)
            ws.write(renglon, 0, vi)
            ws.col(0).width=3000
            ws.write(renglon, 1, fecha)
            ws.col(0).width=3000
            ws.write(renglon, 2, v.maniobras_locales)
            ws.col(0).width=3000
            ws.write(renglon, 3, v.maniobras_foraneas, formatoNum)
            ws.col(0).width=3000
            ws.write(renglon, 4, v.maniobras_retrabajos, formatoNum)
            ws.col(0).width=3000
            ws.write(renglon, 5, a, formatoNum)
            ws.col(0).width=3000
            ws.write(renglon, 6, valml, formatoNum)
            ws.col(0).width=3000
            ws.write(renglon, 7, valmf, formatoNum)
            ws.col(0).width=3000
            ws.write(renglon, 8, valmr, formatoNum)
            ws.col(0).width=3000                       
            renglon += 1

    elif porCP:
        via = Viaje.objects.get(pk=filtersViaje['pk'])
        filtersSol['sol_bill'] = via.pk
        solicitudes = Solicitudes.objects.filter(**filtersSol)
        if solicitudes.count():
            print 'Tiene solicitudes relacionadas, obtener movimientos'
            for sol in solicitudes:
                filterMov['ref_solicitud'] = sol.pk
                movSol = Movimientos.objects.filter(**filterMov)
                for mov in movSol:
                    ws.write(renglon, 0, via.pk)
                    ws.col(0).width=3000
                    ws.write(renglon, 1, mov.pk)
                    ws.col(0).width=3000
                    ws.write(renglon, 2, via.maniobras_locales)
                    ws.col(0).width=3000
                    ws.write(renglon, 3, via.maniobras_foraneas)
                    ws.col(0).width=3000
                    ws.write(renglon, 4, via.maniobras_retrabajos)
                    ws.col(0).width=3000
                    ws.write(renglon, 5, mov.concepto)
                    ws.col(0).width=3000
                    ws.write(renglon, 6, mov.importe,formatoNum)
                    ws.col(0).width=3000
                    ws.write(renglon, 7, sol.cliente_paga,formatoNum)
                    ws.col(0).width=3000
                    ws.write(renglon, 8, sol.importe_asig,formatoNum)
                    ws.col(0).width=3000
                    renglon += 1
                del filterMov['ref_solicitud']


        print 'Obtener directamente los movimientos'
        filterMov['ref_viaje'] = via.pk
        movViaje = Movimientos.objects.filter(**filterMov)
        for mov in movViaje:
            ws.write(renglon, 0, via.pk)
            ws.col(0).width=3000
            ws.write(renglon, 1, mov.pk)
            ws.col(0).width=3000
            ws.write(renglon, 2, via.maniobras_locales)
            ws.col(0).width=3000
            ws.write(renglon, 3, via.maniobras_foraneas)
            ws.col(0).width=3000
            ws.write(renglon, 4, via.maniobras_retrabajos)
            ws.col(0).width=3000
            ws.write(renglon, 5, mov.concepto)
            ws.col(0).width=3000
            ws.write(renglon, 6, mov.importe,formatoNum)
            ws.col(0).width=3000
            ws.write(renglon, 7, 'Sin solicitud')
            ws.col(0).width=3000
            ws.write(renglon, 8, 'Sin solicitud')
            ws.col(0).width=3000
            renglon += 1
    else:
        messages.warning(request, 'No se encontraron registros con los criterios seleccionados')
        form = reporteForm(request.POST)
        return render(request,'reporteManiobras.html', {'form': form,})


    wb.save(response)
    return response
    
@login_required
def index_movs(request):
#Obtener catálogos    
    movimientos_list = Movimiento.objects.all()             
    paginator = Paginator(movimientos_list, 10)
    page = request.GET.get('page')
    
    try:
        movi = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        movi = paginator.page(1)
    except EmptyPage:
        # If page is out of range (e.g. 9999), deliver last page of results.
        movi = paginator.page(paginator.num_pages)
    
        
    context = RequestContext(request,{             
        'movi': movi,
    })    
    
    return render_to_response('tipomov.html', context)
@login_required
def agregarMov(request):

    userid = request.user.id 
    user = request.user
    profile=user.get_profile()
    u1= Perfil.objects.get(user_id=userid)

    if request.method=='POST':
        form = movimientoForm(request.POST)
        if form.is_valid():
            #mov=form.save(commit=False)
            #mov.usuario= u1.id
            mov=form.save()
            messages.success(request, 'Se ha agregado el tipo Movimiento %s'%mov)
            return HttpResponseRedirect(reverse('viajes.views.index_movs',))
    else:
        form = movimientoForm()
    context = RequestContext(request,{
        'form':form,
        'action':'Agregar',
        })
    return render_to_response('Movimiento_add.html', context)


def editMov(request, movimiento_id):

    movimientos=get_object_or_404(Movimiento, pk=movimiento_id)

    if request.method=='POST':

        form = movimientoForm(request.POST, instance = movimientos)

        if form.is_valid():

            form.save()
            messages.success(request, 'Se ha modificado el Movimiento')
            return HttpResponseRedirect(reverse('viajes.views.index_movs',))
        else:
            messages.error(request, 'Ha ocurrido algun error, favor de revisar los datos.')
    else:
        form = movimientoForm(instance = movimientos)
    context=RequestContext(request,{
        'form':form,
        'action':'Editar',
        'action2':'Cancelar',
        })
    return render_to_response('movedit.html', context)

@login_required
def verMovimientos(request):

    movunidad_list = MovUnidad.objects.all().order_by('-fechai')

    paginator = Paginator(movunidad_list, 100)
    page = request.GET.get('page')
    
    try:
        movunidad = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        movunidad = paginator.page(1)
    except EmptyPage:
        # If page is out of range (e.g. 9999), deliver last page of results.
        movunidad = paginator.page(paginator.num_pages)
    
        
    context = RequestContext(request,{             
        'movunidad': movunidad,
    })    
    
    return render_to_response('movimientos.html', context)
@login_required
def indexIncidentes (request):
    incidente_list = ReporteUnidad.objects.all()             
    paginator = Paginator(incidente_list, 100)
    page = request.GET.get('page')
    
    try:
        incidente = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        incidente = paginator.page(1)
    except EmptyPage:
        # If page is out of range (e.g. 9999), deliver last page of results.
        incidente = paginator.page(paginator.num_pages)
    
        
    context = RequestContext(request,{             
        'incidente': incidente,
    })        
    return render_to_response('incidentes.html', context)


##### incidencias #####

@login_required
def agregaMovimiento (request):  
    userid = request.user.id 
    user = request.user
    u1= Perfil.objects.get(user_id=userid)

    if request.method=='POST':
        form = repUnidad(request.POST)
        
        if form.is_valid():
            movimiento=form.save(commit = False)
            movimiento.usuario = u1
            movimiento.status= 1
            #movimiento = form.save()
            #messages.success(request, 'Se ha guardado el Movimiento')
            #return HttpResponseRedirect(reverse('viajes.views.indexIncidentes',))
        formImagen = UploadForm(request.POST, request.FILES)
       
        try: 
            IMG1=str(request.FILES['imagen1'])
            cut_IMG1 = IMG1.split('.')
            extIMG1 = cut_IMG1[1]
            idm = ReporteUnidad
            print 'Este es el nombre de la imagen'
            movimiento.imagen1 = request.FILES['imagen1']
            print movimiento.imagen1

            if extIMG1 !='jpg':
                messages.error(request, 'Error Foto 1: el archivo (%s) no es de extension .jpg '%request.FILES['imagen1'])
            else:
                if request.FILES['imagen1'] is not None or extIMG1=='jpg':
                    #newdoc1 = ReporteUnidad(imagen1 = request.FILES['imagen1'],imagen2 ='imgIncidente/default.jpg',imagen3 ='imgIncidente/default.jpg',imagen4 = 'imgIncidente/default.jpg')
                    #newdoc1.nombre = movimiento.pk
                    #newdoc1.save(formImagen)
                    #newdoc1.delete()
                    messages.success(request, 'La Foto 1: (%s) se agrego correctamente '%request.FILES['imagen1'])    
                    
                    lista1 = os.listdir(settings.MEDIA_ROOT + 'imgEconomicos/')
                    for lis1 in lista1:
                        PARTE1 = str(lis1)
                        num1=500
                        if PARTE1 == str(request.FILES['imagen1']):
                            llave1 = str(economico.pk)
                            NUEVAPARTE1 = PARTE1.split('.')
                            newimg1=llave1+'_'+'1.'+NUEVAPARTE1[1]
                            os.rename (settings.MEDIA_ROOT + 'imgEconomicos/%s'%PARTE1, settings.MEDIA_ROOT + 'imgEconomicos/%s'%newimg1)
                        num1 +=1
        except Exception:
            print "error foto 1"
        
        try:
            IMG2=str(request.FILES['imagen2'])
            cut_IMG2 = IMG2.split('.')
            extIMG2 = cut_IMG2[1]
            movimiento.imagen2 = request.FILES['imagen2']
            if extIMG2 !='jpg':
                messages.error(request, 'Error Foto 2: el archivo (%s) no es de extension .jpg '%request.FILES['imagen2'])
            else:
                
                if request.FILES['imagen2'] is not None or extIMG2=='jpg':
                    #newdoc2 = Imagen(imagen1 ='imgEconomicos/default.jpg' ,imagen2 = request.FILES['imagen2'],imagen3 ='imgEconomicos/default.jpg',imagen4 = 'imgEconomicos/default.jpg')
                    #newdoc2.nombre = economico.pk
                    #newdoc2.save(formImagen)
                    #newdoc2.delete()
                    messages.success(request, 'La Foto 2: (%s) se agrego correctamente '%request.FILES['imagen2'])    
                    lista2 = os.listdir(settings.MEDIA_ROOT + 'imgEconomicos/')
                    for lis2 in lista2:
                        PARTE2 = str(lis2)
                        num2=1000
                        if PARTE2 == str(request.FILES['imagen2']):
                            llave2 = str(economico.pk)
                            NUEVAPARTE2 = PARTE2.split('.')
                            newimg2=llave2+'_'+'2.'+NUEVAPARTE2[1]
                            os.rename (settings.MEDIA_ROOT + 'imgEconomicos/%s'%PARTE2, settings.MEDIA_ROOT + 'imgEconomicos/%s'%newimg2)
                        num2 +=1
        except Exception:
            print "error foto 2"    
        
        try:
            IMG3=str(request.FILES['imagen3'])
            cut_IMG3 = IMG3.split('.')
            extIMG3 = cut_IMG3[1]
            movimiento.imagen3 = request.FILES['imagen3']
            if extIMG3 !='jpg':
                messages.error(request, 'Error Foto 3: el archivo (%s) no es de extension .jpg '%request.FILES['imagen3'])
            else:    
                if request.FILES['imagen3'] is not None or extIMG3=='jpj':
                    #newdoc3 = Imagen(imagen1 = 'imgEconomicos/default.jpg',imagen2 ='imgEconomicos/default.jpg',imagen3 = request.FILES['imagen3'],imagen4 = 'imgEconomicos/default.jpg')
                    #newdoc3.nombre = economico.pk
                    #newdoc3.save(formImagen)
                    #newdoc3.delete()
                    messages.success(request, 'La Foto 3: (%s) se agrego correctamente '%request.FILES['imagen3'])
                    lista3 = os.listdir(settings.MEDIA_ROOT + 'imgEconomicos/')
                    for lis3 in lista3:
                        PARTE3 = str(lis3)
                        num3=1500
                        if PARTE3 == str(request.FILES['imagen3']):
                            llave3 = str(economico.pk)
                            NUEVAPARTE3 = PARTE3.split('.')
                            newimg3=llave3+'_'+'3.'+NUEVAPARTE3[1]
                            os.rename (settings.MEDIA_ROOT + 'imgEconomicos/%s'%PARTE3, settings.MEDIA_ROOT + 'imgEconomicos/%s'%newimg3)
                            num3 +=1
        except Exception:
            print "error foto 3"

        try:
            IMG4=str(request.FILES['imagen4'])
            cut_IMG4 = IMG4.split('.')
            extIMG4 = cut_IMG4[1]
            movimiento.imagen4 = request.FILES['imagen4']
            if extIMG4 !='jpg':
                messages.error(request, 'Error Foto 4: el archivo (%s) no es de extension .jpg '%request.FILES['imagen4'])
            else:    
                if request.FILES['imagen4'] is not None or extIMG4=='jpj':
                    #newdoc4 = Imagen(imagen1 = 'imgEconomicos/default.jpg',imagen2 ='imgEconomicos/default.jpg',imagen3 ='imgEconomicos/default.jpg',imagen4 = request.FILES['imagen4'])
                    #newdoc4.nombre = economico.pk
                    #newdoc4.save(formImagen)
                    #newdoc4.delete()
                    messages.success(request, 'La Foto 4: (%s) se agrego correctamente '%request.FILES['imagen4'])
                    lista4 = os.listdir(settings.MEDIA_ROOT + 'imgEconomicos/')
                    for lis4 in lista4:
                        PARTE4 = str(lis4)
                        num4=2000
                        if PARTE4 == str(request.FILES['imagen4']):
                            llave4 = str(economico.pk)
                            NUEVAPARTE4 = PARTE4.split('.')
                            newimg4=llave4+'_'+'4.'+NUEVAPARTE4[1]
                            os.rename (settings.MEDIA_ROOT + 'imgEconomicos/%s'%PARTE4, settings.MEDIA_ROOT + 'imgEconomicos/%s'%newimg4)
                            num4 +=1
        except Exception:
            print "error foto 4"

        movimiento.status = 1
        movimiento = form.save()
        idm = str(movimiento.pk)
        eco = str(movimiento.reporta)
        obs = str(movimiento.obs)
        dano = movimiento.dano
        
        messages.success(request, 'Se ha guardado el Movimiento') 
        #### Enviar correo del reporte ####

        messages.success(request,'El reporte se ha creado de forma correcta, se enviara un correo con la informacion.')
        today = date.today()
        msg1 = 'El Incidente ' + idm + ' se ha creado favor de verificarlo  \n\n'
        msg2 = 'El motivo  es : ' + dano + ' \n '
        msg3 = '\n\n Contacto : \n\n  Departamento de Sistemas. '
        msg4 = 'El economico es: ' + eco +'  \n\n '
        msg5 = 'Las Observaciones del incidente es :' + obs +'  \n\n'
        msg  = msg1 +  msg4 + msg5 + msg2 + msg3 
        send_mail('IMPORTANTE CONTROL ADMINISTRATIVO, TE INFORMA QUE EXISTE UN REPORTE DE UNIDAD PENDIENTE POR  REVISAR Y AUTORIZAR : ' + idm , msg, 'controladministrativo@logisticavitesse.com.mx',
        ['julio@logisticavitesse.com.mx', 'victor@logisticavitesse.com.mx', 'alejandro@logisticavitesse.com.mx', 'griselda@logisticavitesse.com.mx'], fail_silently=False)
        
        return HttpResponseRedirect(reverse('viajes.views.indexIncidentes',))       
          
    else:

        form=repUnidad()
        formImagen=UploadForm()
    context=RequestContext(request,{
        'form':form,
        'formImagen':formImagen,
        'action':'Agregar', 
        })
    return render_to_response('MovUni.html', context)

def statusIncidente (request, idReporteUnidad, estatus):
    userid = request.user.id 
    user = request.user
    u1= Perfil.objects.get(user_id=userid)
    try: 
        incidente=ReporteUnidad.objects.get(pk=idReporteUnidad)
        s = int(estatus)
        ####referencia = str(incidente.obs)
        e = int(incidente.reporta_id)
        if s == 2:
            ##ReporteUnidad.objects.create(usuario = 'Oscar Farias', reporta_id = 2)
            incidente.status = estatus
            incidente.save()
            #Solicitudes.objects.create(solicitante_id = userid, fecha =incidente.fecha, status = 1, importe = 0.00, importe_asig = 0.00, refer = referencia)
            messages.success(request, 'Se actualizo el status del incidente correctamente, Favor de Realizar la Solicitud Correspondiente')
            return HttpResponseRedirect(reverse('solicitudes.views.agregar' ))
        elif s == 3:
            #movimientosUnidad.objects.create(unidad_id = incidente.reporta_id, fecha=incidente.fecha, tipo = 8, obs = referencia, operador_id = userid)
            #messages.success(request, 'Se ha creado un Movimiento a la unidad, favor de terminarla desde Control Vehicular')
            incidente.status = estatus
            incidente.save()
            messages.success(request, 'Se ha actualizado el Estatus del incidente. Debe de crear el Movimiento correspondiente')
            return HttpResponseRedirect(reverse('catalogos.views.movUnidad'))
        incidente.status = estatus
        incidente.save()
        messages.success(request, 'Se ha cambiado el status del incidente, correctamente.')
    except Exception:
        messages.error(request, 'No se logro Actulizar el Estado del Incidente, favor de reportar a Sistemas')
    return HttpResponseRedirect(reverse('viajes.views.indexIncidentes'))

@login_required
def liberarViaje(request):
    userid = request.user.id 
    user = request.user
    u1= Perfil.objects.get(user_id=userid)
    form=libviajeForm()
    return render(request,'liberaviaje.html', {'form': form})

def liberar(request):
    userid = request.user.id 
    user = request.user
    u1= Perfil.objects.get(user_id=userid)
   
    viaje = request.POST['pk']

    valida=Viaje.objects.filter(id=viaje)
    vi = None
    for v in valida:
        vi=v.fecha_ent
    if vi:
        try: 
            Viaje.objects.filter(id = viaje).update(fecha_ent = None)
            info=Viaje.objects.filter(id=viaje)
            for datos in info:
                ope=datos.operador_id
                opera=datos.operador
                uni=datos.economico_id
                econo=datos.economico
                a=empleado.objects.filter(id=ope)
                for s in a:
                    sera=s.servicios
                b=Economico.objects.filter(id=uni)
                for ecovia in b:
                    envia=ecovia.enviaje


                empleado.objects.filter(id=ope).update(servicios= sera + 1)
                Economico.objects.filter(id=uni).update(enviaje= envia + 1)
                MovUnidad.objects.filter(unidad_id=uni).update(enviaje= envia +1)
            
        except Exception:

            messages.error(request, 'No se logo Actualizar')
            return HttpResponseRedirect(reverse('viajes.views.liberarViaje'))

    else: 
        messages.error(request, 'El viaje %s no esta cerrado o no existe el viaje, favor de revisar los datos.' %viaje)
        return HttpResponseRedirect(reverse('viajes.views.liberarViaje'))

    messages.success(request,'Se libero el viaje %s, Se sumo un servicio al operador %s y se sumo viaje a la unidad %s ; al cerrar el viaje se restaran nuevamente los valores.' % (viaje, opera, econo)) 
    return HttpResponseRedirect(reverse('viajes.views.liberarViaje'))

class IncidenteView(PDFTemplateView):
    template_name = "repIncidencia.html"
    
    def get_context_data(self, pk, **kwargs):

        incidente = ReporteUnidad.objects.get(pk=pk)
        
        return super(IncidenteView, self).get_context_data(
            pagesize="A4",
            title="Incidente " + pk,
            incidente=incidente,

            **kwargs
        )

@login_required
def RepIncidentes(request):
    form = repIncidenteForm()
    conceptos = json.dumps(CONCEPTO_GASTO)
    return render(request,'reporteIncidentes.html', {'form': form,})

@login_required
def imprimeIncidentes(request):

    filters = {}  
    inicial = None  
    final = None  

    if ('fecha_ini') in request.POST and request.POST['fecha_ini'].strip():
        inicial = request.POST['fecha_ini']

    if ('fecha_fin') in request.POST and request.POST['fecha_fin'].strip():
        final = request.POST['fecha_fin']

   
    if inicial and final:
        filters['fecha__range'] = (inicial, final)
    elif inicial or final:
        messages.error(request, 'Seleccione completo rango de fechas.')
        form = reporteForm(request.POST)
        return render(request,'reporteIncidentes.html', {'form': form,})

    if not len(filters):
        messages.error(request, 'Seleccione por lo menos un filtro de búsqueda o un rango de fechas')
        form = reporteForm(request.POST)
        return render(request,'reporteIncidentes.html', {'form': form,})

    response = HttpResponse(mimetype="application/vnd.ms-excel")
    response['Content-Disposition'] = 'attachment; filename=reporte.xls'

    wb = xlwt.Workbook(encoding='utf8')
    ws = wb.add_sheet('Reporte Incidencias')
    date_format = xlwt.XFStyle()
    date_format.num_format_str = 'dd/mm/yyyy'
    formatoNum = XFStyle()
    formatoNum.num_format_str = '$#,##0.00'
    font = xlwt.Font() 
    font.name = 'Arial'
    font.height = 20 * 12  
    font.bold = True
    font.italic = True
    style = xlwt.XFStyle() 
    style = xlwt.easyxf('pattern: pattern solid, fore_colour light_blue;') 
    style.font = font 

    columnas = [
            (u"Incidencia", 70),
            (u"Usuario", 70),
            (r'Fecha', 70),
            (r'Daño',70),
            (r'Observaciones', 100),
            (u"Status Actual", 70),
            (u"Motivo", 100),
            (u"Fecha Auto/Rech",70)
    ]

    posx = 0
    for col in xrange(len(columnas)):
        ws.write(0, posx, columnas[col][0])
        posx += 1
    movimientos = None
    renglon = 1

    movimientos = ReporteUnidad.objects.filter(**filters)
    
    if movimientos:
        for mov in movimientos:
            print str(mov) + ' <--'
            movimiento = mov.id        
            usuario = mov.usuario
            fecham = str(mov.fecha)
            fechaar= None
            if mov.fecha_AR:
                fechaar = str(mov.fecha_AR)
            st=mov.get_status_display()

            ws.write(renglon, 0, mov.pk)
            ws.col(0).width=3000
            ws.write(renglon, 1, mov.usuario)
            ws.col(0).width=3000
            ws.write(renglon, 2, fecham )
            ws.col(0).width=3000
            ws.write(renglon, 3, mov.dano)
            ws.col(0).width=3000
            ws.write(renglon, 4, mov.obs)
            ws.col(0).width=3000
            ws.write(renglon, 5, st)
            ws.col(0).width=3000
            ws.write(renglon, 6, mov.motivo)
            ws.col(0).width=3000
            ws.write(renglon, 7, fechaar)
            ws.col(0).width=3000
            renglon += 1

    else:
        messages.warning(request, 'No se encontraron registros con los criterios seleccionados')
        form = reporteForm(request.POST)
        return render(request,'reporteIncidentes.html', {'form': form,})


    wb.save(response)
    return response

@login_required
def editarMov(request, incidente_id):

    incidente=get_object_or_404(ReporteUnidad, pk=incidente_id)
    current_user = request.user
    valor=incidente.status

    if request.method=='POST':
        form=editarMovForm(request.POST)
        a=request.POST['motivo']
        b=request.POST['status2']
        c=incidente_id
        b = int(b)
        incidente= str(c)
        fecha=date.today()
 
        incidentes=ReporteUnidad.objects.filter(pk=c)
        for i in incidentes:
            obs=i.obs
            eco=i.reporta
            ecostr = str(eco)

        if b == 1:
            b=4 ##Motivo Autorizado
            sta='Autorizado'
            sta2='de la Autorizacion.'
        else: 
            b=6 ## Motivo Rechazados
            sta='Rechazado'
            sta2='del Rechazo'
        if form.is_valid():
            ReporteUnidad.objects.filter(id= c).update(status= b, motivo = a, fecha_AR=fecha)
            messages.success(request,'El Incidente se ha '+ sta + ' de forma adecuada. Se enviara un correo a los responsables para seguimiento')
            today       = date.today()
            msg1 = 'El Incidente ' + incidente + ' se ha ' + sta + ' favor de verificarlo  \n\n'
            msg2 = 'El motivo '+ sta2 + ' es : ' + a + ' \n '
            msg3 = '\n\n Contacto: \n\n  Departamento de Sistemas. '
            msg4 = 'El economico es: ' + ecostr + '\n\n '
            msg5 = 'Las Observaciones del incidente es : ' + obs + ' \n\n'
            msg  = msg1 +  msg4 + msg5 + msg2 + msg3 
            send_mail('Aviso incidente : ' + incidente , msg, 'controladministrativo@logisticavitesse.com.mx',
            ['julio@logisticavitesse.com.mx', 'victor@logisticavitesse.com.mx','alejandro@logisticavitesse.com.mx', 'mara@logisticavitesse.com.mx', 'griselda@logisticavitesse.com.mx'], fail_silently=False)

            return HttpResponseRedirect(reverse('viajes.views.indexIncidentes'))
        
        else:
            render_to_response('editarIncidente.html',{'form':form})
            messages.error(request,'El incidente necesita un motivo para Rechazar o Autorizar')
    else:
        form=editarMovForm()

    context=RequestContext(request,{
        'form':form,
        'action':'Editar',
        })

    return render_to_response('editarIncidente.html', context)

@login_required
def revisarMov(request, incidente_id):
    incidente=get_object_or_404(ReporteUnidad, pk=incidente_id)
    current_user = request.user
    valor=incidente.status
    userid = request.user.id 
    u1= Perfil.objects.get(user_id=userid)

    if request.method=='POST':
        form=revisarMovForm(request.POST)
        a=request.POST['obsrevision']
        c=incidente_id
        incidente= str(c)
        fecha=date.today()
        incidentes=ReporteUnidad.objects.filter(pk=c)
        for i in incidentes:
            obs=i.obs
            eco=i.reporta
            ecostr = str(eco)

        if form.is_valid():
            ReporteUnidad.objects.filter(id= c).update(status= 7, obsrevision = a, fecha_rev=fecha, urevision = current_user)
            messages.success(request,'El Incidente se ha Revisado. Se enviara un correo a los responsables para seguimiento')
            today = date.today()
            msg1 = 'El Incidente ' + incidente + ' se ha revisado favor de verificarlo  \n\n'
            msg2 = 'Las Observaciones de la revision es : ' + a + ' \n '
            msg3 = '\n\n Contacto: \n\n  Departamento de Sistemas. '
            msg4 = 'El economico es: ' + ecostr + '\n\n '
            msg5 = 'Las Observaciones del incidente es : ' + obs + ' \n\n'
            msg  = msg1 +  msg4 + msg5 + msg2 + msg3 
            send_mail('IMPORTANTE CONTROL ADMINISTRATIVO, TE INFORMA QUE EXISTE UN REPORTE DE UNIDAD PENDIENTE POR AUTORIZAR : ' + incidente , msg, 'controladministrativo@logisticavitesse.com.mx',
            ['julio@logisticavitesse.com.mx', 'victor@logisticavitesse.com.mx','alejandro@logisticavitesse.com.mx', 'griselda@logisticavitesse.com.mx'], fail_silently=False)
            #['genseg@hotmail.com'], fail_silently=False)
            return HttpResponseRedirect(reverse('viajes.views.indexIncidentes'))

        else:
            render_to_response('revisarIncidente.html',{'form':form})
            messages.error(request,'El incidente necesita una Observacion de la Revision')
    else:
        form=revisarMovForm()

    context=RequestContext(request,{
        'form':form,
        'action':'Editar',
        })

    return render_to_response('revisarIncidente.html', context)

@login_required
def indexXLS (request):
    
    lista_xls = ExcelViaje.objects.all()             
    paginator = Paginator(lista_xls, 100)
    page = request.GET.get('page')
    
    try:
        lista_xls = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        lista_xls = paginator.page(1)
    except EmptyPage:
        # If page is out of range (e.g. 9999), deliver last page of results.
        lista_xls = paginator.page(paginator.num_pages)
    
        
    context = RequestContext(request,{             
        'lista_xls': lista_xls,
    })        
    return render_to_response('listaExcel.html', context)

def descargarXLS(request, nombreArchivo):
    ruta = settings.MEDIA_ROOT + str(nombreArchivo)
    print ruta
    wrapper = FileWrapper( open( ruta, "r" ) )
    content_type = mimetypes.guess_type( ruta )[0]

    response = HttpResponse(wrapper, content_type = content_type)
    response['Content-Length'] = os.path.getsize( ruta )
    response['Content-Disposition'] = 'attachment; filename=%s' % \
                                       smart_str( os.path.basename( ruta ) )

    return response

########
########  CARGA DE ARCHIVOS 
########
@login_required
def indexArchivos (request):
    userid = request.user.id 
    user = request.user
    u1= Perfil.objects.get(user_id=userid)

    archivos_list = Archivos.objects.all()             
    paginator = Paginator(archivos_list, 100)
    page = request.GET.get('page')
    
    try:
        archivos = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        archivos = paginator.page(1)
    except EmptyPage:
        # If page is out of range (e.g. 9999), deliver last page of results.
        archivos = paginator.page(paginator.num_pages)
    
        
    context = RequestContext(request,{             
        'archivos': archivos,
    })        
    if userid == 81 or userid == 80:
        return render_to_response('listaArchivos.html', context)
    else:        
        return render_to_response('lista_archivos.html', context)
##### incidencias #####

@login_required
def agregaArchivo(request):  
    userid = request.user.id 
    user = request.user
    u1= Perfil.objects.get(user_id=userid)

    if request.method=='POST':
        form = archivosForm(request.POST)
        formImagen = UploadFilesForm(request.POST, request.FILES)

        if form.is_valid():
            print 'El Formato es Valido'
            archivo=form.save(commit = False)
            archivo.fecha = datetime.now()
       
            try: 
                IMG1=str(request.FILES['archivo'])
                cut_IMG1 = IMG1.split('.')
                extIMG1 = cut_IMG1[1]
                nombre=cut_IMG1[0]
                print 'Nombre de la imagen:'
                print nombre
                archivo.archivo = request.FILES['archivo']
                print archivo.archivo

                if extIMG1 !='jpg' or extIMG1 != 'pdf' or extIMG1 != 'xlsx' or extIMG1 != 'docx' or extIMG1 != 'png' or extIMG1 != 'pptx' :
                    messages.error(request, 'Error Archivo : el archivo (%s) no es de extension reconocida '%request.FILES['imagen1'])
                else:
                    if request.FILES['archivo'] is not None or (extIMG1=='jpg' or extIMG1 == 'xlsx' or extIMG1 == 'docx' or extIMG1 == 'png' or extIMG1 == 'pptx'):
                        messages.success(request, 'El Archivo: (%s) se agrego correctamente '%request.FILES['archivo'])
            except Exception:
                print "error Archivo"
            if extIMG1 == 'xlsx':
                tipo = 4
            if extIMG1 == 'pdf':
                tipo = 5
            if extIMG1 == 'png':
                tipo = 2
            if extIMG1 == 'jpg':
                tipo = 1
            if extIMG1 == 'pptx':
                tipo = 3
            if extIMG1 == 'docx':
                tipo = 6    
            else: 
                tipo = 7


            archivo.tipo=tipo
            archivo.nombre = nombre
            archivo = form.save()
        
            messages.success(request, 'Se ha guardado el Archivo') 
        
            return HttpResponseRedirect(reverse('viajes.views.indexArchivos',))       
          
    else:

        form=archivosForm()
        formImagen=UploadFilesForm()
    context=RequestContext(request,{
        'form':form,
        'formImagen':formImagen,
        'action':'Agregar', 
        })
    return render_to_response('add_file.html', context)

def del_file(request, archivo_id):
    archivo=get_object_or_404(Archivos, pk=archivo_id)
    current_user = request.user
    Archivos.objects.filter(id=archivo_id).delete()

    ##print'AL MENOS LLEGA AQUI¡'

    context=RequestContext(request)
    return HttpResponseRedirect(reverse('viajes.views.indexArchivos'))

def descargarArchivo(request, nombreArchivo):
    ruta = settings.MEDIA_ROOT + str(nombreArchivo)
    print ruta
    wrapper = FileWrapper( open( ruta, "r" ) )
    content_type = mimetypes.guess_type( ruta )[0]

    response = HttpResponse(wrapper, content_type = content_type)
    response['Content-Length'] = os.path.getsize( ruta )
    response['Content-Disposition'] = 'attachment; filename=%s' % \
                                       smart_str( os.path.basename( ruta ) )

    return response

def viajesAbiertos(request):
    filters={}
    inicial = '2016-08-01'
    final = '2018-12-31'
    filters['fecha_salida__range']=(inicial, final)
    
    lista_viajes= Viaje.objects.filter(fecha_ent = None, **filters).order_by('-pk')

    print lista_viajes

    paginator = Paginator(lista_viajes, 100)
    page=request.GET.get('page')

    try:
        lista_viajes = paginator.page(page)
    except PageNotAnInteger:
        lista_viajes = paginator.page(1)
    except EmptyPage:
        lista_viajes = paginator.page(paginator.num_pages)

    context = RequestContext(request,{
        'lista_viajes':lista_viajes,
        })
    return render_to_response('lista_viajes_abiertos.html', context)

def facturasIndex(request):

    listaFacturas=Facturas.objects.all()

    print listaFacturas

    paginator = Paginator(listaFacturas, 100)
    page=request.GET.get('page')

    try:
        listaFacturas = paginator.page(page)
    except PageNotAnInteger:
        listaFacturas=paginator.page(1)
    except EmptyPage:
        listaFacturas=paginator.page(paginator.num_pages)
    context=RequestContext(request,{
        'listaFacturas':listaFacturas,
        })
    return render_to_response('listaFacturas.html', context)

#Modificacion Carga Excel
def subirXLS(request): 
    user = request.user
    profile = user.get_profile()
    participant = user.participant_set.first()
    user = request.user

    if request.method == 'POST':
        form = UploadExcelForm(request.POST, request.FILES)
        #form.save()
        if form.is_valid():
            newdoc = ExcelViaje(docfile = request.FILES['docfile'], usuario = user)
            nombreArchivo = str(newdoc.docfile)
            nombreArchivo = nombreArchivo.upper()


            if nombreArchivo.endswith('.XLSX'):
                newdoc.save()
                actFactura(request, str(newdoc.docfile))
                ##ExcelViaje.objects.all().delete() Evitar que se borre el registro.
            else: 
                print 'Archivo con formato invalido'
                txtmsg = 'Archivo con formato inválido.'
                messages.error(request, txtmsg)  
              
        return render(request,'subirXLS.html', {'form': form,})
    else:
        form = UploadExcelForm()
    return render(request,'subirXls.html', {'form': form,})


def actFactura(request, datoArchivo):
    user = request.user
    profile = user.get_profile()
    participant = user.participant_set.first()
    workbook = openpyxl.load_workbook(filename=settings.MEDIA_ROOT + datoArchivo, use_iterators=True)

    try:
        sheet = workbook.get_sheet_by_name('pagos')
        exitos = 0
        numError = 0
        total = 0    
        for row in range(3, sheet.get_highest_row()):
            total += 1
            print sheet['A' + str(row)].value            
            if sheet['A' + str(row)].value == None: break
            try:
                participant = user.participant_set.first()
                
                if Facturas.objects.filter(factura=sheet['A'+str(row)].value).exists():
                    monto = sheet['D'+str(row)].value
                    facturas=Facturas.objects.filter(factura=sheet['A'+str(row)].value)
                    for factura in facturas:
                        saldoi = factura.saldo
                        saldof = saldoi - long(monto)
                        if saldof < 0:
                            messages.warning(request, u'Error ' + str(sheet['A' + str(row)].value))
                            numError += 1
                    Facturas.objects.filter(factura=sheet['A'+str(row)].value).update(
                        fecha_pago = sheet['C'+str(row)].value,
                        pagado = sheet['D'+str(row)].value,
                        banco = sheet['E'+str(row)].value,
                        cuenta = sheet['F'+str(row)].value,
                        saldo = saldof,
                        )
                    Pagos.objects.create(factura = sheet['A'+str(row)].value,
                         fecha_factura = sheet['B'+str(row)].value,
                         fecha_pago = sheet['C'+str(row)].value,
                         monto = sheet['D'+str(row)].value,
                         banco= sheet['E'+str(row)].value,
                         cuenta = sheet['F'+str(row)].value,
                         )
                    Viaje.objects.filter(factura=sheet['A' + str(row)].value).update(
                            facturacion_fecha_pago = sheet['C' + str(row)].value,
                            )               
                    exitos += 1
                else :
                    messages.warning(request, u'No existe la factura ' + str(sheet['A' + str(row)].value))
                    numError += 1
            except Exception:
                messages.error(request, u'Error al registrar el Pago de la Factura' + str(sheet['A' + str(row)].value) + u' Verifique los datos o estatus')
                numError += 1
    
        archivoF = settings.MEDIA_ROOT + 'upload/pagofacturas.xlsx'
        #os.remove(archivoF)  ## Evitar que se elimine el Excel.
        if exitos == 0: 
            messages.error(request, 'No se realizaron actualizaciones')
        if numError == 0:
            messages.success(request, 'Registros actualizados exitosamente.')
        if exitos > 0 and numError > 0:
            messages.success(request, 'Actualización parcial realizada.')
       
    except KeyError:
        messages.error(request, 'Archivo no contiene la hoja pagos')

def index_viajes_doc(request):

    filters={}
    inicial='2016-10-01'
    final = datetime.now()
    filters['fecha_salida__range']=(inicial, final)

    searchform = SearchForm(request.GET)
    
    filters_1 = {}    
    if ('pk' in request.GET) and request.GET['pk'].strip():        
        filters_1['pk'] = request.GET['pk']
    if ('status' in request.GET) and request.GET['status'].strip():
        filters_1['pk'] = request.GET['status']
    if ('fecha_from' in request.GET) and request.GET['fecha_from'].strip():
        filters_1['fecha_salida__gte'] = request.GET['fecha_from']
    if ('fecha_to' in request.GET) and request.GET['fecha_to'].strip():
        filters_1['fecha_salida__lte'] = request.GET['fecha_to']
    if ('cliente' in request.GET) and request.GET['cliente'].strip():
        filters_1['cliente__pk'] = request.GET['cliente']
    if ('departamento' in request.GET) and request.GET['departamento'].strip():
        filters_1['departamento__pk'] = request.GET['departamento']        
    if ('referencia' in request.GET) and request.GET['referencia'].strip():
        filters_1['referencia__icontains'] = request.GET['referencia']
    if ('economico' in request.GET) and request.GET['economico'].strip():
        filters_1['economico__pk'] = request.GET['economico']
    if ('operador' in request.GET) and request.GET['operador'].strip():
        filters_1['operador__pk'] = request.GET['operador']
    if ('destino' in request.GET) and request.GET['destino'].strip():
        destinos = Destino.objects.filter(destino_clave_municipio=request.GET['destino']).values_list('viaje_id', flat=True)        
        filters_1['id__in'] = destinos        
        
    ##viajes_list = ViajeStatus.objects.filter(**filters)
    viajes_cerrados=Viaje.objects.filter(factura__isnull= True, **filters).order_by('pk')
  
    #if not bool(filters_1):
    #    print 'entra a la validacion de filtros'
        #viajes_list = viajes_list.order_by('-workflowactivity__pk')                   
    #    viajes_cerrados=viajes_cerrados.filter(**filters_1)
    viajes_cerrados=viajes_cerrados.filter(**filters_1).order_by('-pk')

    paginator = Paginator(viajes_cerrados, 200)
    page=request.GET.get('page')

    try:
        viajes_cerrados = paginator.page(page)
    except PageNotAnInteger:
        viajes_cerrados=paginator.page(1)
    except EmptyPage:
        viajes_cerrados=paginator.page(paginator.num_pages)
    context=RequestContext(request,{
        'viajes_cerrados':viajes_cerrados,
        'searchform':searchform,
        })
    return render_to_response('viajes_cerrados.html', context)

def cambia_doc(request, pk, status_doc):
    new_status= int(status_doc) + 1
    hoy = time.strftime("%c")
    x = datetime.now()
    user = request.user
    profile = user.get_profile()
    u1= str(profile)
    #print str(profile)

    if status_doc == '1':
        Viaje.objects.filter(pk=pk).update(libera_chofer = x, status_doc = new_status, u_libera_chofer = u1)
        viaje= get_object_or_404(Viaje, pk=pk)
        current = viaje.flujo.current_state()
        viaje2 = viaje.id 
        print viaje2
        noFlujo = viaje.flujo_id
        wfh = WorkflowHistory.objects.filter(workflowactivity_id = noFlujo)
        wfhf = WorkflowHistory.objects.filter(id = wfh[0].id)
        for wf in wfhf:
            wfh_id=wf.id
            wfh_st=wf.state_id
            wf.state_id = 12
            wf.save()
    elif status_doc == '2':
        Viaje.objects.filter(pk=pk).update(liberado_cliente = x, status_doc = new_status, u_liberado_cliente= u1)
    elif status_doc == '3':
        Viaje.objects.filter(pk=pk).update(entrego_cliente = x, status_doc = new_status, u_entrego_cliente = u1)
    elif status_doc == '4':
        Viaje.objects.filter(pk=pk).update(recibido_facturacion = x, status_doc = new_status, u_recibido_facturacion=u1)
    elif status_doc == '5':
        Viaje.objects.filter(pk=pk).update(sabana = x, status_doc = new_status, u_sabana= u1)
   
    return redirect(index_viajes_doc)

def impRepCerrados(request):
    
    response = HttpResponse(mimetype="application/vnd.ms-excel")
    response['Content-Disposition'] = 'attachment; filename=reporte.xls'

    wb = xlwt.Workbook(encoding='utf8')
    ws = wb.add_sheet('Reporte Viajes Cerrados')
    date_format = xlwt.XFStyle()
    date_format.num_format_str = 'dd/mm/yyyy'
    formatoNum = XFStyle()
    formatoNum.num_format_str = '$#,##0.00'
    font = xlwt.Font() 
    font.name = 'Arial'
    font.height = 20 * 12  
    font.bold = True
    font.italic = True
    style = xlwt.XFStyle() 
    style = xlwt.easyxf('pattern: pattern solid, fore_colour light_blue;') 
    style.font = font 

    columnas = [
            (u"Viaje", 70),
            (u"Fecha Salida", 70),
            (r'Cliente', 70),
            (r'Departamento',70),
            (r'Referencia', 100),
            (u"Destino", 70),
            (u"Economico", 100),
            (u"Operador",70),
            (u"Entrega Chofer",70),
            (u"Libera Cliente",70),
            (u"Cliente a Vitesse",70), 
            (u"Recibe Facturacion", 70),
            (u"Autorizacion Cliente / Facturar", 70),
    ]
    posx = 0
    for col in xrange(len(columnas)):
        ws.write(0, posx, columnas[col][0])
        posx += 1
    movimientos = None
    renglon = 1
    
    filters={}
    inicial='2016-10-01'
    final = datetime.now()
    filters['fecha_salida__range']=(inicial, final)
    viajes = Viaje.objects.filter(factura__isnull= True, **filters).order_by('-pk')
    ##viajes = Viaje.objects.filter(factura__isnull= True)
    if viajes:
        for v in viajes:
            print str(v) + ' <--'
            viaje = v.id        
            #st=mov.get_status_display()
            ws.write(renglon, 0, viaje)
            ws.col(0).width=3000
            ws.write(renglon, 1, str(v.fecha_salida))
            ws.col(0).width=3000
            ws.write(renglon, 2, str(v.cliente))
            ws.col(0).width=3000
            ws.write(renglon, 3, str(v.departamento))
            ws.col(0).width=3000
            ws.write(renglon, 4, v.referencia)
            ws.col(0).width=3000
            ws.write(renglon, 5, 'Destino')
            ws.col(0).width=3000
            ws.write(renglon, 6, str(v.economico))
            ws.col(0).width=3000
            ws.write(renglon, 7, str(v.operador))
            ws.col(0).width=3000
            ws.write(renglon, 8, str(v.libera_chofer))
            ws.col(0).width=3000
            ws.write(renglon, 9, str(v.liberado_cliente))
            ws.col(0).width=3000
            ws.write(renglon, 10, str(v.entrego_cliente))
            ws.col(0).width=3000
            ws.write(renglon, 11, str(v.recibido_facturacion))
            ws.col(0).width=3000
            ws.write(renglon, 12, str(v.sabana))
            ws.col(0).width=3000
            
            renglon += 1

    else:
        messages.warning(request, 'No se encontraron registros con los criterios seleccionados')
        return redirect(index_viajes_doc)
    wb.save(response)

    return response


@login_required
def agregarSolicitudes(request):


    SolicitudesFormSet = modelformset_factory(SolicitudesViaje, form=SolicitudesViajeForm, max_num=30, extra=1, can_delete=True)

    if request.method=='POST':
        print 'tratando de guardar'
        formset_solicitudes = SolicitudesFormSet(request.POST, prefix='solicitudes')
        print formset_solicitudes
        if formset_solicitudes.is_valid():
            print 'pasa el segundo if 2do'

            formset_solicitudes.save()

    else:
        estatus_inicial=[{'estatus': 1,},{'estatus': 1,},]
        formset_solicitudes = SolicitudesFormSet(queryset=SolicitudesViaje.objects.none(),initial=estatus_inicial, prefix='solicitudes')
        form = SolicitudesViajeForm()
    context = RequestContext(request,{
        'formset_solicitudes':formset_solicitudes,
        'action':'Agregar',
        })
    return render_to_response('solicitudes_viaje.html', context)



@login_required
def prefacturasCargadas(request):
    if request.method == 'POST':
        print 'post'
    else:
        print 'get request..'
        form = prefacturasForm()
        fechasCarga = Prefacturas.objects.values('fechaCarga').distinct()
        print 'recorre'
        for fecha in fechasCarga:
            print fecha['fechaCarga']

    context = RequestContext(request,{
        'form':form,
        'fechasCarga':fechasCarga

    })

    return render_to_response('prefacturas.html', context)

@login_required
def revisionPrefactura(request, fecha):
    prefacturas = Prefacturas.objects.filter(fechaCarga = fecha)
    if request.method == 'POST':
        print 'post'
    else:
        print 'nothing'

    context = RequestContext(request,{
        'prefacturas':prefacturas,
        'fecha':fecha,

    })

    return render_to_response('resumeprefacturas.html', context)

@login_required
def descargarMachote(request, nombreArchivo):
    ruta = settings.MEDIA_ROOT  + 'prefactura/' + 'prefactura_' + str(nombreArchivo) + '.xls'
    wrapper = FileWrapper( open( ruta, "r" ) )
    content_type = mimetypes.guess_type( ruta )[0]
    response = HttpResponse(wrapper, content_type = content_type)
    response['Content-Length'] = os.path.getsize( ruta )
    response['Content-Disposition'] = 'attachment; filename=%s' % \
                                       smart_str( os.path.basename( ruta ) )

    return response

@login_required
def subirPrefacturas(request):
    if request.method == 'POST':
        print 'post request'
        #response = HttpResponse(mimetype="application/vnd.ms-excel")
        #response['Content-Disposition'] = 'attachment; filename=reporte.xls'

        form = prefacturasFileForm(request.POST, request.FILES)
        #form.save()
        if form.is_valid():
            newdoc = ArchivoPrefacturas(archivo = request.FILES['archivo'])
            nombreArchivo = str(newdoc.archivo)
            nombreArchivo = nombreArchivo.upper()
            print nombreArchivo

            if nombreArchivo.endswith('.XLSX'):
                try:
                    newdoc.save()
                except Exception:
                    print 'no se pudo guardar...'
                workbook = openpyxl.load_workbook(filename=settings.MEDIA_ROOT +  str(newdoc.archivo), use_iterators=True)
                #print workbook.get_sheet_names()[0]
                sheets = workbook.get_sheet_names()
                sheet = workbook.get_sheet_by_name(sheets[0])
                print str(sheet)
                ABC = str(sheet)
                print ABC
                c = workbook.get_sheet_names()[0]
                
                ahora = datetime.now()
                print ahora
                ahora = ahora.strftime("%Y%m%d%H%M%S")
                print ahora

                wb = xlwt.Workbook(encoding='utf8')
                ws = wb.add_sheet('Prefactura')
                date_format = xlwt.XFStyle()
                date_format.num_format_str = 'dd/mm/yyyy'
                formatoNum = XFStyle()
                formatoNum.num_format_str = '$#,##0.00'
                font = xlwt.Font()
                font.name = 'Arial'
                font.height = 20 * 12
                font.bold = True
                font.italic = True
                style = xlwt.XFStyle() # Crar Style
                style = xlwt.easyxf('pattern: pattern solid, fore_colour light_blue;')
                style.font = font

                columnas = [
                        (u"Clave", 70),
                        (r'Cliente', 70),
                        (r'Fecha de elaboración', 70),
                        (u"Descuento financiero", 70),
                        (u"Observaciones", 70),
                        (r'Moneda', 50),
                        (r'Tipo de Cambio', 50),
                        (u"Clave de vendedor",70),
                        (r'Su pedido', 70),
                        (u"Fecha de entrega", 70),
                        (u"Fecha de vencimiento", 70),
                        (u"Precio", 70),
                        (r'Desc. 1', 70),
                        (r'Desc. 2', 70),
                        (r'Desc. 3', 70),
                        (r'Comisión', 70),
                        (u"Clave de esquema de impuestos", 70),
                        (r'Clave de artículo', 70),
                        (u"Cantidad", 70),
                        (u"I.E.P.S.", 70),
                        (u"Impuesto 2", 70),
                        (u"Impuesto 3", 70),
                        (u"I.V.A.", 70),
                        (u"Observaciones de partida", 70),
                ]

                posx = 0
                for col in xrange(len(columnas)):
                    ws.write(0, posx, columnas[col][0])
                    posx += 1
                exitos = 0
                numError = 0
                total = 0
                renglon=1

                ######## Nuevo codigo para identificar Formato
                print 'inicia nuevo codigo.'
                if 'lg' in ABC.lower():
                    print 'Debe entrar al formato de LG'
                    for row in range(2, sheet.get_highest_row()):
                        print sheet['A' + str(row)].value
                        cv = sheet['A' + str(row)].value
                        ##### Aqui metemos la diferencia del formato:
                        test = str(sheet.get_highest_row())
                        print test
                        print ' Valor de Test'
                        if cv:
                            try:
                                if cv:
                                    viaje = Viaje.objects.get(referencia=cv)##.exclude(factura = 'Cancelado')
                                    print viaje
                                    lineas = {}
                                    fletes = sheet['F' + str(row)].value
                                    print sheet['F' + str(2)].value
                                    print 'fletes ' + str(fletes)
                                    if float(fletes) > 0:
                                        print 'flete mayor'
                                        lineas[sheet['F' + str(2)].value] = sheet['F' + str(row)].value

                                    maniobras = sheet['K' + str(row)].value
                                    print 'maniobras ' + str(maniobras)
                                    if float(maniobras) > 0:
                                        print 'maniobras mayor'
                                        lineas[sheet['K' + str(2)].value] = sheet['K' + str(row)].value

                                    #estadias = sheet['F' + str(row)].value
                                    #print 'estadias ' + str(estadias)
                                    #if float(estadias) > 0:
                                    #    print 'estadias mayor'
                                    #    lineas[sheet['F' + str(3)].value] = sheet['F' + str(row)].value

                                    casetas = sheet['M' + str(row)].value
                                    print 'estadias ' + str(casetas)
                                    if float(casetas) > 0:
                                        print 'estadias mayor'
                                        lineas[sheet['M' + str(2)].value] = sheet['M' + str(row)].value


                                    otros = sheet['L' + str(row)].value
                                    print 'otros ' + str(otros)
                                    if float(otros) > 0.00:
                                        print 'otros mayor'
                                        lineas[sheet['L' + str(2)].value] = sheet['L' + str(row)].value
                                    print 'tamaño'
                                    #print len(lineas)
                                    #Guardamos el model
                                    prefactura = Prefacturas()
                                    prefactura.viaje = str(viaje.pk)
                                    prefactura.cliente= viaje.cliente
                                    #prefactura.tc = sheet['A' + str(row)].value
                                    #prefactura.cuenta = sheet['C' + str(row)].value
                                    prefactura.descripcion = sheet['S' + str(row)].value
                                    prefactura.cv = sheet['A' + str(row)].value
                                    prefactura.fte = sheet['D' + str(row)].value
                                    prefactura.flete = sheet['F' + str(row)].value
                                    prefactura.maniobras = sheet['K' + str(row)].value
                                    #prefactura.estadias = sheet['L' + str(row)].value
                                    prefactura.casetas = sheet['M' + str(row)].value
                                    #prefactura.cve = sheet['J' + str(row)].value
                                    prefactura.otros = sheet['L' + str(row)].value
                                    #prefactura.cargosImport = sheet['I' + str(row)].value
                                    prefactura.subtotalImporte = sheet['E' + str(row)].value
                                    prefactura.porcentajeIva = '16 %'
                                    #prefactura.importeIva = sheet['K' + str(row)].value
                                    prefactura.porcentaje = '4 %'
                                    #prefactura.retencionImporte = sheet['M' + str(row)].value
                                    #prefactura.guiaImporte = sheet['S' + str(row)].value
                                    #prefactura.lineaTransporte = sheet['B3'].value
                                    #prefactura.guia = sheet['T3'].value
                                    #prefactura.origen = sheet['O' + str(row)].value
                                    prefactura.nombreOrigen = sheet['O' + str(row)].value
                                    #prefactura.fecha = sheet['M4'].value
                                    #prefactura.parentInv = sheet['M5'].value
                                    prefactura.fechaCarga = ahora
                                    prefactura.save()

                                    total += 1

                                    mm = viaje.cliente.pk
                                    print  mm

                                    #cli = Departamento.objects.get(departamento = c.upper())
                                    #clinom = cli.claveSAE
                                    #print clinom

                                    for key, value in lineas.iteritems():
                                        ws.write(renglon, 0, 1)
                                        ws.write(renglon, 1, '1')
                                        ws.write(renglon, 2, date.today(), date_format)
                                        ws.write(renglon, 3, 0)
                                        ws.write(renglon, 4, 'Viaje ' + str(viaje.pk) + ' '+ str(viaje.fecha_salida))
                                        ws.write(renglon, 5, 1)
                                        ws.write(renglon, 6, 1)
                                        ws.write(renglon, 7, 0)
                                        ws.write(renglon, 8, cv)
                                        ws.write(renglon, 9, date.today(), date_format)
                                        ws.write(renglon, 10, datetime.now() + timedelta(days=7), date_format)
                                        ws.write(renglon, 11, value)
                                        ws.write(renglon, 12, 0)
                                        ws.write(renglon, 13, 0)
                                        ws.write(renglon, 14, 0)
                                        ws.write(renglon, 15, 0)
                                        ws.write(renglon, 16, 0)
                                        if 'total_freight' in key.lower():
                                            ws.write(renglon, 17, 'F001')
                                        elif 'extra_ship_cost' in key.lower():
                                            ws.write(renglon, 17, 'OTROS 010')
                                        elif 'highway_fee' in key.lower() :
                                            ws.write(renglon, 17, 'CAS 006')
                                        elif 'estadias' in key.lower() :
                                            ws.write(renglon, 17, 'ESTAD 007')
                                        elif 'labor' in key.lower():
                                            ws.write(renglon, 17, 'MANIO 009')
                                        ws.write(renglon, 18, 1)
                                        ws.write(renglon, 19, 0)
                                        ws.write(renglon, 20, 0)
                                        if 'total_freight' in key.lower() :
                                            ws.write(renglon, 21, -4)
                                        else:
                                            ws.write(renglon, 21, 0)
                                        ws.write(renglon, 22, 16)
                                        ws.write(renglon, 23, 'Viaje: ' + str(viaje.pk) + ', '+ str(viaje.fecha_salida) + ', ' + str(cv))
                                        renglon += 1
                                        print 'exito'
                            except ObjectDoesNotExist:
                                print 'no existe viaje'
                    print total
                    if total > 0:
                        wb.save(settings.MEDIA_ROOT  + 'prefactura/' +'prefactura_' + ahora + '.xls')
                        messages.info(request, 'Archivo cargado corretamente')
                        return HttpResponseRedirect (reverse('viajes.views.prefacturasCargadas'))
                    else:
                        messages.warning(request, 'Verifique la informacion')
                    #wb.save(response)
                    #return response


                elif 'american' in ABC.lower():
                    print 'Entra al formato de American'
                    for row in range(7, sheet.get_highest_row()):
                        print sheet['B' + str(row)].value
                        cv = sheet['B' + str(row)].value
                        ##### Aqui metemos la diferencia del formato:
                        test = str(sheet.get_highest_row())
                        print test
                        print ' Valor de Test'
                        if cv:
                            try:
                                if cv:
                                    cv = cv[2:]
                                    viaje = Viaje.objects.get(referencia=cv)##.exclude(factura = 'Cancelado')
                                    print viaje
                                    lineas = {}
                                    fletes = sheet['D' + str(row)].value
                                    print sheet['D' + str(7)].value
                                    print 'fletes ' + str(fletes)
                                    if float(fletes) > 0:
                                        print 'flete mayor'
                                        lineas[sheet['D' + str(7)].value] = sheet['D' + str(row)].value

                                    maniobras = sheet['E' + str(row)].value
                                    print 'maniobras ' + str(maniobras)
                                    if float(maniobras) > 0:
                                        print 'maniobras mayor'
                                        lineas[sheet['E' + str(7)].value] = sheet['E' + str(row)].value

                                    estadias = sheet['F' + str(row)].value
                                    print 'estadias ' + str(estadias)
                                    if float(estadias) > 0:
                                        print 'estadias mayor'
                                        lineas[sheet['F' + str(7)].value] = sheet['F' + str(row)].value

                                    casetas = sheet['G' + str(row)].value
                                    print 'estadias ' + str(casetas)
                                    if float(casetas) > 0:
                                        print 'estadias mayor'
                                        lineas[sheet['G' + str(7)].value] = sheet['G' + str(row)].value


                                    otros = sheet['H' + str(row)].value
                                    print 'otros ' + str(otros)
                                    if float(otros) > 0.00:
                                        print 'otros mayor'
                                        lineas[sheet['H' + str(7)].value] = sheet['H' + str(row)].value
                                    print 'tamaño'
                                    #print len(lineas)
                                    #Guardamos el model
                                    prefactura = Prefacturas()
                                    prefactura.viaje = str(viaje.pk)
                                    prefactura.cliente= viaje.cliente
                                    prefactura.tc = sheet['A' + str(row)].value
                                    #prefactura.cuenta = sheet['C' + str(row)].value
                                    prefactura.descripcion = sheet['B3'].value
                                    prefactura.cv = sheet['B' + str(row)].value
                                    prefactura.fte = sheet['C' + str(row)].value
                                    prefactura.flete = sheet['D' + str(row)].value
                                    prefactura.maniobras = sheet['E' + str(row)].value
                                    prefactura.estadias = sheet['F' + str(row)].value
                                    prefactura.casetas = sheet['G' + str(row)].value
                                    #prefactura.cve = sheet['J' + str(row)].value
                                    prefactura.otros = sheet['H' + str(row)].value
                                    prefactura.cargosImport = sheet['I' + str(row)].value
                                    prefactura.subtotalImporte = sheet['J' + str(row)].value
                                    prefactura.porcentajeIva = sheet['L' + str(row)].value
                                    prefactura.importeIva = sheet['K' + str(row)].value
                                    prefactura.porcentaje = sheet['N' + str(row)].value
                                    prefactura.retencionImporte = sheet['M' + str(row)].value
                                    #prefactura.guiaImporte = sheet['S' + str(row)].value
                                    #prefactura.lineaTransporte = sheet['B3'].value
                                    #prefactura.guia = sheet['T3'].value
                                    #prefactura.origen = sheet['B4'].value
                                    #prefactura.nombreOrigen = sheet['F4'].value
                                    #prefactura.fecha = sheet['M4'].value
                                    #prefactura.parentInv = sheet['M5'].value
                                    prefactura.fechaCarga = ahora
                                    prefactura.save()

                                    total += 1

                                    mm = viaje.cliente.pk
                                    print  mm

                                    #cli = Departamento.objects.get(departamento = c.upper())
                                    #clinom = cli.claveSAE
                                    #print clinom

                                    for key, value in lineas.iteritems():
                                        ws.write(renglon, 0, 1)
                                        ws.write(renglon, 1, '2')
                                        ws.write(renglon, 2, date.today(), date_format)
                                        ws.write(renglon, 3, 0)
                                        ws.write(renglon, 4, 'Viaje ' + str(viaje.pk) + ' '+ str(viaje.fecha_salida))
                                        ws.write(renglon, 5, 1)
                                        ws.write(renglon, 6, 1)
                                        ws.write(renglon, 7, 0)
                                        ws.write(renglon, 8, cv)
                                        ws.write(renglon, 9, date.today(), date_format)
                                        ws.write(renglon, 10, datetime.now() + timedelta(days=7), date_format)
                                        ws.write(renglon, 11, value)
                                        ws.write(renglon, 12, 0)
                                        ws.write(renglon, 13, 0)
                                        ws.write(renglon, 14, 0)
                                        ws.write(renglon, 15, 0)
                                        ws.write(renglon, 16, 0)
                                        if 'flete' in key.lower():
                                            ws.write(renglon, 17, 'F001')
                                        elif 'otros' in key.lower():
                                            ws.write(renglon, 17, 'OTROS 010')
                                        elif 'casetas' in key.lower() :
                                            ws.write(renglon, 17, 'CAS 006')
                                        elif 'estadias' in key.lower() :
                                            ws.write(renglon, 17, 'ESTAD 007')
                                        elif 'maniobras' in key.lower():
                                            ws.write(renglon, 17, 'MANIO 009')
                                        ws.write(renglon, 18, 1)
                                        ws.write(renglon, 19, 0)
                                        ws.write(renglon, 20, 0)
                                        if 'flete' in key.lower() :
                                            ws.write(renglon, 21, -4)
                                        else:
                                            ws.write(renglon, 21, 0)
                                        ws.write(renglon, 22, 16)
                                        ws.write(renglon, 23, 'Viaje: ' + str(viaje.pk) + ', '+ str(viaje.fecha_salida) + ', ' + str(cv))
                                        renglon += 1
                                        print 'exito'
                            except ObjectDoesNotExist:
                                print 'no existe viaje'
                    print total
                    if total > 0:
                        wb.save(settings.MEDIA_ROOT  + 'prefactura/' +'prefactura_' + ahora + '.xls')
                        messages.info(request, 'Archivo cargado corretamente')
                        return HttpResponseRedirect (reverse('viajes.views.prefacturasCargadas'))
                    else:
                        messages.warning(request, 'Verifique la informacion')
                    #wb.save(response)
                    #return response

                elif 'nike' in ABC.lower():
                    print 'Entra en el formato de Nike'
                    for row in range(8, sheet.get_highest_row()):
                        print sheet['G' + str(row)].value
                        cv = sheet['G' + str(row)].value
                        ##### Aqui metemos la diferencia del formato:
                        test = str(sheet.get_highest_row())
                        print test
                        print ' Valor de Test'
                        if cv:
                            try:
                                if cv:
                                    viaje = Viaje.objects.get(referencia=cv)##.exclude(factura = 'Cancelado')
                                    print viaje
                                    lineas = {}
                                    fletes = sheet['D' + str(row)].value
                                    print sheet['D' + str(8)].value
                                    print 'fletes ' + str(fletes)
                                    if float(fletes) > 0:
                                        print 'flete mayor'
                                        lineas[sheet['D' + str(8)].value] = sheet['D' + str(row)].value

                                    ##maniobras = sheet['H' + str(row)].value
                                    ##print 'maniobras ' + str(maniobras)
                                    ##if float(maniobras) > 0:
                                    ##    print 'maniobras mayor'
                                    ##    lineas[sheet['H' + str(8)].value] = sheet['H' + str(row)].value

                                    ##estadias = sheet['I' + str(row)].value
                                    ##print 'estadias ' + str(estadias)
                                    ##if float(estadias) > 0:
                                    ##    print 'estadias mayor'
                                    ##    lineas[sheet['I' + str(8)].value] = sheet['I' + str(row)].value

                                    casetas = sheet['C' + str(row)].value
                                    print 'estadias ' + str(casetas)
                                    if float(casetas) > 0:
                                        print 'estadias mayor'
                                        lineas[sheet['C' + str(8)].value] = sheet['C' + str(row)].value


                                    #otros = sheet['K' + str(row)].value
                                    #print 'otros ' + str(otros)
                                    #if float(otros) > 0.00:
                                    #    print 'otros mayor'
                                    #    lineas[sheet['K' + str(8)].value] = sheet['K' + str(row)].value
                                    #print 'tamaño'
                                    #print len(lineas)
                                    #Guardamos el model
                                    prefactura = Prefacturas()
                                    prefactura.viaje = str(viaje.pk)
                                    prefactura.cliente= viaje.cliente
                                    #prefactura.tc = sheet['B' + str(row)].value
                                    #prefactura.cuenta = sheet['C' + str(row)].value
                                    prefactura.descripcion = sheet['B' + str(row)].value
                                    prefactura.cv = sheet['G' + str(row)].value
                                    prefactura.fte = sheet['F' + str(row)].value
                                    prefactura.flete = sheet['D' + str(row)].value
                                    #prefactura.maniobras = sheet['H' + str(row)].value
                                    #prefactura.estadias = sheet['I' + str(row)].value
                                    #prefactura.cve = sheet['J' + str(row)].value
                                    prefactura.otros = sheet['C' + str(row)].value
                                    #prefactura.cargosImport = sheet['L' + str(row)].value
                                    #prefactura.subtotalImporte = sheet['N' + str(row)].value
                                    #prefactura.porcentajeIva = sheet['O' + str(row)].value
                                    #prefactura.importeIva = sheet['P' + str(row)].value
                                    #prefactura.porcentaje = sheet['Q' + str(row)].value
                                    #prefactura.retencionImporte = sheet['R' + str(row)].value
                                    #prefactura.guiaImporte = sheet['S' + str(row)].value
                                    #prefactura.lineaTransporte = sheet['B3'].value
                                    #prefactura.guia = sheet['T3'].value
                                    #prefactura.origen = sheet['B4'].value
                                    #prefactura.nombreOrigen = sheet['F4'].value
                                    #prefactura.fecha = sheet['M4'].value
                                    #prefactura.parentInv = sheet['M5'].value
                                    prefactura.fechaCarga = ahora
                                    prefactura.save()

                                    total += 1

                                    mm = viaje.cliente.pk
                                    print  mm

                                    #cli = Departamento.objects.get(departamento = c.upper())
                                    #clinom = cli.claveSAE
                                    #print clinom

                                    for key, value in lineas.iteritems():
                                        ws.write(renglon, 0, 1)
                                        ws.write(renglon, 1, '2')
                                        ws.write(renglon, 2, date.today(), date_format)
                                        ws.write(renglon, 3, 0)
                                        ws.write(renglon, 4, 'Viaje ' + str(viaje.pk) + ' '+ str(viaje.fecha_salida))
                                        ws.write(renglon, 5, 1)
                                        ws.write(renglon, 6, 1)
                                        ws.write(renglon, 7, 0)
                                        ws.write(renglon, 8, cv)
                                        ws.write(renglon, 9, date.today(), date_format)
                                        ws.write(renglon, 10, datetime.now() + timedelta(days=7), date_format)
                                        ws.write(renglon, 11, value)
                                        ws.write(renglon, 12, 0)
                                        ws.write(renglon, 13, 0)
                                        ws.write(renglon, 14, 0)
                                        ws.write(renglon, 15, 0)
                                        ws.write(renglon, 16, 0)
                                        if 'flete' in key.lower():
                                            ws.write(renglon, 17, 'F001')
                                        elif 'otros' in key.lower():
                                            ws.write(renglon, 17, 'OTROS 010')
                                        elif 'casetas' in key.lower() :
                                            ws.write(renglon, 17, 'CAS 006')
                                        elif 'estadias' in key.lower() :
                                            ws.write(renglon, 17, 'ESTAD 007')
                                        elif 'maniobra' in key.lower():
                                            ws.write(renglon, 17, 'MANIO 009')
                                        ws.write(renglon, 18, 1)
                                        ws.write(renglon, 19, 0)
                                        ws.write(renglon, 20, 0)
                                        if 'flete' in key.lower() :
                                            ws.write(renglon, 21, -4)
                                        else:
                                            ws.write(renglon, 21, 0)
                                        ws.write(renglon, 22, 16)
                                        ws.write(renglon, 23, 'Viaje: ' + str(viaje.pk) + ', '+ str(viaje.fecha_salida) + ', ' + str(cv))
                                        renglon += 1
                                        print 'exito'
                            except ObjectDoesNotExist:
                                print 'no existe viaje'
                    print total
                    if total > 0:
                        wb.save(settings.MEDIA_ROOT  + 'prefactura/' +'prefactura_' + ahora + '.xls')
                        messages.info(request, 'Archivo cargado corretamente')
                        return HttpResponseRedirect (reverse('viajes.views.prefacturasCargadas'))
                    else:
                        messages.warning(request, 'Verifique la informacion')
                    #wb.save(response)
                    #return response
                elif 'lenovo' in ABC.lower():
                    print 'Entra en el formato de Lenovo... Formato principal'
                    for row in range(9, sheet.get_highest_row()):
                        print sheet['E' + str(row)].value
                        cv = sheet['E' + str(row)].value

                        ##### Aqui metemos la diferencia del formato:
                        if cv:
                            try:
                                if cv:
                                    viaje = Viaje.objects.get(referencia=cv)#.exclude(factura='Cancelado')
                                    print viaje
                                    lineas = {}
                                    fletes = sheet['G' + str(row)].value
                                    fletes = str(fletes).replace(',','a')
                                    fletes = str(fletes).replace('.','')
                                    fletes = str(fletes).replace('a','.')
                                    print 'fletes ' + str(fletes)
                                    if float(fletes) > 0:
                                        print 'flete mayor'
                                        #lineas[sheet['G' + str(7)].value] = sheet['G' + str(row)].value
                                        lineas[sheet['G' + str(7)].value] = float(fletes)
                                    maniobras = sheet['H' + str(row)].value
                                    maniobras = maniobras.replace(',','a')
                                    maniobras = maniobras.replace('.','')
                                    maniobras = maniobras.replace('a','.')
                                    print 'maniobras ' + str(maniobras)
                                    if float(maniobras) > 0:
                                        print 'maniobras mayor'
                                        #lineas[sheet['H' + str(7)].value] = sheet['H' + str(row)].value
                                        lineas[sheet['H' + str(7)].value] = float(maniobras)
                                    estadias = sheet['I' + str(row)].value
                                    estadias = estadias.replace(',','a')
                                    estadias = estadias.replace('.','')
                                    estadias = estadias.replace('a','.')
                                    print 'estadias ' + str(estadias)
                                    if float(estadias) > 0:
                                        print 'estadias mayor'
                                        #lineas[sheet['I' + str(7)].value] = sheet['I' + str(row)].value
                                        lineas[sheet['I' + str(7)].value] = float(estadias)

                                    otros = sheet['K' + str(row)].value
                                    otros = otros.replace(',','a')
                                    otros = otros.replace('.','')
                                    otros = otros.replace('a','.')
                                    print 'otros ' + str(otros)
                                    if float(otros) > 0:
                                        print 'otros mayor'
                                        #lineas[sheet['K' + str(7)].value] = sheet['K' + str(row)].value
                                        lineas[sheet['K' + str(7)].value] = float(otros)
                                    print 'tamaño'
                                    print len(lineas)
                                    #Guardamos el model
                                    prefactura = Prefacturas()
                                    prefactura.viaje = str(viaje.pk)
                                    prefactura.cliente= viaje.cliente
                                    prefactura.tc = sheet['B' + str(row)].value
                                    prefactura.cuenta = sheet['C' + str(row)].value
                                    prefactura.descripcion = sheet['D' + str(row)].value
                                    prefactura.cv = sheet['E' + str(row)].value
                                    prefactura.fte = sheet['F' + str(row)].value
                                    prefactura.flete = fletes  ##sheet['G' + str(row)].value
                                    prefactura.maniobras = maniobras ##sheet['H' + str(row)].value
                                    prefactura.estadias = estadias ## sheet['I' + str(row)].value
                                    prefactura.cve = sheet['J' + str(row)].value
                                    prefactura.otros = otros ##sheet['K' + str(row)].value
                                    prefactura.cargosImport = sheet['L' + str(row)].value
                                    prefactura.subtotalImporte = sheet['N' + str(row)].value
                                    prefactura.porcentajeIva = sheet['O' + str(row)].value
                                    prefactura.importeIva = sheet['P' + str(row)].value
                                    prefactura.porcentaje = sheet['Q' + str(row)].value
                                    prefactura.retencionImporte = sheet['R' + str(row)].value
                                    prefactura.guiaImporte = sheet['S' + str(row)].value
                                    prefactura.lineaTransporte = sheet['B3'].value
                                    prefactura.guia = sheet['T3'].value
                                    prefactura.origen = sheet['B4'].value
                                    prefactura.nombreOrigen = sheet['F4'].value
                                    prefactura.fecha = sheet['M4'].value
                                    prefactura.parentInv = sheet['M5'].value
                                    prefactura.fechaCarga = ahora
                                    prefactura.save()

                                    total += 1

                                    mm = viaje.cliente.pk
                                    print  mm

                                    for key, value in lineas.iteritems():
                                        ws.write(renglon, 0, 1)
                                        ws.write(renglon, 1, '2')
                                        ws.write(renglon, 2, date.today(), date_format)
                                        ws.write(renglon, 3, 0)
                                        ws.write(renglon, 4, 'Viaje ' + str(viaje.pk) + ' '+ str(viaje.fecha_salida))
                                        ws.write(renglon, 5, 1)
                                        ws.write(renglon, 6, 1)
                                        ws.write(renglon, 7, 0)
                                        ws.write(renglon, 8, cv)
                                        ws.write(renglon, 9, date.today(), date_format)
                                        ws.write(renglon, 10, datetime.now() + timedelta(days=7), date_format)
                                        ws.write(renglon, 11, value)
                                        ws.write(renglon, 12, 0)
                                        ws.write(renglon, 13, 0)
                                        ws.write(renglon, 14, 0)
                                        ws.write(renglon, 15, 0)
                                        ws.write(renglon, 16, 0)
                                        if key == 'Flete':
                                            ws.write(renglon, 17, 'F001')
                                        elif key == 'Otros':
                                            ws.write(renglon, 17, 'OTROS 010')
                                        elif key == 'Casetas':
                                            ws.write(renglon, 17, 'CAS 006')
                                        elif key == 'Estadias':
                                            ws.write(renglon, 17, 'ESTAD 007')
                                        elif key == 'Maniobra':
                                            ws.write(renglon, 17, 'MANIO 009')
                                        ws.write(renglon, 18, 1)
                                        ws.write(renglon, 19, 0)
                                        ws.write(renglon, 20, 0)
                                        if key == 'Flete':
                                            ws.write(renglon, 21, -4)
                                        else:
                                            ws.write(renglon, 21, 0)
                                        ws.write(renglon, 22, 16)
                                        ws.write(renglon, 23, 'Viaje: ' + str(viaje.pk) + ', '+ str(viaje.fecha_salida) + ', ' + str(cv))
                                        renglon += 1
                                        print 'exito'
                            except ObjectDoesNotExist:
                                print 'no existe viaje'
                    print total
                    if total > 0:
                        wb.save(settings.MEDIA_ROOT  + 'prefactura/' +'prefactura_' + ahora + '.xls')
                        messages.info(request, 'Archivo cargado corretamente')
                        return HttpResponseRedirect (reverse('viajes.views.prefacturasCargadas'))
                    else:
                        messages.warning(request, 'Verifique la informacion')
                    #wb.save(response)
                
                #print 'Archivo con formato invalido'
                txtmsg = 'El nombre de la hoja no coincide con ningun formato preestablecido, favor de revisar el nombre de la hoja.'

                messages.error(request, txtmsg)    #return response

            else:
                
                txtmsg = 'Archivo con formato inválido.'
                messages.error(request, txtmsg)

        return render(request,'subirPrefacturas.html', {'form': form,})

    else:
        print 'get request..'
        form = prefacturasFileForm()

    context = RequestContext(request,{
        'form':form
    })

    return render_to_response('subirPrefacturas.html', context)

def cambia_doc_hold_screen(request):

    param = request.GET.get('param')
    valores = param.split('|')
    pk = valores[0]
    status_doc = valores[1]

    new_status= int(status_doc) + 1
    hoy = time.strftime("%c")
    x = datetime.now()
    user = request.user
    profile = user.get_profile()
    u1= str(profile)
    #print str(profile)

    if status_doc == '1':
        Viaje.objects.filter(pk=pk).update(libera_chofer = x, status_doc = new_status, u_libera_chofer = u1)
        viaje= get_object_or_404(Viaje, pk=pk)
        current = viaje.flujo.current_state()
        viaje2 = viaje.id 
        print viaje2
        noFlujo = viaje.flujo_id
        wfh = WorkflowHistory.objects.filter(workflowactivity_id = noFlujo)
        wfhf = WorkflowHistory.objects.filter(id = wfh[0].id)
        for wf in wfhf:
            wfh_id=wf.id
            wfh_st=wf.state_id
            wf.state_id = 12
            wf.save()        
    elif status_doc == '2':
        Viaje.objects.filter(pk=pk).update(liberado_cliente = x, status_doc = new_status, u_liberado_cliente= u1)
    elif status_doc == '3':
        Viaje.objects.filter(pk=pk).update(entrego_cliente = x, status_doc = new_status, u_entrego_cliente = u1)
    elif status_doc == '4':
        Viaje.objects.filter(pk=pk).update(recibido_facturacion = x, status_doc = new_status, u_recibido_facturacion=u1)
    elif status_doc == '5':
        Viaje.objects.filter(pk=pk).update(sabana = x, status_doc = new_status, u_sabana= u1)

    return HttpResponse()


