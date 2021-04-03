# coding: utf-8
from __builtin__ import open
from django.shortcuts import render_to_response, get_object_or_404, render
from django.template.context import RequestContext
from django.contrib.auth.decorators import permission_required, login_required
from django.contrib import messages
from models import ModeloEconomico, Economico, Operador, Archivos, Documentos, movimientosUnidad
from forms import EconomicoForm, ModeloEconomicoForm, OperadorForm, AseguradoraForm, movUnidadForm, ArchivosForm, SearchDocsForm, documentosForm, reporteForm
from django.http.response import HttpResponseRedirect, HttpResponse
from django.core.urlresolvers import reverse
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from catalogos.forms import CasetaForm, GastoViajeForm, ConceptoFacturacionForm, UploadEconomicoForm,UploadExcelForm,UploadForm, movimientosUnidad
from catalogos.models import Caseta, GastoViaje, ConceptoFacturacion, ExcelEconomico,Aseguradoras,ExcelViaje,Imagen
from cuentas.models import Perfil
from viajes.models import MovUnidad, Movimiento
from empleados.models import empleado

import openpyxl
from django.conf import settings
import os
import mimetypes
import time
import datetime
from datetime import datetime, date 

from django.conf import settings
from django.utils.encoding import smart_str
from django.core.servers.basehttp import FileWrapper
from PyPDF2 import PdfFileMerger
from easy_pdf.views import PDFTemplateView

## Ajuste para reportes 
import json
import xlwt
from xlwt import Workbook,XFStyle,Borders, Pattern, Font, easyxf
from django.views.decorators.csrf import csrf_exempt
from django.http import QueryDict

#Ajuste para pdf
from django.template.loader import get_template
from django.template import Context
from django.views.generic import TemplateView
from xhtml2pdf import pisa

@login_required
def modelos(request):
    modelos_list = ModeloEconomico.objects.all()
    paginator = Paginator(modelos_list, 10)
    page = request.GET.get('page')

    try:
        modelos = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        modelos = paginator.page(1)
    except EmptyPage:
        # If page is out of range (e.g. 9999), deliver last page of results.
        modelos = paginator.page(paginator.num_pages)
    
    context = RequestContext(request,{ 
        'modelos': modelos,                         
    })    
    
    return render_to_response('modelos.html', context)

@login_required
def modelosadd(request):    
    if request.method == 'POST':
        form = ModeloEconomicoForm(request.POST)
        if form.is_valid():
            modelo = form.save(commit = False)
            alto = modelo.alto
            largo = modelo.largo
            ancho = modelo.ancho
            modelo.capacidad_volumen = ancho * largo * alto
            modelo = form.save()
            #update 
            messages.success(request, 'Se ha agregado el modelo %s'%modelo.modelo)
            return HttpResponseRedirect(reverse('catalogos.views.modelos',))
        else:
            messages.error(request, 'Ha ocurrido un error. Por favor verifique los datos.')
    else:
        form = ModeloEconomicoForm()
    
    context = RequestContext(request,{ 
        'form': form,
        'action': 'Agregar',
    })    
    
    return render_to_response('modelos_form.html', context)

@login_required
def modelosedit(request, modelo_id):
    modelo = get_object_or_404(ModeloEconomico, pk=modelo_id)
    
    if request.method == 'POST':
        form = ModeloEconomicoForm(request.POST, instance=modelo)
        if form.is_valid():

            modelo = form.save(commit = False)
            alto = modelo.alto
            largo = modelo.largo
            ancho = modelo.ancho
            modelo.capacidad_volumen = ancho * largo * alto
            modelo = form.save()

            messages.success(request, 'Se ha editado el modelo %s'%modelo.modelo)
            return HttpResponseRedirect(reverse('catalogos.views.modelos',))
        else:
            messages.error(request, 'Ha ocurrido un error. Por favor verifique los datos.')
    else:
        form = ModeloEconomicoForm(instance=modelo)
        
    context = RequestContext(request,{ 
        'form': form,
        'action': 'Editar',
    })    
    
    return render_to_response('modelos_form.html', context)

@login_required
def modelosdelete(request, modelo_id):
    modelo = get_object_or_404(ModeloEconomico, pk=modelo_id)
    
    message = 'El modelo %s ha sido eliminado.'%modelo.modelo
    modelo.delete()
    
    messages.success(request, message)
    return HttpResponseRedirect(reverse('catalogos.views.modelos',))          

@login_required
def economicos(request):
    economicos_list = Economico.objects.all()
    paginator = Paginator(economicos_list, 100)
    page = request.GET.get('page')

    try:
        economicos = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        economicos = paginator.page(1)
    except EmptyPage:
        # If page is out of range (e.g. 9999), deliver last page of results.
        economicos = paginator.page(paginator.num_pages)        
    
    context = RequestContext(request,{ 
        'economicos': economicos,                         
    })    
    
    return render_to_response('economicos.html', context)

@login_required
def economicosadd(request):    
    if request.method == 'POST':
        form = EconomicoForm(request.POST)
        if form.is_valid():
            economico = form.save(commit = False)
            economico.status = 1
            economico = form.save ()
            messages.success(request, 'Se ha agregado el economico %s'%economico.placas)
            return HttpResponseRedirect(reverse('catalogos.views.economicos',))
        else:
            messages.error(request, 'Ha ocurrido un error. Por favor verifique los datos.')

        formImagen = UploadForm(request.POST, request.FILES)

        try: 

            newdoc = Imagen(imagen1 = request.FILES['imagen1'],imagen2 = request.FILES['imagen2'],imagen3 = request.FILES['imagen3'],imagen4 = request.FILES['imagen4'])
            newdoc.nombre = economico.pk
            newdoc.save(formImagen)
            messages.success(request, 'Las imagenes se cargaron correctamente')
            newdoc.delete()
	 
        except Exception:
            print "error foto 2"    
    else:

        form = EconomicoForm()
        formImagen = UploadForm()
    context = RequestContext(request,{ 
        'form': form,
        'formImagen':formImagen,
        'action': 'Agregar',
    })    
    
    return render_to_response('economicos_add.html', context)

@login_required
def economicosedit(request, economico_id):
    economico = get_object_or_404(Economico, pk=economico_id)

    if request.method == 'POST':
        form = EconomicoForm(request.POST, instance=economico)   
        if form.is_valid():
            form.save()
            messages.success(request, 'Se ha editado el economico %s'%economico.placas)
        else:
            messages.error(request, 'Ha ocurrido un error. Por favor verifique los datos.')
        
        formImagen = UploadForm(request.POST, request.FILES)
        
        try:

            IMG1=str(request.FILES['imagen1'])
            print 'estas es la imagen 1: ' + IMG1
            cut_IMG1 = IMG1.split('.')
            extIMG1 = cut_IMG1[1]

            if extIMG1 !='jpg':
                messages.error(request, 'Error Foto 1: el archivo (%s) no es de extension .jpg '%request.FILES['imagen1'])

            else:

              if request.FILES['imagen1'] is not None or extIMG1=='jpj':
                    newdoc1 = Imagen(imagen1 = request.FILES['imagen1'],imagen2 ='imgEconomicos/default.jpg',imagen3 ='imgEconomicos/default.jpg',imagen4 = 'imgEconomicos/default.jpg')
                    newdoc1.nombre = economico.pk
                    newdoc1.save(formImagen)
                    newdoc1.delete()
                    messages.success(request, 'La Foto 1: (%s) se agrego correctamente '%request.FILES['imagen1'])    
                    
                    lista1 = os.listdir(settings.MEDIA_ROOT + 'imgEconomicos/')
                    for lis1 in lista1:
                        PARTE1 = str(lis1)
                        num1=500
                        if PARTE1 == str(request.FILES['imagen1']):
                            llave1 = str(economico.pk)
                            NUEVAPARTE1 = PARTE1.split('.')
                            newimg1=llave1+'_'+'1.'+NUEVAPARTE1[1]
                            os.rename (settings.MEDIA_ROOT + 'imgEconomicos/%s'%PARTE1, settings.MEDIA_ROOT + 'imgEconomicos/%s'%newimg1)
                        num1 +=1
        except Exception:

            print "error foto 1"
        
        try:
            IMG2=str(request.FILES['imagen2'])
            cut_IMG2 = IMG2.split('.')
            extIMG2 = cut_IMG2[1]
            if extIMG2 !='jpg':
                messages.error(request, 'Error Foto 2: el archivo (%s) no es de extension .jpg '%request.FILES['imagen2'])
            else:
                
                if request.FILES['imagen2'] is not None or extIMG2=='jpj':
                    newdoc2 = Imagen(imagen1 ='imgEconomicos/default.jpg' ,imagen2 = request.FILES['imagen2'],imagen3 ='imgEconomicos/default.jpg',imagen4 = 'imgEconomicos/default.jpg')
                    newdoc2.nombre = economico.pk
                    newdoc2.save(formImagen)
                    newdoc2.delete()
                    messages.success(request, 'La Foto 2: (%s) se agrego correctamente '%request.FILES['imagen2'])    
                    lista2 = os.listdir(settings.MEDIA_ROOT + 'imgEconomicos/')
                    for lis2 in lista2:
                        PARTE2 = str(lis2)
                        num2=1000
                        if PARTE2 == str(request.FILES['imagen2']):
                            llave2 = str(economico.pk)
                            NUEVAPARTE2 = PARTE2.split('.')
                            newimg2=llave2+'_'+'2.'+NUEVAPARTE2[1]
                            os.rename (settings.MEDIA_ROOT + 'imgEconomicos/%s'%PARTE2, settings.MEDIA_ROOT + 'imgEconomicos/%s'%newimg2)
                        num2 +=1
        except Exception:
            print "error foto 2"    
        
        try:
            IMG3=str(request.FILES['imagen3'])
            cut_IMG3 = IMG3.split('.')
            extIMG3 = cut_IMG3[1]
            if extIMG3 !='jpg':
                messages.error(request, 'Error Foto 3: el archivo (%s) no es de extension .jpg '%request.FILES['imagen3'])
            else:    
                if request.FILES['imagen3'] is not None or extIMG3=='jpj':
                    newdoc3 = Imagen(imagen1 = 'imgEconomicos/default.jpg',imagen2 ='imgEconomicos/default.jpg',imagen3 = request.FILES['imagen3'],imagen4 = 'imgEconomicos/default.jpg')
                    newdoc3.nombre = economico.pk
                    newdoc3.save(formImagen)
                    newdoc3.delete()
                    messages.success(request, 'La Foto 3: (%s) se agrego correctamente '%request.FILES['imagen3'])
                    lista3 = os.listdir(settings.MEDIA_ROOT + 'imgEconomicos/')
                    for lis3 in lista3:
                        PARTE3 = str(lis3)
                        num3=1500
                        if PARTE3 == str(request.FILES['imagen3']):
                            llave3 = str(economico.pk)
                            NUEVAPARTE3 = PARTE3.split('.')
                            newimg3=llave3+'_'+'3.'+NUEVAPARTE3[1]
                            os.rename (settings.MEDIA_ROOT + 'imgEconomicos/%s'%PARTE3, settings.MEDIA_ROOT + 'imgEconomicos/%s'%newimg3)
                            num3 +=1
        except Exception:
            print "error foto 3"

        try:
            IMG4=str(request.FILES['imagen4'])
            cut_IMG4 = IMG4.split('.')
            extIMG4 = cut_IMG4[1]
            if extIMG4 !='jpg':
                messages.error(request, 'Error Foto 4: el archivo (%s) no es de extension .jpg '%request.FILES['imagen4'])
            else:    
                if request.FILES['imagen4'] is not None or extIMG4=='jpj':
                    newdoc4 = Imagen(imagen1 = 'imgEconomicos/default.jpg',imagen2 ='imgEconomicos/default.jpg',imagen3 ='imgEconomicos/default.jpg',imagen4 = request.FILES['imagen4'])
                    newdoc4.nombre = economico.pk
                    newdoc4.save(formImagen)
                    newdoc4.delete()
                    messages.success(request, 'La Foto 4: (%s) se agrego correctamente '%request.FILES['imagen4'])
                    lista4 = os.listdir(settings.MEDIA_ROOT + 'imgEconomicos/')
                    for lis4 in lista4:
                        PARTE4 = str(lis4)
                        num4=2000
                        if PARTE4 == str(request.FILES['imagen4']):
                            llave4 = str(economico.pk)
                            NUEVAPARTE4 = PARTE4.split('.')
                            newimg4=llave4+'_'+'4.'+NUEVAPARTE4[1]
                            os.rename (settings.MEDIA_ROOT + 'imgEconomicos/%s'%PARTE4, settings.MEDIA_ROOT + 'imgEconomicos/%s'%newimg4)
                            num4 +=1
        except Exception:
            print "error foto 4"       
                      
    else:
        form = EconomicoForm(instance=economico)
        formImagen = UploadForm()
    context = RequestContext(request,{
        'form': form,
        'economicopk':economico.pk,
        'formImagen':formImagen,
        'action': 'Editar',
    })   
    return render_to_response('economicos_form.html', context)







@login_required
def economicosdelete(request, economico_id):
    economico = get_object_or_404(Economico, pk=economico_id)
    
    message = 'El economico %s ha sido eliminado.'%economico.placas
    economico.delete()
    
    messages.success(request, message)
    return HttpResponseRedirect(reverse('catalogos.views.economicos',))  

@login_required
def operadores(request):
    operadores_list = Operador.objects.all()
    paginator = Paginator(operadores_list, 10)
    page = request.GET.get('page')

    try:
        operadores = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        operadores = paginator.page(1)
    except EmptyPage:
        # If page is out of range (e.g. 9999), deliver last page of results.
        operadores = paginator.page(paginator.num_pages)   
    
    context = RequestContext(request,{ 
        'operadores': operadores,                         
    })    
    
    return render_to_response('operadores.html', context)

@login_required
def operadoresadd(request):    
    if request.method == 'POST':
        form = OperadorForm(request.POST)
        if form.is_valid():
            operador = form.save()
            messages.success(request, 'Se ha agregado al operador: %s'%operador)
            return HttpResponseRedirect(reverse('catalogos.views.operadores',))
        else:
            messages.error(request, 'Ha ocurrido un error. Por favor verifique los datos.')
    else:
        form = OperadorForm()
    
    context = RequestContext(request,{ 
        'form': form,
        'action': 'Agregar',
    })    
    
    return render_to_response('operadores_form.html', context)

@login_required
def operadoresedit(request, operador_id):
    operador = get_object_or_404(Operador, pk=operador_id)
    
    if request.method == 'POST':
        form = OperadorForm(request.POST, instance=operador)
        if form.is_valid():
            form.save()
            messages.success(request, 'Se ha editado al operador %s'%operador)
            return HttpResponseRedirect(reverse('catalogos.views.operadores',))
        else:
            messages.error(request, 'Ha ocurrido un error. Por favor verifique los datos.')
    else:
        form = OperadorForm(instance=operador)
        
    context = RequestContext(request,{ 
        'form': form,
        'action': 'Editar',
    })    
    
    return render_to_response('operadores_form.html', context)

@login_required
def operadoresdelete(request, operador_id):
    operador = get_object_or_404(Operador, pk=operador_id)
    
    message = 'El operador %s ha sido eliminado.'%operador
    operador.delete()
    
    messages.success(request, message)
    return HttpResponseRedirect(reverse('catalogos.views.operadores',))  

@login_required
def casetas(request):
    casetas_list = Caseta.objects.all()
    paginator = Paginator(casetas_list, 10)
    page = request.GET.get('page')

    try:
        casetas = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        casetas = paginator.page(1)
    except EmptyPage:
        # If page is out of range (e.g. 9999), deliver last page of results.
        casetas = paginator.page(paginator.num_pages)   
    
    context = RequestContext(request,{ 
        'casetas': casetas,                         
    })    
    
    return render_to_response('casetas.html', context)

@login_required
def casetasadd(request):    
    if request.method == 'POST':
        form = CasetaForm(request.POST)
        if form.is_valid():
            caseta = form.save()
            messages.success(request, 'Se ha agregado la caseta: %s'%caseta)
            return HttpResponseRedirect(reverse('catalogos.views.casetas',))
        else:
            messages.error(request, 'Ha ocurrido un error. Por favor verifique los datos.')
    else:
        form = CasetaForm()
    
    context = RequestContext(request,{ 
        'form': form,
        'action': 'Agregar',
    })    
    
    return render_to_response('casetas_form.html', context)

@login_required
def casetasedit(request, caseta_id):
    caseta = get_object_or_404(Caseta, pk=caseta_id)
    
    if request.method == 'POST':
        form = CasetaForm(request.POST, instance=caseta)
        if form.is_valid():
            form.save()
            messages.success(request, 'Se ha editado la caseta: %s'%caseta)
            return HttpResponseRedirect(reverse('catalogos.views.casetas',))
        else:
            messages.error(request, 'Ha ocurrido un error. Por favor verifique los datos.')
    else:
        form = CasetaForm(instance=caseta)
        
    context = RequestContext(request,{ 
        'form': form,
        'action': 'Editar',
    })    
    
    return render_to_response('casetas_form.html', context)

@login_required
def casetasdelete(request, caseta_id):
    caseta = get_object_or_404(Caseta, pk=caseta_id)
    
    message = 'La caseta %s ha sido eliminado.'%caseta
    caseta.delete()
    
    messages.success(request, message)
    return HttpResponseRedirect(reverse('catalogos.views.casetas',))

@login_required
def gastos(request):
    gastos_list = GastoViaje.objects.all()
    paginator = Paginator(gastos_list, 10)
    page = request.GET.get('page')

    try:
        gastos = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        gastos = paginator.page(1)
    except EmptyPage:
        # If page is out of range (e.g. 9999), deliver last page of results.
        gastos = paginator.page(paginator.num_pages)   
    
    context = RequestContext(request,{ 
        'gastos': gastos,                         
    })    
    
    return render_to_response('gastos.html', context)

@login_required
def gastosadd(request):    
    if request.method == 'POST':
        form = GastoViajeForm(request.POST)
        if form.is_valid():
            gasto = form.save()
            messages.success(request, 'Se ha agregado el tipo de gasto: %s'%gasto)
            return HttpResponseRedirect(reverse('catalogos.views.gastos',))
        else:
            messages.error(request, 'Ha ocurrido un error. Por favor verifique los datos.')
    else:
        form = GastoViajeForm()
    
    context = RequestContext(request,{ 
        'form': form,
        'action': 'Agregar',
    })    
    
    return render_to_response('gastos_form.html', context)

@login_required
def gastosedit(request, gasto_id):
    gasto = get_object_or_404(GastoViaje, pk=gasto_id)
    
    if request.method == 'POST':
        form = CasetaForm(request.POST, instance=gasto)
        if form.is_valid():
            form.save()
            messages.success(request, 'Se ha editado el tipo de gasto: %s'%gasto)
            return HttpResponseRedirect(reverse('catalogos.views.gastos',))
        else:
            messages.error(request, 'Ha ocurrido un error. Por favor verifique los datos.')
    else:
        form = GastoViajeForm(instance=gasto)
        
    context = RequestContext(request,{ 
        'form': form,
        'action': 'Editar',
    })    
    
    return render_to_response('gastos_form.html', context)

@login_required
def gastosdelete(request, gasto_id):
    gasto = get_object_or_404(GastoViaje, pk=gasto_id)
    
    message = 'El tipo de gasto %s ha sido eliminado.'%gasto
    gasto.delete()
    
    messages.success(request, message)
    return HttpResponseRedirect(reverse('catalogos.views.gastos',))  

@login_required
def conceptos(request):
    conceptos_list = ConceptoFacturacion.objects.all()
    paginator = Paginator(conceptos_list, 10)
    page = request.GET.get('page')

    try:
        conceptos = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        conceptos = paginator.page(1)
    except EmptyPage:
        # If page is out of range (e.g. 9999), deliver last page of results.
        conceptos = paginator.page(paginator.num_pages)   
    
    context = RequestContext(request,{ 
        'conceptos': conceptos,                         
    })    
    
    return render_to_response('conceptos.html', context)

@login_required
def conceptosadd(request):    
    if request.method == 'POST':
        form = ConceptoFacturacionForm(request.POST)
        if form.is_valid():
            concepto = form.save()
            messages.success(request, 'Se ha agregado el concepto: %s'%concepto)
            return HttpResponseRedirect(reverse('catalogos.views.conceptos',))
        else:
            messages.error(request, 'Ha ocurrido un error. Por favor verifique los datos.')
    else:
        form = ConceptoFacturacionForm()
    
    context = RequestContext(request,{ 
        'form': form,
        'action': 'Agregar',
    })    
    
    return render_to_response('conceptos_form.html', context)

@login_required
def conceptosedit(request, concepto_id):
    concepto = get_object_or_404(ConceptoFacturacion, pk=concepto_id)
    
    if request.method == 'POST':
        form = ConceptoFacturacionForm(request.POST, instance=concepto)
        if form.is_valid():
            form.save()
            messages.success(request, 'Se ha editado el concepto: %s'%concepto)
            return HttpResponseRedirect(reverse('catalogos.views.conceptos',))
        else:
            messages.error(request, 'Ha ocurrido un error. Por favor verifique los datos.')
    else:
        form = CasetaForm(instance=concepto)
        
    context = RequestContext(request,{ 
        'form': form,
        'action': 'Editar',
    })    
    
    return render_to_response('conceptos_form.html', context)

@login_required
def conceptosdelete(request, concepto_id):
    concepto = get_object_or_404(ConceptoFacturacion, pk=concepto_id)
    
    message = 'El concepto %s ha sido eliminado.'%concepto
    concepto.delete()
    
    messages.success(request, message)
    return HttpResponseRedirect(reverse('catalogos.views.conceptos',))

def ExcelEconomicoView(request):

    if request.method == 'POST':
        form = UploadEconomicoForm(request.POST, request.FILES)
        print request.POST
        print request.FILES
        print request.FILES['docfile']
        #form.save()
        if form.is_valid():
            newdoc = ExcelEconomico(docfile = request.FILES['docfile'])
            nombreArchivo = str(newdoc.docfile)
            nombreArchivo = nombreArchivo.upper()
            if nombreArchivo.endswith('.XLSX'):
                newdoc.save()
                insertaEconomicos(request, str(newdoc.docfile))
                ExcelEconomico.objects.all().delete()
            else:
                txtmsg = 'Archivo con formato inválido.'
                messages.error(request, txtmsg)

        return render(request,'subirEconomicos.html', {'form': form,})
    else:
        form = UploadEconomicoForm()
    return render(request,'subirEconomicos.html', {'form': form,})


def insertaEconomicos(request, datoArchivo):
    user = request.user
    workbook = openpyxl.load_workbook(filename=settings.MEDIA_ROOT + datoArchivo, use_iterators=True)
    try:
        sheet = workbook.get_sheet_by_name('pvehicular')
        exitos = 0
        numError = 0
        total = 0
        for row in range(6, sheet.get_highest_row()):
            total += 1
            print sheet['B' + str(row)].value
            #if sheet['B' + str(row)].value == 'FIN': break
            try:
                economico = Economico()
                economico.pk = sheet['B' + str(row)].value
                economico.color = 'Rojo'
                mdo = ModeloEconomico.objects.get(modelo = sheet['C' + str(row)].value)
                print mdo
                economico.modelo = mdo
                economico.placas = sheet['K' + str(row)].value
                economico.activo = 1
                economico.pasa_como = 1
                economico.marca = sheet['D' + str(row)].value
                economico.submarca = sheet['E' + str(row)].value
                economico.placas = sheet['K' + str(row)].value
                economico.tipoPlaca = sheet['L' + str(row)].value
                economico.caja = sheet['M' + str(row)].value
                economico.medidallantas = sheet['O' + str(row)].value
                economico.tipollantas = sheet['N' + str(row)].value
                economico.carga = sheet['R' + str(row)].value
                economico.aseguradora = Aseguradoras.objects.get(nombre = sheet['S' + str(row)].value)
                economico.poliza = sheet['T' + str(row)].value
                economico.fecha_vencimiento = sheet['U' + str(row)].value
                economico.iave = sheet['V' + str(row)].value
                economico.cerradura = 1
                economico.operador = Operador.objects.get(nombre = sheet['Y' + str(row)].value)
                economico.cctv = 1
                economico.kilometrajeServicio = sheet['AA' + str(row)].value
                economico.ultimoKilometraje = sheet['AB' + str(row)].value
                economico.save()
            except KeyError:
                print 'Error'
                messages.error(request, u'Error al registrar económico ' + str(sheet['B' + str(row)].value) + u' Verifique los datos')
                numError += 1

        archivoF = settings.MEDIA_ROOT + 'upload/' + datoArchivo
        os.remove(archivoF)
        if exitos == 0:
            messages.error(request, 'No se realizaron actualizaciones')
        if numError == 0:
            messages.success(request, 'Registros actualizados exitosamente.')
        if exitos > 0 and numError > 0:
            messages.success(request, 'Actualización parcial realizada.')

    except KeyError:
        messages.error(request, 'Archivo no contiene la hoja fac_cm')


#############################################################################
def subir222222IMG(request): 
    if request.method == 'POST':
        form = UploadExcelForm(request.POST, request.FILES)
        #form.save()
        if form.is_valid():
            newdoc = ExcelViaje(docfile = request.FILES['docfile'])
            nombreArchivo = str(newdoc.docfile)
            nombreArchivo = nombreArchivo.upper()
            if nombreArchivo.endswith('.XLSX'):
                print "ok"
                #newdoc.save()
                #actualizaFacturas(request, str(newdoc.docfile))
                #ExcelViaje.objects.all().delete()
            else: 
                print 'Archivo con formato invalido'
                txtmsg = 'Archivo con formato inválido.'
                messages.error(request, txtmsg)  
              
        return render(request,'economicos_Fotos.html', {'form': form,})
    else:
        form = UploadExcelForm()
    return render(request,'economicos_Fotos.html', {'form': form,})
@login_required
def economicosImg333333333(request, economico_id):
    economico = get_object_or_404(Economico, pk=economico_id)
    print "ID=",economico.pk
    if request.method == 'POST':
        form = UploadExcelForm(request.POST, request.FILES)
        #form.save()
        if form.is_valid():
            #newdoc = ExcelViaje(docfile = request.FILES['docfile'])
            #print newdoc
            #nombreArchivo = str(newdoc.docfile)
            #nombreArchivo = nombreArchivo.upper()
            #if nombreArchivo.endswith('.XLSX'):
            #    print "ok"
                #newdoc.save()
                #actualizaFacturas(request, str(newdoc.docfile))
                #ExcelViaje.objects.all().delete()
            #else: 
            #    print 'Archivo con formato invalido'
            #    txtmsg = 'Archivo con formato inválido.'
            #    messages.error(request, txtmsg)  
        
            print 'Archivo VALIDO'

        return render(request,'economicos_Fotos.html', {'form': form,})
    else:
        form = UploadExcelForm()
    return render(request,'economicos_Fotos.html', {'form': form,})

def subirArchivoPdf(f):
    #destino = open('/var/www/PDFS/%s'%f.name, 'wb+')
    destino = open(settings.MEDIA_ROOT + 'pdfs/%s'%f.name, 'wb+')
    for chunk in f.chunks():
        destino.write(chunk)
    destino.close()
#######################################3
def economicosImg(request, economico_id):
    economico = get_object_or_404(Economico, pk=economico_id)
    print "ID=",economico.pk
    
    if request.method == 'POST':
        form = UploadForm(request.POST, request.FILES)
        if form.is_valid():
            newdoc = Imagen(nombreImagen = request.POST['nombreImagen'],imagen = request.FILES['imagen'])
            newdoc.save(form)
            messages.success(request, 'La imagen: %s se subio correctamente'%newdoc.nombreImagen)
            return render(request, 'economicos_Fotos.html', {'form': form})
    else:
        form = UploadForm()
    return render(request, 'economicos_Fotos.html', {'form': form})

@login_required
def asegura(request):
    aseguradoras_list = Aseguradoras.objects.all()
    paginator = Paginator(aseguradoras_list, 10)
    page = request.GET.get('page')

    try:
        aseguradoras = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        aseguradoras = paginator.page(1)
    except EmptyPage:
        # If page is out of range (e.g. 9999), deliver last page of results.
        aseguradoras = paginator.page(paginator.num_pages)

    context = RequestContext(request,{
        'aseguradoras': aseguradoras,
    })

    return render_to_response('aseguradoras.html', context)

@login_required
def aseguradoraadd(request):
    if request.method == 'POST':
        form = AseguradoraForm(request.POST)
        if form.is_valid():
            asegura = form.save()
            messages.success(request, 'Se ha agregado la aseguradora %s' % asegura.nombre)
            return HttpResponseRedirect(reverse('catalogos.views.asegura',))
        else:
            messages.error(request, 'Ha ocurrido un error. Por favor verifique los datos.')
    else:
        form = AseguradoraForm()

    context = RequestContext(request,{
        'form': form,
        'action': 'Agregar',
    })

    return render_to_response('aseguradora_form.html', context)

@login_required
def aseguradorasedit(request, aseguradora_id):
    aseguradoraDato = get_object_or_404(Aseguradoras, pk=aseguradora_id)

    if request.method == 'POST':
        form = AseguradoraForm(request.POST, instance=aseguradoraDato)
        if form.is_valid():
            form.save()
            messages.success(request, 'Se ha editado la aseguradora %s' % aseguradoraDato.nombre)
            return HttpResponseRedirect(reverse('catalogos.views.asegura',))
        else:
            messages.error(request, 'Ha ocurrido un error. Por favor verifique los datos.')
    else:
        form = AseguradoraForm(instance=aseguradoraDato)

    context = RequestContext(request,{
        'form': form,
        'action': 'Editar',
    })

    return render_to_response('aseguradora_form.html', context)

@login_required
def aseguradoradelete(request, aseguradora_id):
    aseguradoraDato = get_object_or_404(Aseguradoras, pk=aseguradora_id)

    message = 'La aseguradora %s ha sido eliminada.'%  aseguradoraDato.nombre
    aseguradoraDato.delete()

    messages.success(request, message)
    return HttpResponseRedirect(reverse('catalogos.views.asegura',))


def movUnidad (request):
    userid = request.user.id 
    user = request.user
    u1= Perfil.objects.get(user_id=userid)
    
    if request.method == 'POST':

       form = movUnidadForm(request.POST)

       if form.is_valid():

          uni = request.POST['unidad']
          newstatus = request.POST ['tipo']
          unidadBase = Economico.objects.get(id = uni)
          unidadBase.status = newstatus
          unidadBase.save()
          form.save(commit=False)

          o=request.POST['operador']
          t=request.POST['tipo']
          ob=request.POST['obs']
          
          fechai=request.POST['fecha']
          fi = datetime.strptime(fechai,"%d-%m-%Y").strftime("%Y-%m-%d")
          print fi
          fechar=request.POST['tiemporep']
          fr = datetime.strptime(fechar,"%d-%m-%Y").strftime("%Y-%m-%d")
          print fr
          ob2=request.POST['alternativa']
          obs= ob + ' , ' + ob2
          #f=request.POST['fecha']
          ts1= time.time()
          st = datetime.fromtimestamp(ts1).strftime('%Y-%m-%d')
          model=Economico.objects.filter(id=uni)
          for modelos in model:
            m=modelos.modelo

          MovUnidad.objects.filter(unidad=uni).update(
               usuario = u1.id, operador = o, tipo = t, modelo=m, obs = obs, ts=st, cliente ='', departamento='', destino2='', destino = '',  
               fechai= fi, fechaf = fr, fechamcia = None )
          
          form.save()

          messages.success(request,'Se ha actualizado el Estado de la unidad')
          return HttpResponseRedirect (reverse('catalogos.views.economicos')) 
       else:
          messages.error(request,'No se ha podido actualizar el Estado, por favor revisa la informacion')

    else: 

       form = movUnidadForm()
    context = RequestContext (request, {
            'form': form,
            'action': 'Registrar',
})
    return render_to_response ('movunidad.html', context)


@login_required
def documentosEconomico(request, economico_id):
    archivos = []
    if request.method == 'POST':

        formDoc = ArchivosForm(request.POST, request.FILES)

        if formDoc.is_valid():

            try:
                tipoDoc = Documentos.objects.get(nombreDoc = formDoc.cleaned_data['idDoc'])
                print tipoDoc.nombreDoc
                archivo = Archivos.objects.get(idEconomico = formDoc.cleaned_data['idEconomico'], idDoc = tipoDoc.pk)
                print 'eliminando archivo anterior' + str(archivo.nombreDoc)
                ruta = settings.MEDIA_ROOT + str(archivo.nombreDoc)
                os.remove(ruta)
                formDoc = ArchivosForm(request.POST, request.FILES, instance=archivo)
                formDoc.save()
                messages.success(request, u'Archivos actualizados exitosamente')
            except Exception:
                formDoc.save()
                messages.success(request, u'Archivos cargados exitosamente')


    else:
        formDoc = ArchivosForm(initial={'idEconomico':economico_id})
        print 'NO Es valido'



    archivos = Archivos.objects.filter(idEconomico =Economico.objects.get(pk=economico_id))

    context = RequestContext(request,{
        'formDoc':formDoc,
        'action': 'Editar',
        'archivos': archivos,
        'idEconomico' : economico_id
    })
    return render_to_response('economicosDoc.html', context)

@login_required
def descargarDoc(request, nombreArchivo):
    ruta = settings.MEDIA_ROOT + str(nombreArchivo)
    print ruta
    wrapper = FileWrapper( open( ruta, "r" ) )
    content_type = mimetypes.guess_type( ruta )[0]

    response = HttpResponse(wrapper, content_type = content_type)
    response['Content-Length'] = os.path.getsize( ruta )
    response['Content-Disposition'] = 'attachment; filename=%s' % \
                                       smart_str( os.path.basename( ruta ) )

    return response

@login_required
def borrarDoc(request, nombreArchivo, economico_id):

    formDoc = ArchivosForm(initial={'idEconomico':economico_id})
    ruta = settings.MEDIA_ROOT + str(nombreArchivo)
    archivo = get_object_or_404(Archivos, nombreDoc=nombreArchivo, idEconomico =Economico.objects.get(pk=economico_id))
    archivo.delete()
    os.remove(ruta)
    messages.success(request, u'Se elimino el archivo ' + nombreArchivo)
    archivos = Archivos.objects.filter(idEconomico =Economico.objects.get(pk=economico_id))


    context = RequestContext(request,{
        'formDoc':formDoc,
        'action': 'Editar',
        'archivos': archivos,
        'idEconomico' : economico_id
    })
    return render_to_response('economicosDoc.html', context)


####Documentos

def index_docs (request):
    searchform = SearchDocsForm (request.GET)
    CatDoc_list = Documentos.objects.all()
    CatDoc_list= CatDoc_list.order_by('-pk')
    paginator = Paginator(CatDoc_list, 80)
    page = request.GET.get('page')
    try:
        unidoc = paginator.page(page)
    except PageNotAnInteger:
        unidoc = paginator.page(1)
    except EmptyPage:
        unidoc = paginator.page(paginator.num_pages)
    context = RequestContext (request, {
        'searchform':searchform,
        'unidoc':unidoc,
    })
    return render_to_response ('catdocumentos.html', context) ### copiar de empdocumentos.html

def documentos_add (request):
     user = request.user
     profile = user.get_profile()   
     if request.method == 'POST':         
         form = documentosForm(request.POST)
         if form.is_valid ():
             documento = form.save()
             messages.success(request, 'Documento registrado de forma exitosa')
             return HttpResponseRedirect (reverse('catalogos.views.index_docs'))
         else:
             messages.error(request, 'El documento ya existe')
     else: 
         form = documentosForm()
     context = RequestContext (request, {
         'form': form,
         'action':'Agregar',
      })
     return render_to_response ('catdocumentos_form.html',context)###copiar de empdocumentos_form.html

def docu_edit (request, catdoc_id):
     user = request.user
     profile = user.get_profile()
     documento = get_object_or_404(Documentos, pk=catdoc_id)
     if request.method == 'POST':
        form = documentosForm(request.POST, instance = documento)
        if form.is_valid():
            doc = form.save()
            messages.success(request, 'Se ha editado el documento')
            return HttpResponseRedirect (reverse('catalogos.views.index_docs'))
        else: 
            messages.error(request,'No se pudo editar el documento')
     else: 
         form = documentosForm(instance=documento)
     context = RequestContext (request, {
           'form': form,
           'action':'Editar',
      })
     return render_to_response ('catdocumentos_form.html', context)  ### copiar de empdocumentos_form.html

@login_required
def docu_del (request, catdoc_id):
    documento = get_object_or_404(Documentos, pk = catdoc_id)
    message = 'El documento %s ha sido borrado.' %documento.nombreDoc
    documento.delete()
    messages.success(request, message)
    return HttpResponseRedirect(reverse('catalogos.views.index_docs',))
 
@login_required
def generaPdfEconomico (request, economico_id):

    templateString = ("economicoResPdf.html")
    template = get_template(templateString)

    economico = get_object_or_404(Economico, pk=economico_id)

    form = economico

    myContextObject = {
        'form': form,
        'economicopk':economico.pk,
    }

    html = template.render(Context(myContextObject))
    file = open(settings.MEDIA_ROOT + "documentosEconomico/EconomicoNo" + economico_id + ".pdf", "w+b")
    links    = lambda uri, rel: os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ''))
    pisaStatus = pisa.CreatePDF(html.encode('utf-8'), dest=file, encoding='utf-8', link_callback=links)
    file.seek(0)
    pdf = file.read()
    file.close()

    archivos = Archivos.objects.all().filter(idEconomico=economico_id)
    merger = PdfFileMerger()
    merger.append(fileobj=open(settings.MEDIA_ROOT + "documentosEconomico/EconomicoNo" + economico_id + ".pdf", "rb"))
    for arch in archivos:
        merger.append(fileobj=open(settings.MEDIA_ROOT + str(arch.nombreDoc), "rb"))

    output = open(settings.MEDIA_ROOT + "documentosEconomico/EconomicoNo" + economico_id + "_Documentos.pdf", "wb")
    merger.write(output)
    merger.close()

    pdfSal =  open(settings.MEDIA_ROOT + "documentosEconomico/EconomicoNo" + economico_id + "_Documentos.pdf", "r")

    response = HttpResponse(pdfSal, mimetype="application/pdf")
    response["Content-Disposition"] = "attachment; filename=" + "EconomicoNo" + economico_id + "_Documentos.pdf"
    return response

    ##### Cambios de los reportes

@login_required
def reportUnidades(request):

    response = HttpResponse(mimetype="application/vnd.ms-excel")
    response['Content-Disposition'] = 'attachment; filename=Unidades.xls'


    wb = xlwt.Workbook(encoding='utf8')
    ws = wb.add_sheet('Unidades')
    date_format = xlwt.XFStyle()
    date_format.num_format_str = 'dd/mm/yyyy'
    formatoNum = XFStyle()
    formatoNum.num_format_str = '$#,##0.00'
    font = xlwt.Font() # Crear Font
    font.name = 'Arial'
    font.height = 20 * 12  #22pt
    font.bold = True
    font.italic = True
    style = xlwt.XFStyle() # Crar Style
    style = xlwt.easyxf('pattern: pattern solid, fore_colour light_blue;') #background color
    style.font = font # Aplicar Font al Style

    columnas = [
            (u"Unidad", 70),
            (u"Color", 70),
            (r'Modelo', 70),
            (r'Placas',70),
            (r'Activo', 70),
            (u"Pasa Como", 70),
            (u"Marca", 70),
            (u"SubMarca", 70),
            (u"Tipo Placa", 70),
            (u"Caja", 70),
            (u"Medida LLantas", 70),
            (u"Tipo Llantas", 70),
            (u"Carga", 70),
            (u"Aseguradora", 70),
            (u"Poliza", 70),
            (u"Fecha Vencimiento", 70),
            (u"IAVE", 70),
            (u"Cerradura", 70),
            (u"Operador", 70),
            (u"CCTV", 70),
            (u"Ultimo Kilometraje", 70),
            (u"Kilometraje Servicio", 70),
            (u"Combustible", 70),
            (u"Cilindros", 70),
            (u"PSI", 70),
            (u"Filtro", 70),
            (u"Tipo Aceite", 70), 
            (u"Litros", 70),
            (u"Rendimiento", 70),
            (u"Antijamer", 70),
            (u"Fecha Antijame", 70),
            (u"Boton Panico", 70),
            (u"Camara Int", 70),
            (u"Camara Ext", 70),
            (u"Lugar Camara Int", 70),
            (u"Lugar Camara Ext", 70),
            (u"Clave", 70),
            (u"Numero de Serie", 70),
            (u"Año", 70),
            (u"Status", 70),
            (u"Filtro de Aire", 70),
            (u"Filtro de Aceite", 70),
            (u"Filtro de Gas", 70),
            (u"Filtro Aire Acondicionado", 70),
    ]

    posx = 0
    for col in xrange(len(columnas)):
        ws.write(0, posx, columnas[col][0])
        posx += 1

    unidades=Economico.objects.all()

    renglon=1
    if unidades:
        for u in unidades:
            operador = str(u.operador)
            modelo= str(u.modelo)
            aseguradora = str(u.aseguradora)
            fechaven=str(u.fecha_vencimiento)
            print u.status
            st=u.status
            print st

            ws.write(renglon, 0, u.id)
            ws.col(0).width=3000
            ws.write(renglon, 1, u.color)
            ws.col(0).width=3000
            ws.write(renglon, 2, modelo)
            ws.col(0).width=3000
            ws.write(renglon, 3, u.placas)
            ws.col(0).width=3000
            ws.write(renglon, 4, u.activo)
            ws.col(0).width=3000
            ws.write(renglon, 5, u.pasa_como)
            ws.col(0).width=3000
            ws.write(renglon, 6, u.marca)
            ws.col(0).width=3000
            ws.write(renglon, 7, u.submarca)
            ws.col(0).width=3000
            ws.write(renglon, 8, u.tipoPlaca)
            ws.col(0).width=3000
            ws.write(renglon, 9, u.caja)
            ws.col(0).width=3000
            ws.write(renglon, 10, u.medidallantas)
            ws.col(0).width=3000
            ws.write(renglon, 11, u.tipollantas)
            ws.col(0).width=3000
            ws.write(renglon, 12, u.carga)
            ws.col(0).width=3000
            ws.write(renglon, 13, aseguradora)
            ws.col(0).width=3000
            ws.write(renglon, 14, u.poliza)
            ws.col(0).width=3000
            ws.write(renglon, 15, fechaven)
            ws.col(0).width=3000
            ws.write(renglon, 16, u.iave)
            ws.col(0).width=3000
            ws.write(renglon, 17, u.cerradura)
            ws.col(0).width=3000
            ws.write(renglon, 18, operador)
            ws.col(0).width=3000
            ws.write(renglon, 19, u.cctv)
            ws.col(0).width=3000
            ws.write(renglon, 20, u.ultimoKilometraje)
            ws.col(0).width=3000
            ws.write(renglon, 21, u.kilometrajeServicio)
            ws.col(0).width=3000
            ws.write(renglon, 22, u.combustible)
            ws.col(0).width=3000
            ws.write(renglon, 23, u.cilindros)
            ws.col(0).width=3000
            ws.write(renglon, 24, u.psi)
            ws.col(0).width=3000
            ws.write(renglon, 25, u.filtro)
            ws.col(0).width=3000
            ws.write(renglon, 26, u.tipo_aceite)
            ws.col(0).width=3000
            ws.write(renglon, 27, u.litros)
            ws.col(0).width=3000
            ws.write(renglon, 28, u.rendimiento)
            ws.col(0).width=3000
            ws.write(renglon, 29, u.antijamer)
            ws.col(0).width=3000
            ws.write(renglon, 30, u.fecha_antijamer)
            ws.col(0).width=3000
            ws.write(renglon, 31, u.boton_panico)
            ws.col(0).width=3000
            ws.write(renglon, 32, u.camara_int)
            ws.col(0).width=3000
            ws.write(renglon, 33, u.camara_ext)
            ws.col(0).width=3000
            ws.write(renglon, 34, u.lugar_camara_int)
            ws.col(0).width=3000
            ws.write(renglon, 35, u.lugar_camara_ext)
            ws.col(0).width=3000
            ws.write(renglon, 36, u.clave)
            ws.col(0).width=3000
            ws.write(renglon, 37, u.ns)
            ws.col(0).width=3000
            ws.write(renglon, 38, u.anio)
            ws.col(0).width=3000
            ws.write(renglon, 39, st)
            ws.col(0).width=3000
            ws.write(renglon, 40, u.filtro_aire)
            ws.col(0).width=3000
            ws.write(renglon, 41, u.filtro_aceite)
            ws.col(0).width=3000
            ws.write(renglon, 42, u.filtro_gas)
            ws.col(0).width=3000
            ws.write(renglon, 43, u.filtro_airea)
            ws.col(0).width=3000
            renglon += 1

    wb.save(response)
    return response


@login_required
def unidadesXLS(request):
    form=reporteForm()
    return render(request,'reporteUnidades.html',{'form':form})

@login_required
def repMovUni(request):
    inicial= None
    final= None
    filters={}
    filterMov={}

    if ('fecha_ini') in request.POST and request.POST['fecha_ini'].strip():
        inicial = request.POST['fecha_ini']

    if ('fecha_fin') in request.POST and request.POST['fecha_fin'].strip():
        final = request.POST['fecha_fin']

    #Validacion de los campos
    if inicial and final:
        filters['fecha__range'] = (inicial, final)
        filterMov['fecha__range'] = (inicial, final)
    elif inicial or final:
        messages.error(request, 'Seleccione completo rango de fechas.')
        form = reporteForm(request.POST)
        return render(request,'reporteUnidades.html', {'form': form,})

    if not len(filters):
        messages.error(request, 'Seleccione por lo menos un filtro de búsqueda o un rango de fechas')
        form = reporteForm(request.POST)
        return render(request,'reporteUnidades.html', {'form': form,})
        

    response = HttpResponse(mimetype="application/vnd.ms-excel")
    response['Content-Disposition'] = 'attachment; filename=movuni.xls'
    wb = xlwt.Workbook(encoding='utf8')
    ws = wb.add_sheet('Moviminetos', cell_overwrite_ok=True)
    date_format = xlwt.XFStyle()
    date_format.num_format_str = 'dd/mm/yyyy'
    formatoNum = XFStyle()
    formatoNum.num_format_str = '$#,##0.00'
    font = xlwt.Font() # Crear Font
    font.name = 'Arial'
    font.height = 20 * 12  #22pt
    font.bold = True
    font.italic = True
    style = xlwt.XFStyle() # Crar Style
    style = xlwt.easyxf('pattern: pattern solid, fore_colour light_blue;') #background color
    style.font = font # Aplicar Font al Style

    columnas = [
            (u"Movimiento", 70),
            (u"Unidad", 70),
            (u"Fecha", 70),
            (u"tipo", 70),
            (u"Observaciones", 70),
            (u"Operador", 70),
            (u"Kilometraje Entrada", 70),
            (u"Kilometraje Salida", 70), 
            (u"Fecha de Reparacion", 70),
            (u"Alternativa", 70), 

             ]
    posx = 0
    for col in xrange(len(columnas)):
        ws.write(0, posx, columnas[col][0])
        posx += 1
    
    movimientos = movimientosUnidad.objects.filter(**filterMov)#.values_list('ref_viaje',flat=True).distinct() Al parecer hasta aqui va bien.

    renglon=1
    if movimientos:
        for mov in movimientos:

            fechamov=str(mov.fecha)
            unidad=str(mov.unidad)
            print mov.operador 
            o=str(mov.operador)
            print mov.tipo
            reparacion = str(mov.tiemporep)
            m=mov.tipo
            movi=Movimiento.objects.filter(id=m)
            for tipo in movi:
                t=tipo.nombre
                print t

                ws.write(renglon, 0, mov.id)
                ws.col(0).width=3000
                ws.write(renglon, 1, unidad)
                ws.col(0).width=3000
                ws.write(renglon, 2, fechamov)
                ws.col(0).width=3000
                ws.write(renglon, 3, t)
                ws.col(0).width=3000
                ws.write(renglon, 4, mov.obs)
                ws.col(0).width=3000
                ws.write(renglon, 5, o)
                ws.col(0).width=3000
                ws.write(renglon, 6, mov.kmen)
                ws.col(0).width=3000
                ws.write(renglon, 7, mov.kmsa)
                ws.col(0).width=3000
                ws.write(renglon, 5, reparacion)
                ws.col(0).width=3000
                ws.write(renglon, 5, mov.alternativa)
                ws.col(0).width=3000            
                renglon += 1

    


    wb.save(response)
    return response


@login_required
def verMovimientos(request):

    movunidad_list = movimientosUnidad.objects.all().order_by('-pk')             
    paginator = Paginator(movunidad_list, 200)
    page = request.GET.get('page')
    
    try:
        movunidad = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        movunidad = paginator.page(1)
    except EmptyPage:
        # If page is out of range (e.g. 9999), deliver last page of results.
        movunidad = paginator.page(paginator.num_pages)
    
        
    context = RequestContext(request,{             
        'movunidad': movunidad,
    })    
    
    return render_to_response('movimientosCat.html', context)



